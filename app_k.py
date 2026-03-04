# ##########################  app_k.py  ###################################################################################
# #########################################################################################################################

# ======================= START: Import Libraries ==============================================================================

# Streamlit ML Studio — sticky top nav, parallax hero background, NO fixed top tables
# pip install streamlit scikit-learn xgboost shap pandas numpy plotly
from sklearn.metrics import classification_report, confusion_matrix, f1_score, precision_score, recall_score
import streamlit as st  # Import Streamlit for building the web app UI
# import warnings
# warnings.filterwarnings(
#     "ignore",
#     message=".*use_container_width.*",
#     category=DeprecationWarning
# )
#

# Store the original button exactly once (in module attr, not in session)
if not hasattr(st, "_original_button"):
    st._original_button = st.button

def _auto_button(label, *args, **kwargs):
    """
    Your wrapped version of st.button. Add any custom behavior here.
    This wrapper will never fail if session_state is empty.
    """
    return st._original_button(label, *args, **kwargs)

# Apply monkey-patch once
st.button = _auto_button

import pandas as pd  # Import Pandas for data handling
import numpy as np  # Import NumPy for numerical operations
import matplotlib.pyplot as plt  # Import Matplotlib for plotting
import plotly.express as px  # Import Plotly Express for interactive visualizations
import plotly.graph_objects as go  # Import Plotly Graph Objects for more customized plots
from sklearn.decomposition import PCA  # Import PCA for dimensionality reduction
from sklearn.cross_decomposition import PLSRegression  # Import PLSRegression for regression analysis with collinear data
from sklearn.model_selection import train_test_split  # Import train_test_split for splitting datasets
from sklearn.svm import SVC  # Import Support Vector Classifier for classification
from sklearn.linear_model import LogisticRegression  # Import Logistic Regression for classification
from sklearn.neighbors import KNeighborsClassifier  # Import KNN for classification
from sklearn.ensemble import RandomForestClassifier  # Import Random Forest for classification
from xgboost import XGBClassifier  # Import XGBoost for classification
from sklearn.metrics import classification_report, confusion_matrix  # Import metrics for model evaluation
import seaborn as sns  # Import Seaborn for additional plotting capabilities
from sklearn.cross_decomposition import PLSRegression
from sklearn.preprocessing import StandardScaler, LabelEncoder
import plotly.express as px
import streamlit as st
import pandas as pd
import numpy as np
import base64
import io
import os
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.preprocessing import StandardScaler, OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score,
    f1_score, roc_auc_score, confusion_matrix, classification_report
)
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.inspection import permutation_importance

# Optional (safe guards if not installed)
try:
    from xgboost import XGBClassifier
    HAS_XGB = True
except Exception:
    HAS_XGB = False

try:
    import shap
    HAS_SHAP = True
except Exception:
    HAS_SHAP = False

import plotly.express as px
import plotly.graph_objects as go

# =======================END: Import Libraries ==============================================================================

# -------------------- Page config --------------------•
st.set_page_config(
    page_title="AI-Powered Vibrational Spectroscopy for Bone Health Analysis ",
    page_icon="🧬",
    layout="wide"
)

theme = "Light"


# --- Method info + selector (sidebar) ---

# --- Keep method panel state once (can be anywhere before the buttons) ---
if "method_panel_active" not in st.session_state:
    st.session_state.method_panel_active = False
if "method_choice_sidebar" not in st.session_state:
    st.session_state.method_choice_sidebar = "SVM"

def _open_method_panel(method: str):
    st.session_state.method_choice_sidebar = method
    st.session_state.method_panel_active = True

# ======================START: Image paths (local file paths or URLs)==============================================================================

# ✅ Set these to your desired images. Local files will be embedded as base64; URLs used directly.
HERO_IMAGE_PATH = "assets/hero.jpg"   # e.g., "assets/hero.jpg" or a full URL
PAPER_IMAGE_PATH = "assets/paper.png" # e.g., "assets/paper.png" or a full URL
# -------------------- Image paths (local file paths or URLs) --------------------
# Use your local DNA image (downloaded and saved as Background_1.jpg in your project folder)
HERO_DEFAULT = "Background_1.jpg"     # ← local image file
PAPER_IMAGE_PATH = "assets/paper.png" # optional page wallpaper

# Fallback URLs if local files are missing
# HERO_FALLBACK_URL = "https://images.unsplash.com/photo-1500530855697-b586d89ba3ee?q=80&w=1400&auto=format&fit=crop"
# PAPER_FALLBACK_URL = "https://images.unsplash.com/photo-1524820824462-64a7f3c6f42b?q=80&w=1200&auto=format&fit=crop"

HERO_FALLBACK_URL = "https://www.freepik.com/premium-photo/glowing-dna-strands-technological-background_246516446.htm#from_element=cross_selling__photo?q=80&w=1400&auto=format&fit=crop"
PAPER_FALLBACK_URL = "https://www.freepik.com/premium-psd/human-dna-molecule-thread_417146727.htm#from_element=cross_selling__psd"



# ======================END: Image paths (local file paths or URLs)==============================================================================



def path_to_css_url(path: str, fallback_url: str) -> str:
    """Return a URL usable in CSS. If `path` is http(s), return it.
    Otherwise try to read a local file and embed as data URI. On failure, return fallback_url."""
    import os, base64
    if isinstance(path, str) and path.lower().startswith(("http://", "https://")):
        return path
    try:
        with open(path, "rb") as f:
            data = f.read()
        ext = os.path.splitext(path)[1].lower()
        if ext in (".jpg", ".jpeg"):
            mime = "jpeg"
        elif ext == ".png":
            mime = "png"
        elif ext == ".webp":
            mime = "webp"
        else:
            mime = "octet-stream"
        b64 = base64.b64encode(data).decode("utf-8")
        return f"data:image/{mime};base64,{b64}"
    except Exception:
        return fallback_url

# Resolve hero + paper URLs (local or fallback)
hero_url = path_to_css_url(HERO_DEFAULT, HERO_FALLBACK_URL)
paper_url = path_to_css_url(PAPER_IMAGE_PATH, PAPER_FALLBACK_URL)

# -------------------- Accent map --------------------
import streamlit as st
import pandas as pd

def coerce_wavenumber_headers(cols):
    """
    Convert numeric-like column headers to whole-number strings (e.g., '3997.030239' -> '3997').
    Non-numeric headers remain as-is. Duplicates get suffixes (_1, _2, ...).
    """
    new_cols = []
    seen = {}
    for c in cols:
        original = str(c)
        try:
            candidate = str(int(round(float(original))))  # round to integer wavenumber
        except Exception:
            candidate = original

        if candidate in seen:
            seen[candidate] += 1
            candidate = f"{candidate}_{seen[candidate]}"
        else:
            seen[candidate] = 0

        new_cols.append(candidate)
    return new_cols


# -------------------- Shell (Navbar + Hero + FAB) — NO TOP DOCK --------------------
def inject_shell():
    st.markdown(
        f"""
<style>
    .stApp {{ background-color: white; }}

  :root {{ 
    --txt: {"#e5e7eb" if theme=="Dark" else "#0f172a"};
    --subtxt: {"#cbd5e1" if theme=="Dark" else "#334155"};
    --bg: {"#0b1220" if theme=="Dark" else "#ffffff"};
    --card: {"#0f172a" if theme=="Dark" else "#ffffff"};
    --muted: {"#0b1220" if theme=="Dark" else "#f8fafc"};
    --border: {"#172033" if theme=="Dark" else "#e6e9f2"};
    --shadow: {"rgba(0,0,0,.65)" if theme=="Dark" else "rgba(0,0,0,.06)"};
  }}
  html {{ scroll-behavior: smooth; }}
  .main {{ padding-top: 0 !important; }}
  body {{
    color: var(--txt);
    background:
      radial-gradient(1200px 600px at -10% -10%, rgba(99,102,241,.20), transparent 60%),
      radial-gradient(1200px 600px at 110% -10%, rgba(20,184,166,.15), transparent 60%),
      url('{paper_url}') repeat fixed;
    background-size: cover;
  }}
  .block-container {{ max-width: 1200px; }}

  /* Sticky navbar */
  .sticky-nav {{
    position: fixed; top: 0; left: 0; right: 0; z-index: 1000;
    backdrop-filter: saturate(140%) blur(10px);
    background: linear-gradient(180deg, rgba(255,255,255,{0.08 if theme=='Dark' else 0.8}), rgba(255,255,255,{0.04 if theme=='Dark' else 0.65}));
    border: 1px solid var(--border);
    box-shadow: 0 6px 20px var(--shadow);
  }}
  .nav-wrap {{
    max-width: 1200px; margin: 0 auto; padding: 10px 16px;
    display: flex; flex-wrap: wrap; gap: 10px; align-items: center;
  }}
  .brand {{ font-weight: 800; letter-spacing:.3px; margin-right: 8px; opacity:.9; }}
  .nav-btn {{
    display:inline-block; padding:10px 14px; border-radius:14px;
    border: 1px solid var(--border); text-decoration:none; font-weight:700; color:var(--txt);
    background: linear-gradient(180deg, var(--card), {('#0b1220' if theme=='Dark' else '#f6f7fb')});
    box-shadow: 0 2px 0 rgba(0,0,0,.04), inset 0 0 0 1px rgba(255,255,255,.08);
    transition: transform .06s ease, border-color .2s ease, box-shadow .2s ease;
  }}
  .nav-btn:hover {{ transform: translateY(-1px); border-color: var(--accent); box-shadow: 0 8px 26px rgba(0,0,0,.08); }}

  .top-spacer {{ height: 2px; }}

  .block-container {{ 
    max-width: 100% !important;
    padding-left: 5 !important;
    padding-right: 5 !important;
  }}
  
    .hero {{
    position: relative;
    height: 560px;                /* keep your chosen height */
    width: 98%;                   /* 90% of page width */
    margin: 0 auto;               /* center horizontally */
    border-radius: 25px;          /* round border all around */
    overflow: hidden;
    border: 1px solid var(--border);
    box-shadow: 0 20px 60px var(--shadow);
    background-image: linear-gradient(
                        180deg,
                        rgba(0,0,0,{0.45 if theme=='Dark' else 0.15}),
                        rgba(0,0,0,{0.35 if theme=='Dark' else 0.10})
                      ),
                      url('{hero_url}');
    background-size: cover;
    background-position: center;
    background-attachment: fixed;
    display: grid;
    place-items: center;
  }}
 
  .hero h1 {{ color: white; text-shadow: 0 10px 30px rgba(0,0,0,.40); font-size: 56px; margin: 0; }}
  .hero p {{ color: #eef2ff; margin-top: 10px; font-weight: 600; letter-spacing:.2px; }}

  /* Floating Back-to-Top */
  .fab {{
    position: fixed; right: 18px; bottom: 18px; z-index: 998;
    width: 48px; height: 48px; border-radius: 50%; display: grid; place-items: center;
    background: var(--accent); color: white; font-size: 20px; font-weight: 800; text-decoration:none;
    box-shadow: 0 16px 40px rgba(0,0,0,.25); transition: transform .15s ease;
  }}
  .fab:hover {{ transform: translateY(-2px) scale(1.02); }}
 
  /* Floating Back-to-Top */
  .fab {{
    position: fixed; right: 18px; bottom: 18px; z-index: 998;
    width: 48px; height: 48px; border-radius: 50%; 
    display: grid; place-items: center;
    background: var(--accent); color: white; 
    font-size: 20px; font-weight: 800; text-decoration:none;
    box-shadow: 0 16px 40px rgba(0,0,0,.25); 
    transition: transform .15s ease;
  }}
  .fab:hover {{ 
    transform: translateY(-2px) scale(1.02); 
  }}
 
  /* Cards */
  .card {{
    border: 1px solid var(--border); border-radius: 16px; padding: 18px 16px;
    background: var(--card); box-shadow: 0 10px 30px var(--shadow); margin-bottom: 18px;
  }}
  .badge {{
    display:inline-block; padding: 4px 10px; border-radius: 999px;
    background: var(--accent); color: white; font-weight:700; font-size:12px; letter-spacing:.3px; margin-bottom:10px;
  }}

  /* Anchor spacing (for sticky nav) */
  h2, h3, [data-anchor] {{ scroll-margin-top: 100px; }}
     scroll-behavior: smooth;
</style>
  
<div id="top" class="sticky-nav">
  <div class="nav-wrap">
    <span class="brand">🤖 ML Studio</span>
    <a class="nav-btn" href="#data">Data</a>
    <a class="nav-btn" href="#preprocess">Preprocess</a>
    <a class="nav-btn" href="#train">Train</a>
    <a class="nav-btn" href="#compare">Compare</a>
    <a class="nav-btn" href="#explain">Explain</a>
    <a class="nav-btn" href="#predict">Predict</a>
  </div>
</div>
 
<div class="hero">
  <div style="text-align:center">
    <h1 id="hero-title">🧬 AI-Powered Vibrational Spectroscopy <br>for Bone Health</h1>
    <br><br> 
    <p style="font-size: 20px;" >• 🔬Bone Spectral Classification and Composition Analysis Workspace</p>
  </div>
</div>
 
""",

        unsafe_allow_html=True,
    )


# <a class="fab" href="#hero-title">↑</a>

# Inject the shell UI (no top dock tables)
inject_shell()

# -------------------- Sections --------------------
def section(title: str, anchor: str):
    st.markdown(f"<a id='{anchor}'></a>", unsafe_allow_html=True)
    st.markdown(f"<div class='card'><span class='badge'>{title.split(':')[0]}</span><h3>{title}</h3>", unsafe_allow_html=True)
    return st.container()



# ===== Dedicated Method Panel (separate from main project UI) =====
if st.session_state.get("method_panel_active", False):
    _method = st.session_state.get("method_choice_sidebar", "SVM")
    st.markdown("<a id='method-panel'></a>", unsafe_allow_html=True)
    # Create two columns on the same line
    col1, col2 = st.columns([6, 1])  # adjust ratio for spacing

    with col1:
        st.markdown(
            f"<h3 style='margin:0'>{_method}</h3>",
            unsafe_allow_html=True
        )

    with col2:
        if st.button("Back"):
            st.session_state.method_panel_active = False
            st.rerun()

    if _method == "SVM":
        st.info("**SVM** — Support Vector Machine. "
                "")

    elif _method == "PCA":
        st.info("**PCA** — Principal Component Analysis ")

    else:  # PLS-DA
        st.info("**PLS-DA** — Partial Least Squares Discriminant Analysis")

    # Optional placeholders you can remove later
    st.markdown("##### 🔎 Quick Notes")
    st.write("- Keep this panel for method-specific outputs, comparisons, or exports.")

    st.markdown("</div>", unsafe_allow_html=True)

    # 🔒 IMPORTANT: stop rendering the rest of the Analysis page while panel is open
    st.stop()

############################################# [ Browse the file ] #############################################################################


############################################# [ Browse the file ] #############################################################################

# ======================= MAIN PAGE ==========================
if "page" not in st.session_state:
    st.session_state.page = "home"

if st.session_state.page == "home":
    # ⬇️ Center the button using columns
    col_left, col_center, col_right = st.columns([4, 4, 1])
    with col_center:

        st.markdown("<div style='margin-top:50px;'>", unsafe_allow_html=True)

        start_click = st.button("Start Analysis")

    if start_click:
        st.session_state.page = "analysis"
        st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)

elif st.session_state.page == "know_more":

    if st.button("⬅️ Back to Analysis"):
        st.session_state.page = "analysis"
        st.rerun()

# ======================= ANALYSIS PAGE ==========================
elif st.session_state.page == "analysis":

    # ===== Sticky in-page tabs (no changes to your sections required) =====
    import streamlit.components.v1 as components

    ####################################### [ Upload the document ] #######################################

    uploaded_file = st.file_uploader(f"**Upload FTIR Excel file**", type=["xlsx"])  # File uploader for FTIR data


    if uploaded_file:

            ####################################### [ # Sidebar header for controls] #######################################

            st.sidebar.header("🔧 Controls")

            ####################################### [ # Sidebar for main information] #######################################
            # < button class ="sb-btn" id="go-methods" > 📘 Overview of Machine Learning Methods < / button >
            # "go-methods": "📘 Overview of Machine Learning Methods",

            with st.sidebar:

                import streamlit.components.v1 as components

                components.html("""
                <style>
                  .sb-jumps { display:flex; flex-direction:column; gap:8px; }
                  .sb-btn {
                    border:1px solid #e5e7eb; border-radius:10px; padding:8px 10px;
                    font-size:14px; background:#f8fafc; cursor:pointer; text-align:left;
                    transition:transform .12s ease, background .12s ease;
                  }
                  .sb-btn:hover{ background:#eef2ff; transform:translateY(-1px) }
                  
                    scroll-behavior: smooth;
                </style>

                <div class="sb-jumps">
                  
                  <button class="sb-btn" id="go-shap">🔎 SHAP Analysis</button>
                  <button class="sb-btn" id="go-models">📋 Model Results with Detailed Metrics</button>
                  <button class="sb-btn" id="go-pca">🔬 PCA: Principal Component Analysis</button>
                  <button class="sb-btn" id="go-pls">📊 PLS: Partial Least Squares Regression</button>
                  <button class="sb-btn" id="go-plsda">📊 PLS-DA: Partial Least Squares Discriminant Analysis</button> 
                  <button class="sb-btn" id="go-bands">🧪 Molecular Interpretation from Spectral Bands</button>
                </div>

                <script>
                  const TARGETS = {

                    "go-shap": "🔎 SHAP Analysis",
                    "go-models": "📋 Model Results with Detailed Metrics",
                    "go-pca": "🔬 PCA: Principal Component Analysis",
                    "go-pls": "📊 PLS: Partial Least Squares Regression",
                    "go-plsda": "📊 PLS-DA: Partial Least Squares Discriminant Analysis",
                    "go-bands": "🧪 Molecular Interpretation from Spectral Bands"
                  };

                  function findHeaderByText(doc, text){
                    const hs = doc.querySelectorAll('h1,h2,h3,h4,h5,h6,summary');
                    for(const h of hs){
                      if(h.innerText.trim().startsWith(text)) return h;
                    }
                    return null;
                  }

                  function smoothScrollTo(text){
                    try {
                      const el = findHeaderByText(window.parent.document, text) || findHeaderByText(document, text);
                      if(el){ el.scrollIntoView({behavior:'smooth', block:'start'}); }
                    } catch(e){
                      const el = findHeaderByText(document, text);
                      if(el){ el.scrollIntoView({behavior:'smooth', block:'start'}); }
                    }
                  }

                  for(const id in TARGETS){
                    const btn = document.getElementById(id);
                    if(btn){ btn.addEventListener('click', ()=> smoothScrollTo(TARGETS[id])); }
                  }
                </script>
                """, height=300)

            st.subheader("📘 Overview of Machine Learning Methods")
            # === “Know More About Methods” (tabs) — show after file upload ===
            with st.expander("📚 Know More", expanded=False):

                # st.title("📚 Know More About Machine Learning Methods")
                know_tab = st.tabs(["SVM", "PCA", "PLS-DA"])

                with know_tab[0]:
                    st.subheader("🧠 Support Vector Machine (SVM)")
                    st.markdown("""
                    - **What it is:** Supervised classifier that finds a decision boundary with the **widest margin**.
                    - **Why it’s good for spectra:** Handles **high-dimensional** FTIR features; robust with the **kernel trick** (RBF).
                    - **Tune:** `C` (regularization), `gamma` (RBF width), `kernel` (linear/RBF/poly).
                    - **Tips:** Scale features; try **Stratified** CV; watch for overfitting with large `gamma`.
                    """)

                with know_tab[1]:
                    st.subheader("📊 Principal Component Analysis (PCA)")
                    st.markdown("""
                    - **What it is:** Unsupervised **dimensionality reduction** to components (PCs) with max variance.
                    - **Why it’s useful:** Visualizing separability, denoising, and **feature compression**.
                    - **Outputs:** Scores, loadings, explained variance; inspect **PC1–PC3** first.
                    """)

                with know_tab[2]:
                    st.subheader("📈 Partial Least Squares Discriminant Analysis (PLS-DA)")
                    st.markdown("""
                    - **What it is:** Supervised projection maximizing covariance between **X (spectra)** and **Y (labels)**.
                    - **Why it’s useful:** Combines dimensionality reduction + classification for spectral workflows.
                    - **Tune:** Number of LVs; validate with **RMSEC/RMSECV**, **R²**, and permutation tests if possible.
                    """)

            ############################################### [upload file ] ######################################################################

            xls = pd.ExcelFile(uploaded_file)  # Load Excel file

            # ✅ Dynamically list available sheet names
            sheet_name = st.selectbox("📄 Select a sheet to analyze:", xls.sheet_names, index=0)

            df = xls.parse(sheet_name)  # Parse selected sheet
            df.columns = df.columns.astype(str)  # Ensure column names are strings

            # ✅ Clean column headers (convert float-like to whole numbers)
            df.columns = coerce_wavenumber_headers(df.columns.astype(str))

            # ✅ Drop rows with missing values (if you want this cleaning step)
            df = df.dropna()  # Remove rows with missing values

#################################################### [Sidebar ] #####################################################

            # selected_range = st.sidebar.slider("Select Wavenumber Range (cm⁻¹)", 400, 8500, (800, 1200),  step=1)  # Slider for wavenumber range
            # ================= Wavenumber Range (two-way sync; no duplicate value setting) =================
            WN_MIN_BOUND, WN_MAX_BOUND = 400, 8500
            DEFAULT_MIN, DEFAULT_MAX = 800, 1200

            # 1) Initialize once (defaults live in session_state)
            if "wn_range" not in st.session_state:
                st.session_state.wn_range = (DEFAULT_MIN, DEFAULT_MAX)
            if "wn_min" not in st.session_state:
                st.session_state.wn_min = DEFAULT_MIN
            if "wn_max" not in st.session_state:
                st.session_state.wn_max = DEFAULT_MAX


            def _clamp(v, lo=WN_MIN_BOUND, hi=WN_MAX_BOUND):
                return max(lo, min(hi, int(v)))


            def _on_numbers_change():
                # When user types numbers, push into slider state
                mn = _clamp(st.session_state.wn_min)
                mx = _clamp(st.session_state.wn_max)
                if mn >= mx:
                    mx = min(WN_MAX_BOUND, mn + 1)
                st.session_state.wn_min = mn
                st.session_state.wn_max = mx
                st.session_state.wn_range = (mn, mx)


            def _on_slider_change():
                # When user drags slider, push into number inputs
                mn, mx = st.session_state.wn_range
                mn = _clamp(mn);
                mx = _clamp(mx)
                if mn >= mx:
                    mx = min(WN_MAX_BOUND, mn + 1)
                st.session_state.wn_range = (mn, mx)
                st.session_state.wn_min = mn
                st.session_state.wn_max = mx


            st.sidebar.markdown("### ⚙️ Select Wavenumber Range (cm⁻¹)")

            # 2) Number inputs (NO `value` since we use keys)
            c1, c2 = st.sidebar.columns(2)
            with c1:
                st.number_input("Min", min_value=WN_MIN_BOUND, max_value=WN_MAX_BOUND,
                                step=1, key="wn_min", on_change=_on_numbers_change)
            with c2:
                st.number_input("Max", min_value=WN_MIN_BOUND, max_value=WN_MAX_BOUND,
                                step=1, key="wn_max", on_change=_on_numbers_change)

            # 3) Slider (NO `value` since we use key)
            st.sidebar.slider("Adjust Range",
                              min_value=WN_MIN_BOUND, max_value=WN_MAX_BOUND,
                              step=1, key="wn_range", on_change=_on_slider_change)

            # 4) Use this everywhere below
            selected_range = st.session_state.wn_range
            st.sidebar.caption(f"Selected: {selected_range[0]} – {selected_range[1]} cm⁻¹")

            ###################################### [ Disply the data ] #########################################################

            st.subheader("🔍 Uploaded Data Preview")
            st.dataframe(df.head(10))
            label_column = df.columns[0]  # First column as label column
            data_columns = df.columns[1:]  # Remaining columns

            # ✅ Extract numeric wavenumber values (first part before '/') for filtering
            wavenumber_map = {}
            for col in data_columns:
                try:
                    numeric_value = float(
                        col.split("/")[0].strip())  # Extract first number (e.g., "1030" from "1030/1670")
                    wavenumber_map[col] = numeric_value
                except ValueError:
                    st.warning(f"⚠ Skipping column: {col} (cannot extract numeric value)")

            # ✅ Filter columns based on selected range using extracted numeric values
            selected_columns = [
                col for col, wn in wavenumber_map.items()
                if selected_range[0] <= wn <= selected_range[1]
            ]

            # ✅ Features (retain original column names) and labels
            X = df[selected_columns].astype(float)
            y = df[label_column]

            # ✅ Encode labels dynamically (works for multi-class)
            from sklearn.preprocessing import LabelEncoder

            label_enc = LabelEncoder()
            y_encoded = label_enc.fit_transform(y)

            ################################################################################################################
            ################################################################################################################


            ################################################################################################################
            ################################################################################################################

            ###################################### [ Display the selected data ] #############################

            st.subheader("📌 Data Selected for Analysis")
            st.markdown(f"**Selected Wavenumber Range:** {selected_range[0]} cm⁻¹ to {selected_range[1]} cm⁻¹")
            st.dataframe(X.head(10))

            ################################################################################################################

            ##### random sampling
            # ✅ Show class counts and enable random sampling BEFORE train-test split
            st.sidebar.markdown("### 🎯 Class Sampling")
            class_counts = df[label_column].value_counts()
            st.sidebar.write("**Class Distribution:**")
            st.sidebar.write(class_counts)

            # Initialize sample counts in session_state if not present
            if "sample_counts" not in st.session_state:
                st.session_state.sample_counts = {cls: int(class_counts[cls]) for cls in class_counts.index}

            # --- Monkeypatch st.button ONLY for the two SHAP buttons in your Final Script
            if "_original_button" not in st.session_state:
                st.session_state._original_button = st.button

            import streamlit as st

            # --- 1) Auto-click only the SHAP plot buttons (works for both Non-Quant & Quant labels) ---
            if "_original_button" not in st.session_state:
                st.session_state._original_button = st.button  # keep original

            def _auto_button(label, *args, **kwargs):
                auto_labels = {
                    "Generate Beeswarm Plot (Non-Quant, All Rows)",
                    "Generate Waterfall Plot (Non-Quant, All Rows)",
                    "Generate Beeswarm Plot",  # Quant block uses these bare labels
                    "Generate Waterfall Plot",  # Quant block uses these bare labels
                }
                if label in auto_labels:
                    return True
                return st.session_state._original_button(label, *args, **kwargs)

            #####################################################################################################################

            st.button = _auto_button

            # Custom small button style
            # Add custom style for the small reset button
            st.markdown("""
            <style>
            /* Force a very small Reset Sample Selection button */
            div.stButton.reset-sample {
                display: inline-block;       /* Make it inline and shrink wrap */
                width: auto !important;      /* Override default width */
                margin: 0 !important;        /* Remove spacing around it */
            }

            div.stButton.reset-sample > button {
                background-color: #9C2007;
                color: white;
                font-size: 12px !important;    /* Extremely small font */
                font-weight: bold;

                border: none;
                cursor: pointer;
                box-shadow: 0px 0px 2px rgba(0, 0, 0, 0.3);
                height: 1px !important;      /* Tiny height */
                min-width: 1px !important;   /* Force narrow width */
                line-height: 1px !important; /* Prevent extra height */
            }

            div.stButton.reset-sample > button:hover {
                background-color: #7a0000;
                transform: scale(1.05);
            }
          scroll-behavior: smooth;
          
        </style> 
                """, unsafe_allow_html=True )

            ################################################################################################################
            #######################################  [ OIM-WT Result ]  ####################################################

            # # Reset button
            # if st.sidebar.button("🔄 Reset"):
            #     st.session_state.sample_counts = {cls: int(class_counts[cls]) for cls in class_counts.index}
            #
            # # Input: number of samples per class
            # sample_counts = {}
            # for cls in class_counts.index:
            #     max_val = int(class_counts[cls])
            #
            #     # ✅ Ensure stored value does not exceed max
            #     if st.session_state.sample_counts.get(cls, max_val) > max_val:
            #         st.session_state.sample_counts[cls] = max_val
            #
            #     # Number input with clamped default
            #     sample_counts[cls] = st.sidebar.number_input(
            #         f"Select samples for '{cls}' (max {max_val})",
            #         min_value=1,
            #         max_value=max_val,
            #         value=int(st.session_state.sample_counts[cls]),  # safe default
            #         step=1,
            #         key=f"sample_{cls}"
            #     )
            #
            #     st.session_state.sample_counts[cls] = sample_counts[cls]


            ################################################################################################################
            ################################################################################################################

            # --- right after computing class_counts ---
            # class_counts = df[label_column].value_counts()

            # Ensure the session dict exists
            if "sample_counts" not in st.session_state:
                st.session_state.sample_counts = {}

            # 1) Re-sync keys with current classes (add missing; drop stale)
            current_classes = list(class_counts.index)
            # add any new classes with full count default
            for cls in current_classes:
                st.session_state.sample_counts.setdefault(cls, int(class_counts[cls]))
            # drop any classes that aren't present anymore
            for stale in list(st.session_state.sample_counts.keys()):
                if stale not in current_classes:
                    del st.session_state.sample_counts[stale]

            # Optional: a HARD reset button that also clears saved choices
            if st.sidebar.button("🔄 Reset"):
                st.session_state.sample_counts = {cls: int(class_counts[cls]) for cls in current_classes}

            # 2) Inputs (safe defaults + clamping)
            sample_counts = {}
            for cls in current_classes:
                max_val = int(class_counts[cls])

                # clamp any stored value to valid range
                stored = int(st.session_state.sample_counts.get(cls, max_val))
                if stored > max_val:
                    stored = max_val
                if stored < 1:
                    stored = 1

                sample_counts[cls] = st.sidebar.number_input(
                    f"Select samples for '{cls}' (max {max_val})",
                    min_value=1,
                    max_value=max_val,
                    value=stored,  # ✅ SAFE DEFAULT
                    step=1,
                    key=f"sample_{cls}"
                )
                st.session_state.sample_counts[cls] = int(sample_counts[cls])

            ################################################################################################################
            ################################################################################################################


            # Random sampling
            df_sampled = pd.concat([
                df[df[label_column] == cls].sample(n=int(sample_counts[cls]), random_state=42)
                for cls in class_counts.index
            ])

            # Update X and y
            X = df_sampled[selected_columns].astype(float)
            y = df_sampled[label_column]
            from sklearn.preprocessing import LabelEncoder

            label_enc = LabelEncoder()
            y_encoded = label_enc.fit_transform(y)

            # Show sampled counts
            st.sidebar.write("✅ **Sampled Class Counts:**")
            st.sidebar.write(df_sampled[label_column].value_counts())

            ################################################################################################################

            ###################################### [ Train-Test Split Slider] ######################################
            # ✅ NEW: Train-Test Split Slider
            split_ratio = st.sidebar.slider(
                "Select Train-Test Split Ratio (Test Size %)",
                min_value=10, max_value=50, value=20, step=5,
                help="Choose percentage of data for testing (default is 20%)"
            )

            st.sidebar.markdown(f"**Training: {100 - split_ratio}% | Testing: {split_ratio}%**")

            model_options = ["SVM", "Logistic Regression", "KNN", "Random Forest", "XGBoost"]  # List of model options
            selected_models = st.sidebar.multiselect("Select Classification Models", model_options, default=["SVM"])  # Multiselect for models

            ##show_histogram = st.sidebar.checkbox("Show Histogram of Raw Features")  # Option to show raw feature histograms

            run_all = st.sidebar.checkbox("Run All Models")  # Checkbox to run all models
            show_models = st.sidebar.checkbox("Show Model Results", value=True)  # Option to show model results
            show_pca = st.sidebar.checkbox("Show PCA Visualizations", value=True)  # Option to show PCA
            show_pls = st.sidebar.checkbox("Show PLS Analysis", value=True)  # Option to show PLS
            show_plsda = st.sidebar.checkbox("Show PLSDA Analysis", value=True)  # Option to show PLSDA

            st.sidebar.header("📊 Spectral Plots Display")  # Sidebar header for spectral plots

            show_mean = st.sidebar.checkbox("Mean of FTIR Spectra", value=False)  # Checkbox to show mean spectra
            show_stddev = st.sidebar.checkbox("Standard Deviation of FTIR Spectra",value=False)  # Checkbox to show standard deviation
            show_mean_stddev = st.sidebar.checkbox("Mean + Standard Deviation of FTIR Spectra", value=False)  # Checkbox for mean+std
            show_minimum = st.sidebar.checkbox("Minimum of FTIR Spectra", value=False)  # Checkbox for minimum spectra
            show_maximum = st.sidebar.checkbox("Maximum of FTIR Spectra", value=False)  # Checkbox for maximum spectra

            #####################################################################################################
            # ====================== Spectrum Plot with Bond Annotations ======================
            show_spectrum = st.sidebar.checkbox("📈 Show Raw Spectrum with Bond Annotations", value=False)

            if show_spectrum:
                st.subheader("📈 FTIR Spectrum with Bond Annotations")

                # --- Plot raw spectra for all samples (X is already filtered by range) ---
                fig, ax = plt.subplots(figsize=(10, 6))

                # Plot each sample
                for idx, row in X.iterrows():
                    ax.plot(X.columns.astype(float), row.values, alpha=0.4, lw=1)

                # Labels
                ax.set_xlabel("Wavenumber (cm⁻¹)")
                ax.set_ylabel("Absorbance")
                ax.set_title("FTIR Spectrum with Bond Annotations")

                # ✅ Bond Annotation Mapping Table
                bond_annotations_table = [
                    (8500, "O–H Stretching and Bending (Water)"),
                    (7000, "O–H Stretching (Water)"),
                    (6688, "N–H Stretching (Protein/Collagen)"),
                    (5800, "CH₂ Stretching (Lipid)"),
                    (5200, "O–H Stretching and Bending (Water)"),
                    (4890, "N–H Bending (Protein/Collagen)"),
                    (4610, "C–H Stretching & Deformation (Protein/Collagen)"),
                    (4310, "Sugar Ring Vibrations (Proteoglycan)"),
                    ((3600, 3200), "O–H Stretching (Water/Hydroxyl)"),
                    ((3500, 3300), "N–H Stretching (Proteins)"),
                    ((3000, 2800), "C–H Stretching (Lipids, CH₂)"),
                    ((1750, 1650), "C=O Stretching (Proteins/Lipids)"),
                    (1550, "Amide II (Proteins)"),
                    (1338, "CH₂ Side Chain Bending (Collagen)"),
                    ((1100, 900), "PO₄³⁻ Stretching (Bone Mineral)"),
                    ((890, 850), "CO₃²⁻ Bending (Carbonate)"),
                    (1740, "C=O Stretching (Ester, Lipids)"),
                    (1650, "Amide I (Proteins)"),
                    (1630, "Water O–H Bending (Water)"),
                    ((1200, 1000), "C–O Stretching (Alcohols/Ethers)"),
                    (1115, "HPO₄²⁻ Stretching (Bone mineral)"),
                    (1060, "Sugar Ring C–O Stretch (Carbohydrates)"),
                    (1030, "PO₄³⁻ Stretching (Bone Mineral)"),
                    (875, "CO₃²⁻ Bending (Carbonates)"),
                    (856, "C–S Bending (Proteoglycans)"),
                    ((900, 800), "Aromatic C–H Bending (Fingerprint region)")
                ]

                # Add vertical lines + labels for annotations
                for entry in bond_annotations_table:
                    if isinstance(entry[0], tuple):  # range
                        wn_min, wn_max = entry[0]
                        ax.axvspan(wn_min, wn_max, color="red", alpha=0.08)
                        ax.text((wn_min + wn_max) / 2, ax.get_ylim()[1] * 0.9, entry[1],
                                rotation=90, fontsize=8, ha="center", va="top", color="red")
                    else:  # single wavenumber
                        wn = entry[0]
                        ax.axvline(wn, color="red", linestyle="--", alpha=0.6)
                        ax.text(wn, ax.get_ylim()[1] * 0.95, entry[1],
                                rotation=90, fontsize=8, ha="center", va="top", color="red")

                st.pyplot(fig)
                plt.close(fig)

            #####################################################################################################

            #
            # ###########################################  [New Script - END: SHAP Analysis ]  ##########################################################
            # ###########################################################################################################################################
            #
            # ================================
            # 🔎 SHAP Analysis (Complete Block) — Selection-Aware (FTIR Project)
            # Shows:
            #   📊 Feature Importance (Mean |SHAP|) – All Rows
            #   📌 SHAP Beeswarm Plot (All Rows)
            #   💧 SHAP Waterfall Plot (Individual Sample)
            # + Adds non-technical interpretation + mineral/collagen band meaning
            # ================================

            st.subheader("🔎 SHAP Analysis")

            with st.expander("🔎 SHAP Analysis (Selection-Aware)", expanded=False):

                # ---------- Imports (local to block) ----------
                import numpy as np
                import pandas as pd
                import matplotlib.pyplot as plt
                import shap
                import hashlib
                from sklearn.model_selection import train_test_split
                from sklearn.preprocessing import LabelEncoder

                # Prefer XGBoost; fallback to RF
                _HAS_XGB = True
                try:
                    from xgboost import XGBClassifier
                except Exception:
                    _HAS_XGB = False
                    from sklearn.ensemble import RandomForestClassifier as _RF


                # ---------- Helpers ----------
                def _filter_cols_by_range(cols, selected_range_tuple):
                    """Keep columns whose numeric wavenumber (supports '1600_1') overlaps selected_range."""
                    lo, hi = float(selected_range_tuple[0]), float(selected_range_tuple[1])
                    keep = []
                    for c in cols:
                        s = str(c)
                        base = s.split("_")[0]
                        try:
                            w = float(base)
                            if lo <= w <= hi:
                                keep.append(c)
                        except Exception:
                            pass
                    return keep


                def _resolve_shap_matrix(shap_vals, class_idx=None, reduce_mode="mean"):
                    """Return SHAP as 2D matrix (n_samples, n_features). Handles list (multiclass), 3D, 2D."""
                    if isinstance(shap_vals, list):
                        if class_idx is not None:
                            return np.asarray(shap_vals[class_idx])
                        stack = np.stack([np.asarray(a) for a in shap_vals], axis=-1)  # (n,d,C)
                        if reduce_mode == "sum":
                            return stack.sum(axis=2)
                        if reduce_mode == "absmax":
                            idx = np.argmax(np.abs(stack), axis=2)
                            rows = np.arange(stack.shape[0])[:, None]
                            cols = np.arange(stack.shape[1])[None, :]
                            return stack[rows, cols, idx]
                        return stack.mean(axis=2)

                    arr = np.asarray(shap_vals)
                    if arr.ndim == 3:
                        if class_idx is not None:
                            return arr[:, :, class_idx]
                        if reduce_mode == "sum":
                            return arr.sum(axis=2)
                        if reduce_mode == "absmax":
                            idx = np.argmax(np.abs(arr), axis=2)
                            rows = np.arange(arr.shape[0])[:, None]
                            cols = np.arange(arr.shape[1])[None, :]
                            return arr[rows, cols, idx]
                        return arr.mean(axis=2)

                    if arr.ndim == 2:
                        return arr

                    raise ValueError(f"Unsupported SHAP shape: {arr.shape}")


                def _reduce_expected_value(expected_value, class_idx=None, reduce_mode="mean"):
                    ev = np.asarray(expected_value)
                    if ev.ndim == 0:
                        return float(ev)
                    if class_idx is not None:
                        return float(ev[class_idx])
                    if reduce_mode == "sum":
                        return float(ev.sum())
                    if reduce_mode == "absmax":
                        return float(ev[np.argmax(np.abs(ev))])
                    return float(ev.mean())


                def _fit_model(X_train, y_train):
                    """Train XGB (or RF fallback)."""
                    if len(np.unique(y_train)) < 2:
                        raise ValueError("Only one class present; need ≥2 classes for SHAP.")
                    if _HAS_XGB:
                        m = XGBClassifier(
                            n_estimators=500,
                            max_depth=3,
                            learning_rate=0.05,
                            subsample=0.9,
                            colsample_bytree=0.9,
                            reg_lambda=1.0,
                            random_state=42,
                            eval_metric="logloss"
                        )
                        m.fit(X_train, y_train)
                        return m
                    rf = _RF(n_estimators=500, random_state=42)
                    rf.fit(X_train, y_train)
                    return rf


                # ============================================================
                # ✅ 1) Read selection (wavenumber range) + base X/y
                # ============================================================
                if "selected_range" not in globals() and "selected_range" not in locals():
                    st.error("selected_range not found. Ensure the wavenumber range slider runs before SHAP.")
                    st.stop()
                selected_range_shap = selected_range

                if "X" not in globals() and "X" not in locals():
                    st.error("X not found. Ensure feature matrix X is created before SHAP.")
                    st.stop()
                if "y" not in globals() and "y" not in locals():
                    st.error("y not found. Ensure label vector y is created before SHAP.")
                    st.stop()

                # Only numeric FTIR features
                X_base = X.copy().select_dtypes(include=[np.number]).astype(float)
                if X_base.shape[1] == 0:
                    st.error("No numeric FTIR features found in X.")
                    st.stop()

                # Prefer selected_columns (if your app already builds it); else infer from range
                if "selected_columns" in globals() or "selected_columns" in locals():
                    try:
                        sc = selected_columns
                        if isinstance(sc, (list, tuple)) and len(sc) > 0:
                            sc_str = [str(c) for c in sc]
                            cols_in_X = [c for c in X_base.columns if (str(c) in sc_str) or (c in sc)]
                            if len(cols_in_X) > 0:
                                X_sel = X_base[cols_in_X].copy()
                            else:
                                X_sel = X_base[_filter_cols_by_range(X_base.columns, selected_range_shap)].copy()
                        else:
                            X_sel = X_base[_filter_cols_by_range(X_base.columns, selected_range_shap)].copy()
                    except Exception:
                        X_sel = X_base[_filter_cols_by_range(X_base.columns, selected_range_shap)].copy()
                else:
                    X_sel = X_base[_filter_cols_by_range(X_base.columns, selected_range_shap)].copy()

                if X_sel.shape[1] == 0:
                    st.error(f"No FTIR features found in range {selected_range_shap[0]}–{selected_range_shap[1]} cm⁻¹.")
                    st.stop()

                y_clean = pd.Series(y).astype(str).values
                le = LabelEncoder()
                y_enc = le.fit_transform(y_clean)

                # ============================================================
                # ✅ 2) Reset plots when selection changes (prevents stale plots)
                # ============================================================
                sig_payload = {
                    "sheet_name": str(sheet_name) if "sheet_name" in globals() else "NA",
                    "range": (float(selected_range_shap[0]), float(selected_range_shap[1])),
                    "n_features": int(X_sel.shape[1]),
                    "feature_first": str(X_sel.columns[0]),
                    "feature_last": str(X_sel.columns[-1]),
                }
                sig = hashlib.md5(repr(sig_payload).encode("utf-8")).hexdigest()

                if st.session_state.get("shap_sig") != sig:
                    st.session_state["shap_sig"] = sig
                    st.session_state["shap_bees_fig"] = None
                    st.session_state["shap_wf_fig"] = None

                # ============================================================
                # ✅ 3) Debug: confirm selection is changing
                # ============================================================
                # st.caption("✅ SHAP is computed on the CURRENT selection (this should change when you move the range slider).")

                st.write("Selected range:", f"{int(selected_range_shap[0])}–{int(selected_range_shap[1])} cm⁻¹")

                # st.write("Selected feature count:", int(X_sel.shape[1]))
                # st.write("Feature span:", f"{X_sel.columns[0]} → {X_sel.columns[-1]}")

                # ============================================================
                # ✅ 4) Train model
                # ============================================================
                stratify_vec = y_enc if len(np.unique(y_enc)) > 1 else None
                try:
                    X_train, X_test, y_train, y_test = train_test_split(
                        X_sel, y_enc, test_size=0.2, stratify=stratify_vec, random_state=42
                    )
                except ValueError:
                    X_train, X_test, y_train, y_test = X_sel, X_sel, y_enc, y_enc
                    st.info("Not enough rows for stratified split — using all rows for training and SHAP.")

                try:
                    model = _fit_model(X_train, y_train)
                except Exception as e:
                    st.error(f"Model training failed: {e}")
                    st.stop()

                # ============================================================
                # ✅ 5) Compute SHAP on ALL rows of CURRENT selection (global)
                # ============================================================
                try:
                    explainer = shap.TreeExplainer(model)
                    shap_raw_all = explainer.shap_values(X_sel)  # ALL rows, selection-aware
                    shap_vals_all = _resolve_shap_matrix(shap_raw_all, class_idx=None, reduce_mode="mean")
                    base_val = _reduce_expected_value(explainer.expected_value, class_idx=None, reduce_mode="mean")
                except Exception as e:
                    st.error(f"SHAP computation failed: {e}")
                    st.stop()

                # ============================================================
                # ✅ 6) Beeswarm + Feature Importance
                # ============================================================
                c1, c2 = st.columns([1, 1])

                with c1:
                    st.subheader("📌 SHAP Beeswarm Plot (All Rows)")

                    if st.button("Generate Beeswarm Plot", key=f"btn_shap_bees_{sig}"):
                        try:
                            plt.close("all")
                            shap.summary_plot(
                                shap_vals_all,
                                X_sel,
                                feature_names=[str(c) for c in X_sel.columns],
                                plot_type="dot",
                                show=False
                            )
                            st.session_state["shap_bees_fig"] = plt.gcf()
                        except Exception as e:
                            st.error(f"Beeswarm failed: {e}")

                    if st.session_state.get("shap_bees_fig") is not None:
                        st.pyplot(st.session_state["shap_bees_fig"])
                        plt.close(st.session_state["shap_bees_fig"])

                    st.markdown("""
            **How to read this plot**
            - Each horizontal row = **one wavenumber band**
            - Each dot = **one bone spectrum (one sample)**

            **X-axis (SHAP value)** → contribution to prediction  
            - **Right (+)** pushes toward one class  
            - **Left (−)** pushes toward the other class  

            **Color (feature value)**  
            - 🔴 **Red** = high absorbance  
            - 🔵 **Blue** = low absorbance  
            """)

                with c2:
                    st.subheader("📊 Feature Importance (Mean |SHAP|) – All Rows")

                    try:
                        mean_abs = np.abs(shap_vals_all).mean(axis=0)
                        feature_importance_all = pd.DataFrame({
                            "Feature (Wavenumber)": [str(c) for c in X_sel.columns],
                            "Mean |SHAP Value|": mean_abs
                        }).sort_values("Mean |SHAP Value|", ascending=False).reset_index(drop=True)

                        st.dataframe(feature_importance_all.style.format({"Mean |SHAP Value|": "{:.6f}"}), width="stretch")
                                     # use_container_width=True)

                        st.download_button(
                            "📥 Download Feature Importance CSV",
                            data=feature_importance_all.to_csv(index=False).encode("utf-8"),
                            file_name=f"shap_feature_importance_{int(selected_range_shap[0])}_{int(selected_range_shap[1])}.csv",
                            mime="text/csv",
                            key=f"dl_shap_fi_{sig}"
                        )
                    except Exception as e:
                        st.error(f"Feature importance failed: {e}")

                    st.markdown("""
            **What this table shows**
            - **Mean |SHAP|** = average absolute contribution of each wavenumber across **all spectra**
            - Higher values = the band is **more important globally** for WT vs OIM discrimination
            - It measures **magnitude**, not direction
            """)

                st.markdown("---")

                # ============================================================
                # ✅ 7) Waterfall (Individual sample explanation)
                # ============================================================
                w1, w2 = st.columns([1, 1])

                with w1:
                    # st.subheader("🎚 Select Sample Index (All Rows)")
                    # sample_idx = st.slider(
                    #     "Select Sample Index for Waterfall Plot",
                    #     0, max(0, len(X_sel) - 1), 0,
                    #     key=f"slider_shap_wf_{sig}"
                    # )
                    st.subheader("🎚 Representative Sample Explanation")

                    sample_idx = st.slider(
                        "Select a representative FTIR spectrum (row index) to explain the model’s prediction",
                        0, max(0, len(X_sel) - 1), 0,
                        key=f"slider_shap_wf_{sig}"
                    )

                    st.caption(
                        "This slider selects a single bone spectrum from the dataset. "
                        "The SHAP Waterfall plot below explains how individual FTIR bands "
                        "contribute to the model’s prediction for this representative sample."
                    )



                with w2:
                    st.subheader("💧 SHAP Waterfall Plot")

                    # st.subheader("💧 SHAP Waterfall Plot (Individual Sample)")

                    if st.button("Generate Waterfall Plot", key=f"btn_shap_wf_{sig}"):
                        try:
                            plt.close("all")
                            vals = np.asarray(shap_vals_all[sample_idx]).ravel()
                            data_row = np.asarray(X_sel.iloc[sample_idx], dtype=float)
                            m = min(len(vals), len(data_row), len(X_sel.columns))

                            explanation = shap.Explanation(
                                values=vals[:m],
                                base_values=base_val,
                                data=data_row[:m],
                                feature_names=[str(c) for c in X_sel.columns[:m]]
                            )

                            shap.plots.waterfall(explanation, show=False)
                            st.session_state["shap_wf_fig"] = plt.gcf()
                        except Exception as e:
                            st.error(f"Waterfall failed: {e}")

                    if st.session_state.get("shap_wf_fig") is not None:
                        st.pyplot(st.session_state["shap_wf_fig"])
                        plt.close(st.session_state["shap_wf_fig"])

                    st.markdown("""
            **What this plot shows (for one spectrum)**
            - Starts at the **baseline** *E[f(X)]* (average model output)
            - Each bar shows how a specific **wavenumber** shifts the prediction
            - Red bars push the prediction up; blue bars push it down
            - The final sum equals the model output for this sample
            """)

                # ============================================================
                # ✅ 8) Region-aware molecular explanation + top-band mapping
                # ============================================================
                st.markdown("---")
                # st.markdown("## ✅ SHAP results explanation")

                min_wn = int(selected_range_shap[0])
                max_wn = int(selected_range_shap[1])

                st.markdown(f"""
                You selected: **{min_wn}–{max_wn} cm⁻¹**, so SHAP is computed **only on those wavenumbers**.

                - **Feature Importance (Mean |SHAP|)** ranks bands by average impact (global importance).
                - **Beeswarm** shows per-sample impacts: dot = spectrum; x-position = contribution; color = absorbance.
                - **Waterfall** explains one sample by accumulating the strongest band contributions.
                """)

                def _region_label(lo, hi):
                    if lo <= 800 and hi <= 1200:
                        return "🧱 Mineral (Phosphate/Carbonate) region (800–1200 cm⁻¹)"
                    if lo >= 1300 and hi <= 1700:
                        return "🧬 Organic matrix / Collagen (Amide) region (1300–1700 cm⁻¹)"
                    if lo >= 3000 and hi <= 3800:
                        return "💧 Water / Hydroxyl stretching region (~3200–3600 cm⁻¹)"
                    return "📍 Mixed / broader FTIR region"

                ####st.markdown(f"### 🔬 Scientific meaning of this selected region\n**{_region_label(min_wn, max_wn)}**")

                # Reference band mapping aligned with your Molecular Interpretation table
                # mapping = [
                #     (960, 1005, "PO₄³⁻ ν₁ (phosphate stretch)", "Bone mineral (hydroxyapatite phosphate groups)"),
                #     (1100, 1135, "PO₄³⁻ ν₃ (asymmetric stretch)", "Mineral phosphate environment / maturity"),
                #     (890, 850, "CO₃²⁻ bending (carbonate)", "Carbonate substitution in apatite"),
                #     (800, 900, "PO₄ bending / lattice modes", "Mineral lattice / crystallinity-related modes"),
                #     (1600, 1700, "Amide I (C=O stretch)", "Collagen secondary structure / organic matrix"),
                #     (1510, 1580, "Amide II (N–H bend + C–N stretch)", "Collagen / protein structure"),
                #     (1330, 1350, "CH₂ side chain vibrations", "Collagen integrity marker"),
                #     (1200, 1300, "C–O / sugar ring bands", "Proteoglycans / organic matrix components"),
                # ]

                mapping = [
                    # =========================
                    # 🧱 MINERAL (Bone apatite): 800–1200 cm⁻¹
                    # =========================
                    (800, 860, "PO₄³⁻ bending / lattice modes", "Mineral lattice / crystallinity-related modes"),
                    (870, 890, "CO₃²⁻ ν₂ (carbonate bending, type B/A)",
                     "Carbonate substitution in apatite; mineral chemistry changes"),
                    # bone carbonate bands widely used :contentReference[oaicite:1]{index=1}
                    (900, 920, "PO₄³⁻ ν₁/ν₃ shoulder region", "Phosphate environment; overlaps with mineral sub-bands"),
                    (940, 970, "PO₄³⁻ ν₁ (phosphate symmetric stretch)",
                     "Primary phosphate marker in hydroxyapatite; mineral content"),
                    # bone FTIR phosphate region :contentReference[oaicite:2]{index=2}
                    (990, 1015, "PO₄³⁻ ν₁ / ν₃ mixed region (~1000)", "Mineral maturity / stoichiometry differences"),
                    (1020, 1090, "PO₄³⁻ ν₃ (asymmetric stretch, broad)",
                     "Mineral phosphate network; sensitive to composition"),
                    (1090, 1135, "PO₄³⁻ ν₃ (high side)", "Crystallinity / mineral order; phosphate environment"),
                    (1110, 1125, "HPO₄²⁻ contribution", "Immature mineral / acid phosphate contribution"),

                    # =========================
                    # 🧬 ORGANIC MATRIX + MIXED: 1200–1500 cm⁻¹
                    # =========================
                    (1200, 1300, "C–O stretching / sugar ring vibrations",
                     "Proteoglycans/glycoproteins; organic matrix carbohydrate modes"),
                    (1240, 1280, "Amide III region (protein)",
                     "Collagen-related vibrations (protein backbone/side chain coupling)"),
                    (1330, 1350, "CH₂ side chain vibrations", "Collagen integrity marker"),
                    (1400, 1470, "CO₃²⁻ ν₃ (carbonate asymmetric stretch) + CH",
                     "Carbonate substitution + matrix CH bending (overlaps)"),
                    # carbonate in apatite commonly used :contentReference[oaicite:3]{index=3}

                    # =========================
                    # 🧬 COLLAGEN / PROTEIN: 1500–1700 cm⁻¹
                    # =========================
                    (1510, 1580, "Amide II (N–H bend + C–N stretch)", "Collagen/protein secondary structure changes"),
                    (1590, 1610, "COO⁻ asymmetric stretching (carboxylate)",
                     "Collagen side chains (Asp/Glu), matrix charge; mineral–matrix interactions"),
                    (1620, 1640, "H₂O O–H bending (water)", "Bound/free water contribution near 1630 cm⁻¹"),
                    (1600, 1700, "Amide I (C=O stretch)", "Collagen secondary structure / organic matrix"),

                    # =========================
                    # 🧪 LIPIDS (mid-IR C–H stretch): 2800–3000 cm⁻¹
                    # =========================
                    (2800, 3000, "C–H stretching (CH₂/CH₃)", "Lipids + protein side chains (organic matrix content)"),

                    # =========================
                    # 💧 WATER / N–H / O–H fundamentals: 3000–3700 cm⁻¹
                    # =========================
                    (3200, 3600, "O–H stretching (water/hydroxyl, broad)",
                     "Hydration and/or hydroxyl content; broad hydrogen-bonded band"),
                    (3300, 3500, "N–H stretching (amide A/B)", "Protein (collagen) N–H stretching contribution"),

                    # =========================
                    # 🌙 NIR COMBINATION REGION: ~4000–5200 cm⁻¹
                    # =========================
                    (4100, 4500, "Combination bands (C–H / O–H / N–H)",
                     "Organic matrix + water combinations; complex overlap typical of NIR"),
                    # NIR combination region :contentReference[oaicite:4]{index=4}
                    (4500, 4900, "Combination bands (C–H stretch + deformation)", "Protein/lipid combinations (NIR)"),
                    (5000, 5200, "O–H combination region (H₂O)",
                     "Water-related combination features leading into ~5200 cm⁻¹ band"),
                    # NIR water region :contentReference[oaicite:5]{index=5}

                    # =========================
                    # 💧 STRONG NIR WATER BAND: ~5200 cm⁻¹
                    # =========================
                    (5100, 5400, "H₂O combination band (νOH + δHOH)",
                     "Major bone water band (loosely + tightly bound components)"),
                    # water combination ~5200 :contentReference[oaicite:6]{index=6}

                    # =========================
                    # 🧬 NIR C–H FIRST OVERTONE: ~5600–6100 cm⁻¹
                    # =========================
                    (5600, 6100, "C–H first overtone (CH₂/CH₃)",
                     "Protein side chains + lipids; organic matrix signal in NIR"),
                    # CH overtone region :contentReference[oaicite:7]{index=7}

                    # =========================
                    # 💧 NIR WATER OVERTONE: ~6800–7400 cm⁻¹
                    # =========================
                    (6800, 7400, "H₂O / O–H first overtone (2νOH) (broad)",
                     "Major water overtone band (~6900–7100+ cm⁻¹); hydration-related"),
                    # water overtone ~6900/7000 :contentReference[oaicite:8]{index=8}

                    # =========================
                    # ☀️ HIGH NIR O–H overtone / edge of your range: 7400–8000 cm⁻¹
                    # =========================
                    (7400, 8000, "O–H overtone shoulder / higher-order NIR",
                     "Higher-energy O–H overtone/shoulders; weaker but water/hydroxyl-linked"),
                    # general NIR band chart + OH overtone ranges :contentReference[oaicite:9]{index=9}
                ]


                def _best_assignment(w):
                    try:
                        w = float(w)
                        for lo, hi, mode, meaning in mapping:
                            a, b = min(lo, hi), max(lo, hi)
                            if a <= w <= b:
                                return mode, meaning

                        return "Band not mapped (add if needed)", "—"
                    except:
                        print("====", w)
                        w = w.split("/")
                        w = (w)
                        for lo, hi, mode, meaning in mapping:
                            a, b = min(lo, hi), max(lo, hi)
                            # if a <= w <= b:
                            #     return mode, meaning
                        return "", ""


                # Show top bands with meaning (non-technical + scientific)
                st.markdown("### 🎯 Top SHAP features")
                top = feature_importance_all.head(10).copy()
                # print('==============', s)

                try:
                    top["Mode / Assignment"] = top["Feature (Wavenumber)"].astype(str).apply(
                        lambda s: _best_assignment(float(s))[0])
                except:
                    print('in except ---- 1 ')
                    top["Mode / Assignment"] = top["Feature (Wavenumber)"].astype(str).apply(
                        lambda s: _best_assignment(s)[0])

                try:
                    top["Biochemical meaning"] = top["Feature (Wavenumber)"].astype(str).apply(
                        lambda s: _best_assignment(float(s))[1])
                except:
                    print('in except ---- 2 ')
                    top["Biochemical meaning"] = top["Feature (Wavenumber)"].astype(str).apply(
                        lambda s: _best_assignment(s)[1])


                st.dataframe(top, width="stretch")  #use_container_width=True)

                st.success(
                    f"✅ Interpretation: In the selected range **{min_wn}–{max_wn} cm⁻¹**, "
                    f"the model’s decision is being driven mainly by the bands listed above. "
                    f"These correspond to the biochemical structures active in this region (mineral vs collagen vs water)."
                )
                ########## Add script to download the SHAP Result ###########

                # ============================================================
                # 📥 Download SHAP Results (ONE Excel file, ONE sheet/tab)
                # Includes:
                #   - Selected range
                #   - Beeswarm image (if generated)
                #   - Feature Importance table
                #   - Waterfall image (if generated)
                #   - Top SHAP features table with biochemical meaning
                # NOTE: Does NOT change your SHAP logic — export only.
                # ============================================================

                # ============================================================
                # 📥 Download SHAP Results (ONE Excel file, 3 sheets)
                # Sheets:
                #   1) "Beeswarm+FI"   -> Selected range + Beeswarm image + FI table + quick analysis text
                #   2) "Waterfall"     -> Selected range + Waterfall image + explanation text
                #   3) "TopFeatures"   -> Selected range + Top SHAP features table (with assignment/meaning)
                # NOTE: Export-only. Does NOT change SHAP training/plot logic.
                # ============================================================

                try:
                    import io
                    import numpy as np
                    from openpyxl import Workbook
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    from openpyxl.styles import Font, Alignment
                    from openpyxl.drawing.image import Image as XLImage

                    # Optional: needed to embed images into Excel
                    try:
                        from PIL import Image as PILImage

                        _HAS_PIL = True
                    except Exception:
                        _HAS_PIL = False


                    def _fig_to_png_bytes(fig):
                        """Convert a matplotlib figure to PNG bytes."""
                        if fig is None:
                            return None
                        buf = io.BytesIO()
                        fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
                        buf.seek(0)
                        return buf.getvalue()


                    def _add_df(ws, df, start_row, start_col=1):
                        """Write a dataframe into a worksheet starting at (start_row, start_col)."""
                        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                            for c_idx, val in enumerate(row, start=start_col):
                                ws.cell(row=r_idx, column=c_idx, value=val)


                    def _set_cols(ws, widths):
                        """Set excel column widths: widths = {"A": 28, "B": 22, ...}"""
                        for col, w in widths.items():
                            ws.column_dimensions[col].width = w


                    # -------- Gather artifacts already computed in your SHAP block --------
                    selected_range_txt = f"{int(selected_range_shap[0])}–{int(selected_range_shap[1])} cm⁻¹"

                    bees_fig = st.session_state.get("shap_bees_fig", None)
                    wf_fig = st.session_state.get("shap_wf_fig", None)

                    bees_png = _fig_to_png_bytes(bees_fig)
                    wf_png = _fig_to_png_bytes(wf_fig)

                    # Optional: Build a short “analysis result” text for Beeswarm+FI sheet
                    # Uses existing computed tables, no retraining.
                    try:
                        top3 = feature_importance_all.head(3)["Feature (Wavenumber)"].astype(str).tolist()
                        top3_txt = ", ".join(top3) if len(top3) else "—"
                    except Exception:
                        top3_txt = "—"

                    # -------- Create workbook with 3 sheets --------
                    wb = Workbook()
                    # Remove default sheet
                    wb.remove(wb.active)

                    # ============================================================
                    # Sheet 1: Beeswarm + Feature Importance (+ analysis text)
                    # ============================================================
                    ws1 = wb.create_sheet("Beeswarm+FI")

                    ws1["A1"] = "SHAP Results — Beeswarm + Feature Importance"
                    ws1["A1"].font = Font(bold=True, size=14)

                    ws1["A3"] = "Selected range"
                    ws1["A3"].font = Font(bold=True)
                    ws1["B3"] = selected_range_txt

                    ws1["A5"] = "📌 SHAP Beeswarm Plot (All Rows)"
                    ws1["A5"].font = Font(bold=True)

                    # Add Beeswarm image (if available)
                    if bees_png is not None and _HAS_PIL:
                        pil_img = PILImage.open(io.BytesIO(bees_png))
                        tmp = io.BytesIO()
                        pil_img.save(tmp, format="PNG")
                        tmp.seek(0)

                        xl_img = XLImage(tmp)
                        xl_img.width = 560
                        xl_img.height = 380
                        ws1.add_image(xl_img, "A6")
                    else:
                        ws1[
                            "A6"] = "Beeswarm image not available (click 'Generate Beeswarm Plot' first, then download)."
                        ws1["A6"].alignment = Alignment(wrap_text=True)

                    # Analysis text (non-technical friendly)
                    ws1["A26"] = "Analysis (quick interpretation)"
                    ws1["A26"].font = Font(bold=True)
                    ws1["A27"] = (
                        f"You selected {selected_range_txt}, so SHAP was computed only on those wavenumbers.\n"
                        f"Feature Importance ranks bands by average impact (Mean |SHAP|).\n"
                        f"Top global driver bands (by Mean |SHAP|): {top3_txt}.\n"
                        f"Beeswarm: each dot is one spectrum; position shows contribution; color shows absorbance (red=high, blue=low)."
                    )
                    ws1["A27"].alignment = Alignment(wrap_text=True)

                    # Feature Importance table
                    fi_start_row = 32
                    ws1[f"A{fi_start_row}"] = "📊 Feature Importance (Mean |SHAP|) – All Rows"
                    ws1[f"A{fi_start_row}"].font = Font(bold=True)

                    _add_df(ws1, feature_importance_all, start_row=fi_start_row + 1, start_col=1)

                    _set_cols(ws1, {"A": 28, "B": 22, "C": 44, "D": 60})

                    # ============================================================
                    # Sheet 2: Waterfall (Representative Sample)
                    # ============================================================
                    ws2 = wb.create_sheet("Waterfall")

                    ws2["A1"] = "SHAP Results — Waterfall (Representative Sample)"
                    ws2["A1"].font = Font(bold=True, size=14)

                    ws2["A3"] = "Selected range"
                    ws2["A3"].font = Font(bold=True)
                    ws2["B3"] = selected_range_txt

                    ws2["A5"] = "💧 SHAP Waterfall Plot (Representative Sample)"
                    ws2["A5"].font = Font(bold=True)

                    # Embed Waterfall image if available
                    if wf_png is not None and _HAS_PIL:
                        pil_img2 = PILImage.open(io.BytesIO(wf_png))
                        tmp2 = io.BytesIO()
                        pil_img2.save(tmp2, format="PNG")
                        tmp2.seek(0)

                        xl_img2 = XLImage(tmp2)
                        xl_img2.width = 560
                        xl_img2.height = 380
                        ws2.add_image(xl_img2, "A6")
                    else:
                        ws2[
                            "A6"] = "Waterfall image not available (click 'Generate Waterfall Plot' first, then download)."
                        ws2["A6"].alignment = Alignment(wrap_text=True)

                    ws2["A26"] = "How to read this plot"
                    ws2["A26"].font = Font(bold=True)
                    ws2["A27"] = (
                        "This explains ONE representative spectrum (one row).\n"
                        "Starts at baseline E[f(X)] (average model output).\n"
                        "Each bar shows how a specific band pushes the prediction up/down.\n"
                        "Final value equals the model output for this sample."
                    )
                    ws2["A27"].alignment = Alignment(wrap_text=True)

                    _set_cols(ws2, {"A": 28, "B": 22, "C": 44, "D": 60})

                    # ============================================================
                    # Sheet 3: Top SHAP features
                    # ============================================================
                    ws3 = wb.create_sheet("TopFeatures")

                    ws3["A1"] = "SHAP Results — Top Features"
                    ws3["A1"].font = Font(bold=True, size=14)

                    ws3["A3"] = "Selected range"
                    ws3["A3"].font = Font(bold=True)
                    ws3["B3"] = selected_range_txt

                    ws3["A5"] = "🎯 Top SHAP features"
                    ws3["A5"].font = Font(bold=True)

                    _add_df(ws3, top, start_row=6, start_col=1)

                    ws3["A20"] = "What this table means"
                    ws3["A20"].font = Font(bold=True)
                    ws3["A21"] = (
                        "These are the strongest global driver bands (highest Mean |SHAP|).\n"
                        "Mode/Assignment and Biochemical meaning map each band to bone FTIR chemistry "
                        "(mineral phosphate/carbonate, collagen/amide, water, NIR combinations)."
                    )
                    ws3["A21"].alignment = Alignment(wrap_text=True)

                    _set_cols(ws3, {"A": 28, "B": 22, "C": 44, "D": 60})

                    # -------- Save to bytes and provide download button --------
                    out = io.BytesIO()
                    wb.save(out)
                    out.seek(0)

                    st.download_button(
                        "📥 Download SHAP Results (Excel — 3 sheets)",
                        data=out.getvalue(),
                        file_name=f"shap_results_{int(selected_range_shap[0])}_{int(selected_range_shap[1])}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_shap_3sheets_{sig}"
                    )

                except Exception as e:
                    st.info(f"Could not create SHAP Excel download: {e}")

            #
            # #################  [19 Feb 2026 Script - START: 📋 Model Results with Detailed Metric ]  ###################################

            #################  [New Script - Start: 📋 Model Results with Detailed Metric ]  #####################################
            #######################################################################################################################

            if show_models:
                st.subheader("📋 Model Results with Detailed Metrics")

                # ---- imports + helpers (local to this block) ----
                import numpy as np, pandas as pd, matplotlib.pyplot as plt, seaborn as sns
                import io, zipfile, textwrap
                from sklearn.metrics import (
                    classification_report, confusion_matrix, accuracy_score,
                    precision_score, recall_score, f1_score
                )

                def _safe_prf(y_true, y_pred):
                    """
                    Precision/Recall/F1 that work for single-class, binary, and multiclass.
                    - single-class: returns 1.0 if predictions match perfectly, else 0.0
                    - binary: uses 'binary' average with consistent positive label
                    - multiclass: uses 'weighted' average
                    """
                    y_true = np.asarray(y_true)
                    y_pred = np.asarray(y_pred)
                    labels_present = np.unique(np.concatenate([y_true, y_pred]))
                    n = len(labels_present)

                    if n == 1:
                        same = bool(np.all(y_true == y_pred))
                        return {"precision": 1.0 if same else 0.0,
                                "recall": 1.0 if same else 0.0,
                                "f1": 1.0 if same else 0.0,
                                "avg": "single-class", "labels": labels_present}

                    if n == 2:
                        pos = labels_present[1]  # deterministic choice
                        return {
                            "precision": precision_score(y_true, y_pred, labels=labels_present, pos_label=pos,
                                                         average="binary", zero_division=0),
                            "recall": recall_score(y_true, y_pred, labels=labels_present, pos_label=pos,
                                                   average="binary", zero_division=0),
                            "f1": f1_score(y_true, y_pred, labels=labels_present, pos_label=pos,
                                           average="binary", zero_division=0),
                            "avg": "binary", "labels": labels_present
                        }

                    # 3+ classes
                    return {
                        "precision": precision_score(y_true, y_pred, labels=labels_present,
                                                     average="weighted", zero_division=0),
                        "recall": recall_score(y_true, y_pred, labels=labels_present,
                                               average="weighted", zero_division=0),
                        "f1": f1_score(y_true, y_pred, labels=labels_present,
                                       average="weighted", zero_division=0),
                        "avg": "weighted", "labels": labels_present
                    }

                # Try to find a LabelEncoder from earlier (quant/non-quant code paths)
                def _pick_encoder():
                    try:
                        return label_enc
                    except NameError:
                        pass
                    try:
                        return le_all
                    except NameError:
                        pass
                    try:
                        return label_enc_nq
                    except NameError:
                        pass
                    return None

                enc = _pick_encoder()

                # ✅ NEW: explanation helpers (no change to model logic)
                def _safe_get(dfi, row, col, default=np.nan):
                    try:
                        return float(dfi.loc[row, col])
                    except Exception:
                        return default

                def _write_model_explanation(model_name, report_df, cm_pct, class_names):
                    """
                    Prints a clear explanation under each model result (paper + non-technical friendly).
                    Uses:
                    - report_df from sklearn classification_report(output_dict=True)
                    - cm_pct row-normalized (%)
                    - class_names (decoded label names)
                    """
                    # Pull key summary stats safely
                    acc = _safe_get(report_df, "accuracy", "precision", np.nan)
                    w_f1 = _safe_get(report_df, "weighted avg", "f1-score", np.nan)
                    m_f1 = _safe_get(report_df, "macro avg", "f1-score", np.nan)

                    # Per-class correctness from diagonal
                    diag = np.diag(cm_pct) if cm_pct is not None and len(cm_pct) else []
                    per_class_correct = {}
                    for i in range(min(len(class_names), len(diag))):
                        per_class_correct[class_names[i]] = float(diag[i])

                    # Biggest confusion (largest off-diagonal %)
                    biggest_err = 0.0
                    biggest_pair = None
                    if cm_pct is not None and cm_pct.size > 0:
                        cm_off = cm_pct.copy()
                        np.fill_diagonal(cm_off, 0)
                        i_err, j_err = np.unravel_index(np.argmax(cm_off), cm_off.shape)
                        biggest_err = float(cm_off[i_err, j_err])
                        if i_err < len(class_names) and j_err < len(class_names):
                            biggest_pair = (class_names[i_err], class_names[j_err])

                    headline = []
                    if not np.isnan(acc):  headline.append(f"Accuracy: **{acc:.3f}**")
                    if not np.isnan(w_f1): headline.append(f"Weighted F1: **{w_f1:.3f}**")
                    if not np.isnan(m_f1): headline.append(f"Macro F1: **{m_f1:.3f}**")

                    if headline:
                        st.markdown(f"**{model_name}** → " + " | ".join(headline))

                    if per_class_correct:
                        st.markdown("**Correct classification rate by class (from confusion matrix diagonal):**")
                        for cls, pct in per_class_correct.items():
                            st.markdown(f"- **{cls}: {pct:.1f}%** of true {cls} samples were predicted correctly.")

                    if biggest_pair is not None:
                        st.markdown(
                            f"**Main confusion:** The most frequent error is **True {biggest_pair[0]} → Predicted {biggest_pair[1]} = {biggest_err:.1f}%**."
                        )

                # Create two columns and use only the first column for 80/20 split
                col1, col2 = st.columns([80, 20])
                with col1:
                    with st.expander("📊 Click to View The Selected Model Results", expanded=True):

                        st.markdown(
                            """
            **Classification Report**
            - **Precision:** When the model predicts a class (e.g., OIM), how often it is correct.
            - **Recall (Sensitivity):** Out of all true samples of a class, how many the model correctly identifies.
            - **F1-score:** Balance of precision and recall (high only if both are high).
            - **Support:** Number of test samples used for that class.
            
            **Confusion Matrix (%)**
            - **Rows = True class**, **Columns = Predicted class**
            - Values are **row-normalized** → each row sums to **100%**
            - **Diagonal cells** = correct classification rate for each true class
            - **Off-diagonal cells** = misclassification rate (confusions)
                            """
                        )

                        results = []
                        artifacts = []  # per-model bundle for ZIP

                        models_to_run = model_options if run_all else selected_models

                        for model_name in models_to_run:
                            # Select model based on user choice (unchanged)
                            if model_name == "SVM":
                                model = SVC()
                            elif model_name == "Logistic Regression":
                                model = LogisticRegression(max_iter=1000)
                            elif model_name == "KNN":
                                model = KNeighborsClassifier(n_neighbors=3)
                            elif model_name == "Random Forest":
                                model = RandomForestClassifier(n_estimators=100)
                            elif model_name == "XGBoost":
                                model = XGBClassifier(eval_metric="logloss")

                            # Split data into train/test sets (keep your params, guard tiny splits)
                            try:
                                X_train, X_test, y_train, y_test = train_test_split(
                                    X, y_encoded, test_size=0.2, stratify=y_encoded, random_state=42
                                )
                            except ValueError:
                                # Fallback if single-class or too few samples for stratify
                                X_train, X_test, y_train, y_test = X, X, y_encoded, y_encoded

                            # Fit model and make predictions
                            model.fit(X_train, y_train)
                            y_pred = model.predict(X_test)

                            # ---- Metrics (dynamic & safe) ----
                            acc = accuracy_score(y_test, y_pred)
                            prf = _safe_prf(y_test, y_pred)  # precision/recall/f1 + labels
                            precision = prf["precision"]
                            recall = prf["recall"]
                            f1 = prf["f1"]
                            labels_present = prf["labels"]

                            # Save results row
                            results.append({
                                "Model": model_name,
                                "Accuracy": acc,
                                "Precision": precision,
                                "Recall": recall,
                                "F1-Score": f1
                            })

                            # ---- Build readable class names (dynamic) ----
                            if np.issubdtype(labels_present.dtype, np.integer) and enc is not None:
                                try:
                                    class_names = [str(x) for x in enc.inverse_transform(labels_present)]
                                except Exception:
                                    class_names = [str(x) for x in labels_present]
                            else:
                                class_names = [str(x) for x in labels_present]

                            st.markdown("---")
                            st.markdown(f"### 🔎 Analysis for {model_name}")
                            report_col, matrix_col = st.columns([2, 3])

                            # Classification report
                            with report_col:
                                st.markdown("**Classification Report:**")
                                report_dict = classification_report(
                                    y_test, y_pred,
                                    labels=labels_present,
                                    target_names=class_names,
                                    output_dict=True,
                                    zero_division=0
                                )
                                report_df = pd.DataFrame(report_dict).transpose()

                                styled_report = report_df.style.set_table_styles([
                                    {'selector': 'th', 'props': [('border', '1px solid black'),
                                                                 ('padding', '8px'), ('background', '#f2f2f2')]},
                                    {'selector': 'td', 'props': [('border', '1px solid black'), ('padding', '8px')]},
                                    {'selector': 'tr', 'props': [('border', '1px solid black')]}
                                ]).format("{:.2f}")

                                st.write(styled_report)

                            # Confusion matrix (%)
                            with matrix_col:
                                cm = confusion_matrix(y_test, y_pred, labels=labels_present)
                                with np.errstate(divide="ignore", invalid="ignore"):
                                    cm_pct = cm.astype(float) / cm.sum(axis=1, keepdims=True) * 100
                                cm_pct = np.nan_to_num(cm_pct)

                                annot_labels = np.array([[f"{v:.1f}%" for v in row] for row in cm_pct])

                                fig_cm, ax = plt.subplots()
                                sns.heatmap(
                                    cm_pct,
                                    annot=annot_labels,
                                    fmt="",
                                    cmap="Blues",
                                    vmin=0, vmax=100,
                                    xticklabels=class_names,
                                    yticklabels=class_names,
                                    cbar_kws={"label": "% of true class"}
                                )
                                ax.set_xlabel("Predicted")
                                ax.set_ylabel("True")
                                ax.set_title(f"{model_name} - Confusion Matrix (%)")
                                st.pyplot(fig_cm)

                                # Save confusion matrix as PNG for ZIP
                                buf_png = io.BytesIO()
                                try:
                                    fig_cm.savefig(buf_png, format="png", dpi=200, bbox_inches="tight")
                                    cm_png_bytes = buf_png.getvalue()
                                except Exception:
                                    cm_png_bytes = b""
                                finally:
                                    buf_png.close()

                                plt.close(fig_cm)

                            # Explanation in UI (unchanged)
                            _write_model_explanation(model_name, report_df, cm_pct, class_names)

                            # Collect per-model artifacts for ZIP
                            cm_pct_df = pd.DataFrame(cm_pct, index=class_names, columns=class_names)

                            # Build a plain-text explanation to export (same content as UI)
                            try:
                                _acc = float(report_df.loc["accuracy", "precision"]) if "accuracy" in report_df.index else float(acc)
                                _w_f1 = float(report_df.loc["weighted avg", "f1-score"]) if "weighted avg" in report_df.index else float(f1)
                                _m_f1 = float(report_df.loc["macro avg", "f1-score"]) if "macro avg" in report_df.index else float(f1)

                                diag = np.diag(cm_pct) if cm_pct is not None and cm_pct.size else np.array([])
                                per_class_lines = []
                                for i in range(min(len(class_names), len(diag))):
                                    per_class_lines.append(
                                        f"{class_names[i]}: {diag[i]:.1f}% correctly classified (true → predicted same class)"
                                    )

                                biggest_err = 0.0
                                biggest_pair = None
                                if cm_pct is not None and cm_pct.size:
                                    cm_off = cm_pct.copy()
                                    np.fill_diagonal(cm_off, 0)
                                    i_err, j_err = np.unravel_index(np.argmax(cm_off), cm_off.shape)
                                    biggest_err = float(cm_off[i_err, j_err])
                                    if i_err < len(class_names) and j_err < len(class_names):
                                        biggest_pair = (class_names[i_err], class_names[j_err])

                                explanation_txt = f"""{model_name} → Accuracy: {_acc:.3f} | Weighted F1: {_w_f1:.3f} | Macro F1: {_m_f1:.3f}
            
            Correct classification rate by class (from confusion matrix diagonal):
            - {"; ".join(per_class_lines) if per_class_lines else "N/A"}
            
            Main confusion:
            - {f"True {biggest_pair[0]} → Predicted {biggest_pair[1]} = {biggest_err:.1f}%"
               if biggest_pair is not None else "N/A"}
            """
                            except Exception:
                                explanation_txt = f"{model_name} → Explanation not available for this run."

                            artifacts.append({
                                "model": model_name,
                                "report_df": report_df.copy(),
                                "cm_pct_df": cm_pct_df.copy(),
                                "cm_png": cm_png_bytes,
                                "explanation": explanation_txt
                            })

                        ################################################################
                        # Show metrics summary table
                        results_df = pd.DataFrame(results)

                        st.markdown("---")
                        st.markdown("#### 🏆 Best Model")

                        # Pick best by F1-score (common biomedical choice)
                        best_row = results_df.sort_values(by="F1-Score", ascending=False).iloc[0]
                        best_model = best_row["Model"]

                        st.markdown(
                            f"**Best model:** `{best_model}`  - with Accuracy: **{best_row['Accuracy'] * 100:.2f}%** "
                        )

                        st.markdown(
                            "**Why this model is best (for FTIR):** "
                            "It achieves the strongest balance between precision and recall (highest F1-score), "
                            "meaning it detects OIM vs WT reliably while keeping misclassifications low. "
                            "This is especially important in biomedical classification where both false positives and false negatives matter."
                        )

                        ################################################################
                        st.markdown("#### 📊 All Selected Models Performance Summary")
                        st.dataframe(results_df.style.background_gradient(cmap='YlGnBu'))

                        # ✅ Keep your original CSV download (unchanged)
                        metrics_csv = results_df.to_csv(index=False).encode('utf-8')
            #             st.download_button("📥 Download Model Metrics CSV", metrics_csv, "model_metrics.csv", "text/csv")
            #
            #             # ✅ NEW: Download FULL bundle (ZIP) with everything shown in UI
            #             try:
            #                 zip_buf = io.BytesIO()
            #                 with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            #
            #                     # Summary table
            #                     zf.writestr("model_metrics.csv", results_df.to_csv(index=False))
            #
            #                     # Best model
            #                     best_txt = f"""Best model: {best_model}
            # Accuracy: {best_row['Accuracy']:.4f}
            # Precision: {best_row['Precision']:.4f}
            # Recall: {best_row['Recall']:.4f}
            # F1-Score: {best_row['F1-Score']:.4f}
            # """
            #                     zf.writestr("BEST_MODEL.txt", best_txt)
            #
            #                     # Per-model files
            #                     for a in artifacts:
            #                         m = a["model"].replace(" ", "_")
            #                         zf.writestr(f"{m}/classification_report.csv", a["report_df"].to_csv())
            #                         zf.writestr(f"{m}/confusion_matrix_percent.csv", a["cm_pct_df"].to_csv())
            #                         zf.writestr(f"{m}/explanation.txt", a["explanation"])
            #                         if a["cm_png"]:
            #                             zf.writestr(f"{m}/confusion_matrix_percent.png", a["cm_png"])
            #
            #                     # README
            #                     zf.writestr(
            #                         "README.txt",
            #                         textwrap.dedent("""\
            #                         This ZIP contains the full 'Model Results with Detailed Metrics' export:
            #                         - model_metrics.csv: summary table for all executed models
            #                         - BEST_MODEL.txt: best model by F1-score
            #                         - <MODEL>/: per-model folder containing:
            #                             - classification_report.csv
            #                             - confusion_matrix_percent.csv (row-normalized, each row sums to 100%)
            #                             - confusion_matrix_percent.png (same heatmap shown in Streamlit)
            #                             - explanation.txt (human-readable summary)
            #                         """)
            #                     )
            #
            #                 zip_buf.seek(0)
            #                 st.download_button(
            #                     "📦 Download FULL Model Results (ZIP)",
            #                     data=zip_buf.getvalue(),
            #                     file_name="model_results_full_bundle.zip",
            #                     mime="application/zip",
            #                     key="dl_full_model_results_zip"
            #                 )
            #             except Exception as e:
            #                 st.info(f"Could not build ZIP download: {e}")
            #     # ✅ NEW: Download EVERYTHING in one Excel file (one sheet per model, with image + tables + explanation)
            #     # Place this block AFTER results_df / best_model are computed and AFTER `artifacts` is filled.
                #
                        try:
                            import io
                            from openpyxl import Workbook
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            from openpyxl.drawing.image import Image as XLImage
                            from openpyxl.styles import Font, Alignment
                            from PIL import Image as PILImage  # pillow is typically available; needed to embed png bytes

                            # Create workbook
                            wb = Workbook()

                            # -----------------------------
                            # 1) Summary sheet
                            # -----------------------------
                            ws_sum = wb.active
                            ws_sum.title = "Summary"

                            ws_sum["A1"] = "All Selected Models Performance Summary"
                            ws_sum["A1"].font = Font(bold=True, size=14)

                            # Write results_df
                            for r_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=True), start=3):
                                for c_idx, val in enumerate(row, start=1):
                                    ws_sum.cell(row=r_idx, column=c_idx, value=val)

                            # Best model info
                            ws_sum["A2"] = f"Best model (by F1-score): {best_model}"
                            ws_sum["A2"].font = Font(bold=True)

                            # Auto width (simple)
                            for col in ["A", "B", "C", "D", "E"]:
                                ws_sum.column_dimensions[col].width = 22

                            # -----------------------------
                            # 2) One sheet per model
                            # -----------------------------
                            for a in artifacts:
                                sheet_name = a["model"]
                                # Excel sheet name max length is 31 chars
                                if len(sheet_name) > 31:
                                    sheet_name = sheet_name[:31]

                                ws = wb.create_sheet(title=sheet_name)

                                # Title
                                ws["A1"] = f"🔎 Analysis for {a['model']}"
                                ws["A1"].font = Font(bold=True, size=14)

                                # Explanation text
                                ws["A3"] = "Explanation"
                                ws["A3"].font = Font(bold=True)
                                ws["A4"] = a["explanation"]
                                ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
                                ws.row_dimensions[4].height = 120
                                ws.column_dimensions["A"].width = 60

                                # Classification report table
                                ws["A6"] = "Classification Report"
                                ws["A6"].font = Font(bold=True)

                                start_row = 7
                                report_df = a["report_df"].copy()
                                report_df.insert(0, "row", report_df.index.astype(str))
                                report_df = report_df.reset_index(drop=True)

                                for r_idx, row in enumerate(dataframe_to_rows(report_df, index=False, header=True),
                                                            start=start_row):
                                    for c_idx, val in enumerate(row, start=1):
                                        ws.cell(row=r_idx, column=c_idx, value=val)

                                # Confusion matrix (%) values table
                                cm_start = start_row + len(report_df) + 3
                                ws.cell(row=cm_start, column=1, value="Confusion Matrix (%) Values").font = Font(bold=True)

                                cm_pct_df = a["cm_pct_df"].copy()
                                cm_pct_df.insert(0, "True\\Pred", cm_pct_df.index.astype(str))

                                for r_idx, row in enumerate(dataframe_to_rows(cm_pct_df, index=False, header=True),
                                                            start=cm_start + 1):
                                    for c_idx, val in enumerate(row, start=1):
                                        ws.cell(row=r_idx, column=c_idx, value=val)

                                # Embed confusion matrix image
                                img_anchor_row = cm_start
                                img_anchor_col = 10  # place image to the right (column J)

                                if a["cm_png"]:
                                    # Convert PNG bytes to a temporary in-memory image file for openpyxl
                                    pil_img = PILImage.open(io.BytesIO(a["cm_png"]))
                                    tmp = io.BytesIO()
                                    pil_img.save(tmp, format="PNG")
                                    tmp.seek(0)

                                    xl_img = XLImage(tmp)
                                    xl_img.width = 520  # adjust size if needed
                                    xl_img.height = 360

                                    # Anchor at J{cm_start}
                                    ws.add_image(xl_img, f"J{img_anchor_row}")

                                # Make some columns wider for tables
                                for col in ["B", "C", "D", "E", "F", "G"]:
                                    ws.column_dimensions[col].width = 18

                            # -----------------------------
                            # 3) Save to bytes + download
                            # -----------------------------
                            xlsx_buf = io.BytesIO()
                            wb.save(xlsx_buf)
                            xlsx_buf.seek(0)

                            st.download_button(
                                "📥 Download FULL Model Results (Excel)",
                                data=xlsx_buf.getvalue(),
                                file_name="Model_Results_with_Detailed_Metrics.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_full_model_results_excel"
                            )

                        except Exception as e:
                            st.info(f"Could not build Excel download: {e}")
            #################  [New Script - END: 📋 Model Results with Detailed Metric ]  #####################################
            #######################################################################################################################
            # #################  [19 Feb 2026 Script - END: 📋 Model Results with Detailed Metric ]  #####################################
            # #######################################################################################################################

            #
            # #################  [New Script - Start: 📋 Model Results with Detailed Metric ]  #####################################
            # #######################################################################################################################
            #
            # if show_models:
            #     st.subheader("📋 Model Results with Detailed Metrics")
            #
            #     # ---- imports + helpers (local to this block) ----
            #     import numpy as np, pandas as pd, matplotlib.pyplot as plt, seaborn as sns
            #     from sklearn.metrics import (
            #         classification_report, confusion_matrix, accuracy_score,
            #         precision_score, recall_score, f1_score
            #     )
            #
            #
            #     def _safe_prf(y_true, y_pred):
            #         """
            #         Precision/Recall/F1 that work for single-class, binary, and multiclass.
            #         - single-class: returns 1.0 if predictions match perfectly, else 0.0
            #         - binary: uses 'binary' average with consistent positive label
            #         - multiclass: uses 'weighted' average
            #         """
            #         y_true = np.asarray(y_true);
            #         y_pred = np.asarray(y_pred)
            #         labels_present = np.unique(np.concatenate([y_true, y_pred]))
            #         n = len(labels_present)
            #
            #         if n == 1:
            #             same = bool(np.all(y_true == y_pred))
            #             return {"precision": 1.0 if same else 0.0,
            #                     "recall": 1.0 if same else 0.0,
            #                     "f1": 1.0 if same else 0.0,
            #                     "avg": "single-class", "labels": labels_present}
            #
            #         if n == 2:
            #             pos = labels_present[1]  # deterministic choice
            #             return {
            #                 "precision": precision_score(y_true, y_pred, labels=labels_present, pos_label=pos,
            #                                              average="binary", zero_division=0),
            #                 "recall": recall_score(y_true, y_pred, labels=labels_present, pos_label=pos,
            #                                        average="binary", zero_division=0),
            #                 "f1": f1_score(y_true, y_pred, labels=labels_present, pos_label=pos,
            #                                average="binary", zero_division=0),
            #                 "avg": "binary", "labels": labels_present
            #             }
            #
            #         # 3+ classes
            #         return {
            #             "precision": precision_score(y_true, y_pred, labels=labels_present,
            #                                          average="weighted", zero_division=0),
            #             "recall": recall_score(y_true, y_pred, labels=labels_present,
            #                                    average="weighted", zero_division=0),
            #             "f1": f1_score(y_true, y_pred, labels=labels_present,
            #                            average="weighted", zero_division=0),
            #             "avg": "weighted", "labels": labels_present
            #         }
            #
            #
            #     # Try to find a LabelEncoder from earlier (quant/non-quant code paths)
            #     def _pick_encoder():
            #         try:
            #             return label_enc
            #         except NameError:
            #             pass
            #         try:
            #             return le_all
            #         except NameError:
            #             pass
            #         try:
            #             return label_enc_nq
            #         except NameError:
            #             pass
            #         return None
            #
            #     enc = _pick_encoder()
            #
            #     # Create two columns and use only the first column for 80/20 split
            #     col1, col2 = st.columns([80, 20])
            #     with col1:
            #         with st.expander("📊 Click to View The Selected Model Results", expanded=True):
            #
            #
            #             results = []
            #             models_to_run = model_options if run_all else selected_models
            #
            #             for model_name in models_to_run:
            #                 # Select model based on user choice (unchanged)
            #                 if model_name == "SVM":
            #                     model = SVC()
            #                 elif model_name == "Logistic Regression":
            #                     model = LogisticRegression(max_iter=1000)
            #                 elif model_name == "KNN":
            #                     model = KNeighborsClassifier(n_neighbors=3)
            #                 elif model_name == "Random Forest":
            #                     model = RandomForestClassifier(n_estimators=100)
            #                 elif model_name == "XGBoost":
            #                     model = XGBClassifier(eval_metric="logloss")
            #
            #
            #                 # Split data into train/test sets (keep your params, guard tiny splits)
            #                 try:
            #                     X_train, X_test, y_train, y_test = train_test_split(
            #                         X, y_encoded, test_size=0.2, stratify=y_encoded, random_state=42
            #                     )
            #                 except ValueError:
            #                     # Fallback if single-class or too few samples for stratify
            #                     X_train, X_test, y_train, y_test = X, X, y_encoded, y_encoded
            #
            #                 # Fit model and make predictions
            #                 model.fit(X_train, y_train)
            #                 y_pred = model.predict(X_test)
            #
            #                 # ---- Metrics (dynamic & safe) ----
            #                 acc = accuracy_score(y_test, y_pred)
            #                 prf = _safe_prf(y_test, y_pred)  # precision/recall/f1 + labels
            #                 precision = prf["precision"]
            #                 recall = prf["recall"]
            #                 f1 = prf["f1"]
            #                 labels_present = prf["labels"]
            #
            #                 # Save results row
            #                 results.append({
            #                     "Model": model_name,
            #                     "Accuracy": acc,
            #                     "Precision": precision,
            #                     "Recall": recall,
            #                     "F1-Score": f1 #,
            #                     # "Averaging": prf["avg"]
            #                 })
            #
            #                 # ---- Build readable class names (dynamic) ----
            #                 if np.issubdtype(labels_present.dtype, np.integer) and enc is not None:
            #                     try:
            #                         class_names = [str(x) for x in enc.inverse_transform(labels_present)]
            #                     except Exception:
            #                         class_names = [str(x) for x in labels_present]
            #                 else:
            #                     class_names = [str(x) for x in labels_present]
            #
            #                 # Display Classification Report & Confusion Matrix Side-by-Side
            #                 st.markdown(f"### 🔎 Analysis for {model_name}")
            #                 report_col, matrix_col = st.columns([2, 3])
            #
            #                 with report_col:
            #                     st.markdown("**Classification Report:**")
            #                     # Use the same labels order + names
            #                     report_dict = classification_report(
            #                         y_test, y_pred,
            #                         labels=labels_present,
            #                         target_names=class_names,
            #                         output_dict=True,
            #                         zero_division=0
            #                     )
            #                     report_df = pd.DataFrame(report_dict).transpose()
            #
            #                     styled_report = report_df.style.set_table_styles([
            #                         {'selector': 'th', 'props': [('border', '1px solid black'),
            #                                                      ('padding', '8px'), ('background', '#f2f2f2')]},
            #                         {'selector': 'td', 'props': [('border', '1px solid black'), ('padding', '8px')]},
            #                         {'selector': 'tr', 'props': [('border', '1px solid black')]}
            #                     ]).format("{:.2f}")
            #
            #                     st.write(styled_report)
            #
            #                 with matrix_col:
            #                     # Confusion Matrix (%), row-normalized
            #                     cm = confusion_matrix(y_test, y_pred, labels=labels_present)
            #                     with np.errstate(divide="ignore", invalid="ignore"):
            #                         cm_pct = cm.astype(float) / cm.sum(axis=1, keepdims=True) * 100
            #                     cm_pct = np.nan_to_num(cm_pct)
            #
            #                     annot_labels = np.array([[f"{v:.1f}%" for v in row] for row in cm_pct])
            #
            #                     fig_cm, ax = plt.subplots()
            #                     sns.heatmap(
            #                         cm_pct,
            #                         annot=annot_labels,
            #                         fmt="",
            #                         cmap="Blues",
            #                         vmin=0, vmax=100,
            #                         xticklabels=class_names,
            #                         yticklabels=class_names,
            #                         cbar_kws={"label": "% of true class"}
            #                     )
            #                     ax.set_xlabel("Predicted")
            #                     ax.set_ylabel("True")
            #                     ax.set_title(f"{model_name} - Confusion Matrix (%)")
            #                     st.pyplot(fig_cm)
            #
            #
            #             ################################################################
            #             # Show metrics summary table
            #             results_df = pd.DataFrame(results)
            #
            #             st.markdown("## 🏆 Best Model")
            #
            #             # Pick best by F1-score (common biomedical choice)
            #             best_row = results_df.sort_values(by="F1-Score", ascending=False).iloc[0]
            #             best_model = best_row["Model"]
            #
            #             st.markdown(
            #                 f"**Best model:** `{best_model}`  - with Accuracy: **{best_row['Accuracy'] * 100:.2f}%** "
            #                 # f"- Accuracy: **{best_row['Accuracy'] * 100:.2f}%**  \n"
            #                 # f"- Precision: **{best_row['Precision']:.2f}**  \n"
            #                 # f"- Recall: **{best_row['Recall']:.2f}**  \n"
            #                 # f"- F1-score: **{best_row['F1-Score']:.2f}**"
            #             )
            #
            #             st.markdown(
            #                 "**Why this model is best (for FTIR):** "
            #                 "It achieves the strongest balance between precision and recall (highest F1-score), "
            #                 "meaning it detects OIM vs WT reliably while keeping misclassifications low. "
            #                 "This is especially important in biomedical classification where both false positives and false negatives matter."
            #             )
            #
            #             ################################################################
            #             st.markdown("### 📊 Model Performance Summary")
            #             st.dataframe(results_df.style.background_gradient(cmap='YlGnBu'))
            #
            #             # Download metrics CSV
            #             metrics_csv = results_df.to_csv(index=False).encode('utf-8')
            #             st.download_button("📥 Download Model Metrics CSV", metrics_csv, "model_metrics.csv", "text/csv")
            #
            # #################  [New Script - END: 📋 Model Results with Detailed Metric ]  #####################################
            # #######################################################################################################################


            ##################################### [ PA Analysis ] ######################################
            if show_pca:
                # Create two columns
                col1, col2 = st.columns([4, 1])  # 4:1 width ratio (adjust as needed)

                with col1:
                    st.subheader("🔬 PCA: Principal Component Analysis")

                # =============================================================================================================

                with st.expander(
                        "PCA (Principal Component Analysis) reduces the high-dimensional FTIR spectral data into 2D and 3D projections to help visualize patterns and separability between classes like WT and OIM.",
                        expanded=False):

                    # ======================================= [ PCA Analysis ] ======================================================================

                    import numpy as np
                    import pandas as pd
                    import streamlit as st
                    from sklearn.decomposition import PCA
                    from sklearn.model_selection import StratifiedKFold, KFold
                    from scipy.signal import savgol_filter
                    from scipy.sparse import diags, eye
                    from scipy.sparse.linalg import spsolve

                    # ---------------- Preprocessing selection (combined) ----------------
                    PREPROCESSING_OPTIONS = ["Mean Center", "Autoscale", "Smoothing", "Normalization", "Baseline",
                                             "Second Derivative"]
                    x_preproc_pca = st.multiselect(
                        "Select Preprocessing Techniques (applied together, in order)",
                        options=PREPROCESSING_OPTIONS,
                        default=["Autoscale"]
                    )

                    # ---- Savitzky–Golay helpers ----
                    def _savgol_params(n_feat: int, window: int, poly: int, deriv: int):
                        w_max_odd = n_feat if (n_feat % 2 == 1) else (n_feat - 1)
                        if w_max_odd < 3:
                            return None
                        w = int(window)
                        if w % 2 == 0: w -= 1
                        if w < 3: w = 3
                        if w > w_max_odd: w = w_max_odd
                        p = int(poly)
                        p = max(p, deriv, 1)
                        if p >= w: p = w - 1
                        return w, p

                    def smooth_savgol(X, window=7, poly=3):
                        Xn = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                        Xn = Xn.copy()
                        params = _savgol_params(Xn.shape[1], window, poly, deriv=0)
                        if params is None: return Xn
                        w, p = params
                        return savgol_filter(Xn, window_length=w, polyorder=p, deriv=0, axis=1, mode="interp")


                    def second_derivative_savgol(X, window=7, poly=3):
                        Xn = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                        Xn = Xn.copy()
                        params = _savgol_params(Xn.shape[1], window, poly, deriv=2)
                        if params is None: return Xn
                        w, p = params
                        return savgol_filter(Xn, window_length=w, polyorder=p, deriv=2, axis=1, mode="interp")


                    # ---- Baseline (ALS, sparse + fast) ----
                    def als_baseline_sparse(y, lam=1e5, p=0.001, niter=10, eps=1e-12):
                        y = np.asarray(y, dtype=float).ravel()
                        L = y.size
                        if L < 3:
                            return np.zeros_like(y)
                        D = diags([1, -2, 1], [0, 1, 2], shape=(L - 2, L))
                        DTD = lam * (D.T @ D)
                        w = np.ones(L)
                        I = eye(L) * eps
                        for _ in range(niter):
                            W = diags(w, 0, shape=(L, L))
                            z = spsolve(W + DTD + I, w * y)
                            w = p * (y > z) + (1 - p) * (y < z)
                        return z

                    # ---- Fit stats for leakage-safe MC/Autoscale ----
                    def fit_center_scale_stats(X_arr, need_mean: bool, need_std: bool):
                        mu = sd = None
                        if need_mean or need_std:
                            mu = X_arr.mean(axis=0, keepdims=True)
                        if need_std:
                            sd = X_arr.std(axis=0, keepdims=True)
                            sd = sd.copy();
                            sd[sd == 0] = 1.0
                        return mu, sd

                    # ---- Apply the combined preprocessing chain (deterministic steps don't need "fit") ----
                    def apply_preprocessing_chain(
                            X,
                            ops, *,
                            smooth_window=7, smooth_poly=3,
                            normalization_mode="l2",
                            baseline_lam=100_000.0, baseline_p=0.001,
                            mean_=None, std_=None
                    ):
                        Xn = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                        Xn = Xn.copy()

                        # 1) Baseline
                        if "Baseline" in ops:
                            Xcorr = np.empty_like(Xn)
                            for i in range(Xn.shape[0]):
                                bl = als_baseline_sparse(Xn[i, :], lam=baseline_lam, p=baseline_p)
                                Xcorr[i, :] = Xn[i, :] - bl
                            Xn = Xcorr

                        # 2) Smoothing / Second Derivative
                        if "Smoothing" in ops and "Second Derivative" in ops:
                            # Prefer derivative if both selected
                            Xn = second_derivative_savgol(Xn, window=smooth_window, poly=smooth_poly)
                        elif "Smoothing" in ops:
                            Xn = smooth_savgol(Xn, window=smooth_window, poly=smooth_poly)
                        elif "Second Derivative" in ops:
                            Xn = second_derivative_savgol(Xn, window=smooth_window, poly=smooth_poly)

                        # 3) Mean Center / Autoscale (column-wise; leakage-safe if mean_/std_ provided)
                        if "Mean Center" in ops or "Autoscale" in ops:
                            mu = mean_ if mean_ is not None else Xn.mean(axis=0, keepdims=True)
                            Xn = Xn - mu
                        if "Autoscale" in ops:
                            sd = std_ if std_ is not None else Xn.std(axis=0, keepdims=True)
                            sd = sd.copy();
                            sd[sd == 0] = 1.0
                            Xn = Xn / sd

                        # 4) Normalization (row-wise)
                        if "Normalization" in ops:
                            mode = normalization_mode.lower()
                            if mode == "l2":
                                denom = np.linalg.norm(Xn, axis=1, keepdims=True)
                            elif mode == "l1":
                                denom = np.sum(np.abs(Xn), axis=1, keepdims=True)
                            else:  # "max"
                                denom = np.max(np.abs(Xn), axis=1, keepdims=True)
                            denom[denom == 0] = 1.0
                            Xn = Xn / denom

                        return Xn

                    # ---- Helper: PCA fit + reconstruction ----
                    def _pca_fit_reconstruct(X_mat, ncomp):
                        pca = PCA(n_components=ncomp, svd_solver="full", random_state=42)
                        T = pca.fit_transform(X_mat)
                        X_hat = pca.inverse_transform(T)
                        return pca, X_hat, T

                    # ========================= Score plots (2D/3D) & controls =========================
                    from sklearn.preprocessing import StandardScaler  # keep if you still need for separate visual plots

                    max_pc = max(2, min(10, X.shape[1]))

                    # (Optional) If you still want a separate plot space using plain StandardScaler:
                    scaler_all_plot = StandardScaler()
                    X_scaled_for_plot = scaler_all_plot.fit_transform(X)
                    ncomp_plot = st.slider("Number of PCs for plotting", 2, max_pc, value=min(5, max_pc), step=1,
                                           key="pca_nplot")
                    pca_plot = PCA(n_components=ncomp_plot, svd_solver="full", random_state=42)
                    T_all = pca_plot.fit_transform(X_scaled_for_plot)

                    # PCA evaluation controls
                    cv_splits_pca = st.slider("CV Splits (PCA)", 3, 15, value=10, step=1)
                    conf_limit_pca = st.selectbox("Confidence Limit (PCA)", [0.90, 0.95, 0.99], index=1)
                    pc_list = st.multiselect(
                        "Ncomp/LVs to evaluate (PCA)",
                        options=list(range(2, max_pc + 1)),
                        default=list(range(2, min(3, max_pc) + 1))
                    )

                    # ========================= CAL & CV with the combined preprocessing =========================
                    rows = []
                    model_id = 1

                    # Build a label for the chosen combination
                    preproc_label = ", ".join(x_preproc_pca) if x_preproc_pca else "None"

                    # ---- Calibration on ALL data using the full chain ----
                    X_all = apply_preprocessing_chain(
                        X, x_preproc_pca,
                        smooth_window=7, smooth_poly=3,
                        baseline_lam=100_000.0, baseline_p=0.001
                    )
                    sst_all = float((X_all ** 2).sum())

                    X_size = f"{X.shape[0]} x {X.shape[1]}"
                    Y_size = "—"

                    for ncomp in pc_list:
                        # CAL
                        pca_all, X_hat_all, _ = _pca_fit_reconstruct(X_all, ncomp)
                        resid_all = X_all - X_hat_all
                        sse_all = float((resid_all ** 2).sum())
                        rmsec = np.sqrt(sse_all / (X.shape[0] * X.shape[1]))
                        r2x_cal = 1.0 - (sse_all / sst_all) if sst_all > 0 else float("nan")
                        bias_cal = float(resid_all.mean())

                        # CV: split and re-apply the same chain with train-only stats for MC/Autoscale
                        if "y" in globals():
                            y_for_split = y
                        elif "y_encoded" in globals():
                            y_for_split = y_encoded
                        else:
                            y_for_split = None

                        if y_for_split is not None:
                            try:
                                folds = StratifiedKFold(n_splits=cv_splits_pca, shuffle=True, random_state=42).split(X,
                                                                                                                     y_for_split)
                            except Exception:
                                folds = KFold(n_splits=cv_splits_pca, shuffle=True, random_state=42).split(X)
                        else:
                            folds = KFold(n_splits=cv_splits_pca, shuffle=True, random_state=42).split(X)

                        sse_cv_total = 0.0
                        sst_cv_total = 0.0
                        bias_cv_accum = 0.0

                        for tr_idx, te_idx in folds:
                            X_tr = X.iloc[tr_idx].values.astype(float)
                            X_te = X.iloc[te_idx].values.astype(float)

                            # Fit stats for MC/Autoscale if selected
                            need_mean = ("Mean Center" in x_preproc_pca) or ("Autoscale" in x_preproc_pca)
                            need_std = ("Autoscale" in x_preproc_pca)
                            mu_tr, sd_tr = fit_center_scale_stats(X_tr, need_mean, need_std)

                            # Apply full chain to train/test with train stats
                            X_tr_s = apply_preprocessing_chain(
                                X_tr, x_preproc_pca,
                                smooth_window=7, smooth_poly=3,
                                baseline_lam=100_000.0, baseline_p=0.001,
                                mean_=mu_tr, std_=sd_tr
                            )
                            X_te_s = apply_preprocessing_chain(
                                X_te, x_preproc_pca,
                                smooth_window=7, smooth_poly=3,
                                baseline_lam=100_000.0, baseline_p=0.001,
                                mean_=mu_tr, std_=sd_tr
                            )

                            pca_cv = PCA(n_components=ncomp, svd_solver="full", random_state=42).fit(X_tr_s)
                            X_te_hat = pca_cv.inverse_transform(pca_cv.transform(X_te_s))

                            resid_te = X_te_s - X_te_hat
                            sse_cv_total += float((resid_te ** 2).sum())
                            sst_cv_total += float((X_te_s ** 2).sum())
                            bias_cv_accum += float(resid_te.mean())

                        rmsecv = np.sqrt(sse_cv_total / (X.shape[0] * X.shape[1]))
                        r2x_cv = 1.0 - (sse_cv_total / sst_cv_total) if sst_cv_total > 0 else float("nan")
                        bias_cv = bias_cv_accum / cv_splits_pca
                        rmse_ratio = rmsecv / rmsec if rmsec > 0 else float("nan")

                        rows.append({
                            "CV Splits": cv_splits_pca,
                            "Confidence Limit": conf_limit_pca,
                            "RMSE Ratio (RMSECV/RMSEC)": rmse_ratio,
                            "RMSEC (Cal)": rmsec,
                            "RMSECV (CV)": rmsecv,
                            "Model Name": f"Model {model_id}",
                            "Ncomp/LVs": ncomp,
                            "X-Preprocessing": preproc_label,
                            "X Include Size": X_size,
                            "Y Include Size": Y_size,
                            "R2C (Cal)": r2x_cal,  # R²X on calibration
                            "R2CV (CV)": r2x_cv,  # R²X on CV
                            "Bias": bias_cal,
                            "Bias (CV)": bias_cv
                        })
                        model_id += 1

                    # ---------------- Table + download ----------------
                    pca_calcv_df = pd.DataFrame(rows, columns=[
                        "CV Splits", "Confidence Limit", "RMSE Ratio (RMSECV/RMSEC)",
                        "RMSEC (Cal)", "RMSECV (CV)", "Model Name", "Ncomp/LVs",
                        "X-Preprocessing", "X Include Size", "Y Include Size",
                        "R2C (Cal)", "R2CV (CV)", "Bias", "Bias (CV)"
                    ])

                    fmt_cols = ["RMSE Ratio (RMSECV/RMSEC)", "RMSEC (Cal)", "RMSECV (CV)", "R2C (Cal)", "R2CV (CV)",
                                "Bias", "Bias (CV)"]

                    st.markdown("### 📋 PCA Analysis Result Summary")
                    st.dataframe(
                        pca_calcv_df.style
                        .format({c: "{:.4f}" for c in fmt_cols})
                        .background_gradient(subset=["R2CV (CV)", "R2C (Cal)"], cmap="YlGn")
                        .background_gradient(subset=["RMSECV (CV)", "RMSEC (Cal)"], cmap="YlOrRd_r")

                    )

                    st.download_button(
                        "📥 Download PCA Summary (CSV)",
                        data=pca_calcv_df.to_csv(index=False).encode("utf-8"),
                        file_name="pca_summary_table.csv",
                        mime="text/csv"
                    )

                    ########################################################################################################

                    # ---- Controls matching PLS/PLS-DA style ----
                    from sklearn.decomposition import PCA
                    from sklearn.model_selection import StratifiedKFold
                    from sklearn.preprocessing import StandardScaler

                    max_pc = max(2, min(10, X.shape[1]))

                    # Pre-scale once for calibration pass
                    scaler_all = StandardScaler()
                    X_scaled_all = scaler_all.fit_transform(X)
                    sst_all = float((X_scaled_all ** 2).sum())

                    # ========================= Score plots (2D/3D) =========================
                    # (Removed UI sliders/selects; use defaults)
                    ncomp_plot = min(5, max_pc)
                    pca_plot = PCA(n_components=ncomp_plot)
                    T_all = pca_plot.fit_transform(X_scaled_all)

                    # default confidence limit for ellipses (since UI removed)
                    conf_limit_pca = 0.95

                    # Fit PCA
                    pca = PCA()
                    X_pca_all = pca.fit_transform(X)
                    explained_var = pca.explained_variance_ratio_
                    eigenvalues = pca.explained_variance_

                    # Create PCA dataframe
                    df_pca_all = pd.DataFrame(X_pca_all, columns=[f"PC{i + 1}" for i in range(X_pca_all.shape[1])])
                    df_pca_all['Label'] = y.values

                    # Build plotting DF (include all PCs to pick axes)
                    pc_cols = [f"PC{i + 1}" for i in range(ncomp_plot)]
                    df_pca_all = pd.DataFrame(T_all, columns=pc_cols)
                    df_pca_all["Label"] = y.values

                    # Confidence ellipses
                    chi2_map = {0.90: 4.605, 0.95: 5.991, 0.99: 9.210}
                    chi2_val = chi2_map[conf_limit_pca]


                    def _ellipse_xy(points_2d, chi2_q, npts=200):
                        pts = np.asarray(points_2d)
                        if pts.shape[0] < 3:
                            return None, None
                        mu = pts.mean(axis=0)
                        S = np.cov(pts.T)
                        vals, vecs = np.linalg.eigh(S)
                        order = np.argsort(vals)[::-1]
                        vals, vecs = vals[order], vecs[:, order]
                        radii = np.sqrt(np.maximum(vals, 1e-12) * chi2_q)
                        t = np.linspace(0, 2 * np.pi, npts)
                        circ = np.stack([np.cos(t) * radii[0], np.sin(t) * radii[1]], axis=0)
                        ell = (vecs @ circ).T + mu
                        return ell[:, 0], ell[:, 1]


                    # Axis selectors
                    col_ax1, col_ax2, col_ax3 = st.columns(3)
                    with col_ax1:
                        pc_x = st.selectbox("X-axis (scores)", pc_cols, index=0, key="pca_x")
                    with col_ax2:
                        pc_y = st.selectbox("Y-axis (scores)", pc_cols, index=1, key="pca_y")
                    with col_ax3:
                        pc_z = st.selectbox("Z-axis (scores)", pc_cols, index=min(2, len(pc_cols) - 1), key="pca_z")

                    x_idx = int(pc_x.replace("PC", "")) - 1
                    y_idx = int(pc_y.replace("PC", "")) - 1
                    z_idx = int(pc_z.replace("PC", "")) - 1

                    # 2D plot
                    st.subheader("📉 PCA 2D Projection")
                    fig_2d = px.scatter(df_pca_all, x=pc_x, y=pc_y, color="Label",
                                        title=f"PCA Projection ({pc_x} vs {pc_y})")
                    for lab in df_pca_all["Label"].unique():
                        d = df_pca_all[df_pca_all["Label"] == lab][[pc_x, pc_y]].values
                        ex, ey = _ellipse_xy(d, chi2_val)
                        if ex is not None:
                            fig_2d.add_trace(
                                go.Scatter(x=ex, y=ey, mode="lines",
                                           name=f"{lab} {int(conf_limit_pca * 100)}% ellipse",
                                           line=dict(width=1), showlegend=False)
                            )
                    st.plotly_chart(fig_2d)

                    # 3D plot
                    st.subheader("🧭 PCA 3D Visualization")
                    if len(pc_cols) >= 3:
                        fig_3d = px.scatter_3d(df_pca_all, x=pc_x, y=pc_y, z=pc_z, color="Label",
                                               title=f"3D PCA: {pc_x} vs {pc_y} vs {pc_z}",
                                               height=520, width=800)
                        st.plotly_chart(fig_3d)
                    else:
                        st.info("Add ≥3 PCs to view the 3D projection.")

                    pca_summary = pd.DataFrame({
                        "PC": [f"PC{i + 1}" for i in range(len(explained_var))],
                        "Eigenvalue of Cov(X)": eigenvalues,
                        "% Variance This PC": explained_var * 100,
                        "% Variance Cumulative": np.cumsum(explained_var) * 100
                    })

                    ################################# display the pca table ######################
                    st.dataframe(pca_summary, height=350)

                    ##### PCA Loadings with Bond Annotation #####

                    st.subheader("📊 PCA Loadings")
                    # Compute PCA Loadings
                    # loadings = pd.DataFrame(pca.components_.T, index=selected_columns,
                    #                         columns=[f"PC{i + 1}" for i in range(pca.n_components_)])

                    # --- NEW: get the exact feature names PCA was fit on ---
                    feature_cols = list(X.columns)  # must match n_features of pca.components_
                    selected_columns = feature_cols  # keep your later code that uses selected_columns

                    # Safety check (optional but helpful during debugging)
                    assert pca.components_.shape[1] == len(feature_cols), \
                        f"Mismatch: PCA features={pca.components_.shape[1]}, labels={len(feature_cols)}"

                    # --- Build loadings with the correct index ---
                    loadings = pd.DataFrame(
                        pca.components_.T,
                        index=feature_cols,
                        columns=[f"PC{i + 1}" for i in range(pca.components_.shape[0])]
                    )

                    # User Selects PC to View Loadings
                    selected_pc_loading = st.selectbox("Select PC to view Loadings:", loadings.columns)

                    st.markdown(f"### 🔑 Loadings Plot for {selected_pc_loading}")

                    ###========================== line plot of loading
                    #
                    # fig_load1 = go.Figure()
                    # fig_load1.add_trace(go.Scatter(
                    #     x=selected_columns,
                    #     y=loadings[selected_pc_loading],
                    #     mode="lines",
                    #     line=dict(color="blue"),
                    #     name=f"{selected_pc_loading} Loadings"
                    # ))
                    # fig_load1.add_hline(y=0, line=dict(color="gray", dash="dash"))  # Zero reference line

                    # ===== Line plot of loadings (use feature_cols on x) =====
                    fig_load1 = go.Figure()
                    fig_load1.add_trace(go.Scatter(
                        x=feature_cols,
                        y=loadings[selected_pc_loading],
                        mode="lines",
                        line=dict(color="blue"),
                        name=f"{selected_pc_loading} Loadings"
                    ))
                    fig_load1.add_hline(y=0, line=dict(color="gray", dash="dash"))  # Zero reference line



                    # ✅ Add Band Annotations
                    band_annotations1 = {
                        1740: "C=O Stretch (Lipid)",
                        1650: "Amide I (Protein)",
                        1550: "Amide II (Protein)",
                        1338: "CH₂ (Collagen)",
                        1200: "C–O Stretch (Matrix)",
                        1115: "HPO₄²⁻ (Mineral)",
                        1060: "Sugar Ring C–O",
                        1030: "PO₄³⁻ (Mineral)",
                        875: "CO₃²⁻ Bending",
                        856: "C–S Proteoglycan",
                        7000: "O–H Stretch",
                        6688: "N–H Stretch",
                        5800: "CH₂ Lipid",
                        5200: "O–H Water",
                    }

                    for wn, label in band_annotations1.items():
                        if selected_range[0] <= wn <= selected_range[1]:
                            fig_load1.add_vline(x=wn, line=dict(color="gray", width=1, dash="dash"))
                            fig_load1.add_annotation(x=wn, y=loadings[selected_pc_loading].max(), text=label,
                                                     showarrow=True, arrowhead=1, yshift=10, font=dict(size=9))

                    fig_load1.update_layout(
                        title=f"PCA Loadings for {selected_pc_loading} ({explained_var[int(selected_pc_loading[-1]) - 1] * 100:.2f}% Variance)",
                        xaxis_title="Wavenumber (cm⁻¹)",
                        yaxis_title="Loading Score",
                        height=400
                    )
                    st.plotly_chart(fig_load1)

                    ###========================== line plot of loading

                    fig_load = go.Figure()
                    fig_load.add_trace(go.Scatter(
                        x=selected_columns,
                        y=loadings[selected_pc_loading],
                        mode="lines",
                        line=dict(color="blue"),
                        name=f"{selected_pc_loading} Loadings"
                    ))
                    fig_load.add_hline(y=0, line=dict(color="gray", dash="dash"))  # Zero reference line

                    # ✅ Bond Annotation Mapping Table
                    bond_annotations_table = [
                        (8500, "O–H Stretching and Bending (Water)"),
                        (7000, "O–H Stretching (Water)"),
                        (6688, "N–H Stretching (Protein/Collagen)"),
                        (5800, "CH₂ Stretching (Lipid)"),
                        (5200, "O–H Stretching and Bending (Water)"),
                        (4890, "N–H Bending (Protein/Collagen)"),
                        (4610, "C–H Stretching & Deformation (Protein/Collagen)"),
                        (4310, "Sugar Ring Vibrations (Proteoglycan)"),
                        ((3600, 3200), "O–H Stretching (Water/Hydroxyl)"),
                        ((3500, 3300), "N–H Stretching (Proteins)"),
                        ((3000, 2800), "C–H Stretching (Lipids, CH₂)"),
                        ((1750, 1650), "C=O Stretching (Proteins/Lipids)"),
                        (1550, "Amide II (Proteins)"),
                        (1338, "CH₂ Side Chain Bending (Collagen)"),
                        ((1100, 900), "PO₄³⁻ Stretching (Bone Mineral)"),
                        ((890, 850), "CO₃²⁻ Bending (Carbonate)"),
                        (1740, "C=O Stretching (Ester, Lipids)"),
                        (1650, "Amide I (Proteins)"),
                        (1630, "Water O–H Bending (Water)"),
                        ((1200, 1000), "C–O Stretching (Alcohols/Ethers)"),
                        (1115, "HPO₄²⁻ Stretching (Bone mineral)"),
                        (1060, "Sugar Ring C–O Stretch (Carbohydrates)"),
                        (1030, "PO₄³⁻ Stretching (Bone Mineral)"),
                        (875, "CO₃²⁻ Bending (Carbonates)"),
                        (856, "C–S Bending (Proteoglycans)"),
                        ((900, 800), "Aromatic C–H Bending (Fingerprint region)")
                    ]


                    # ✅ Function to Assign Bond Annotation
                    def get_bond_annotation(wavenumber):
                        # Check ranges first
                        for entry in bond_annotations_table:
                            if isinstance(entry[0], tuple):  # Range
                                low, high = entry[0]
                                if high <= wavenumber <= low:
                                    return entry[1]
                        # Exact match
                        for entry in bond_annotations_table:
                            if not isinstance(entry[0], tuple) and entry[0] == int(wavenumber):
                                return entry[1]
                        # Nearest match if no exact or range match
                        all_wavenumbers = [
                            val if isinstance(val, int) else sum(val) // 2 for val, _ in bond_annotations_table
                        ]
                        closest_idx = np.argmin([abs(wavenumber - wn) for wn in all_wavenumbers])
                        return bond_annotations_table[closest_idx][1]


                    ########################################################################################################################

                    # ✅ Variable Importance with Bond Annotations in One Table
                    variable_importance = loadings.abs()[selected_pc_loading].sort_values(ascending=False)
                    importance_df = pd.DataFrame(variable_importance.head(20)).reset_index()
                    importance_df.columns = ["Wavenumber (cm⁻¹)", "Importance"]


                    # ✅ Safely extract numeric value (handle cases like '1030/1240')
                    def extract_first_numeric(value):
                        value_str = str(value).strip()
                        if "/" in value_str:  # Take first part if it's a ratio
                            value_str = value_str.split("/")[0].strip()
                        try:
                            return float(value_str)
                        except ValueError:
                            return None  # In case there's any unexpected text


                    importance_df["Numeric Wavenumber"] = importance_df["Wavenumber (cm⁻¹)"].apply(
                        extract_first_numeric)

                    # ✅ Drop rows where numeric conversion failed (if any)
                    importance_df = importance_df.dropna(subset=["Numeric Wavenumber"])

                    # ✅ Assign bond annotations
                    importance_df["Band Annotation"] = importance_df["Numeric Wavenumber"].apply(
                        get_bond_annotation)

                    # ======================= PCA Loadings — Dynamic Insight & Reasoning (ADD-ON v2) =======================
                    import re
                    import numpy as np
                    import pandas as pd
                    import streamlit as st


                    # ---------- 0) Helpers ----------
                    def _pc_index(name):
                        m = re.search(r"\d+", str(name))
                        return max(int(m.group()) - 1, 0) if m else 0


                    def _num(val):
                        try:
                            return float(str(val).split("/")[0])
                        except Exception:
                            return np.nan


                    def _smooth(y, win=7):
                        # simple moving average, odd window
                        win = int(win)
                        if win < 1: return y
                        if win % 2 == 0: win += 1
                        if win == 1: return y
                        k = np.ones(win, dtype=float) / win
                        pad = win // 2
                        ypad = np.pad(np.asarray(y, dtype=float), (pad, pad), mode="edge")
                        return np.convolve(ypad, k, mode="valid")


                    def _local_extrema(y, order=3, mode="max"):
                        # naive local peaks (mode 'max' for peaks, 'min' for dips)
                        y = np.asarray(y, dtype=float)
                        idxs = []
                        n = len(y)
                        for i in range(order, n - order):
                            seg = y[i - order:i + order + 1]
                            if mode == "max":
                                if y[i] == seg.max() and (y[i] > seg[:order].max()) and (
                                        y[i] > seg[order + 1:].max()):
                                    idxs.append(i)
                            else:  # min
                                if y[i] == seg.min() and (y[i] < seg[:order].min()) and (
                                        y[i] < seg[order + 1:].min()):
                                    idxs.append(i)
                        return idxs


                    def _favor_class(value, class_pos, class_neg):
                        if class_pos is None or class_neg is None:
                            return ""
                        return f"favors **{class_pos}**" if value > 0 else f"favors **{class_neg}**"


                    # ---------- 1) Selected PC, variance, scores ----------
                    pc_idx = _pc_index(selected_pc_loading)
                    try:
                        var_pct = float(explained_var[pc_idx]) * 100.0
                    except Exception:
                        var_pct = np.nan

                    try:
                        pc_scores = X_pca_all[:, pc_idx]
                    except Exception:
                        pc_scores = pca.transform(X)[:, pc_idx]

                    # ---------- 2) Determine which class is on + / − side ----------
                    try:
                        df_scores = pd.DataFrame({"Label": np.asarray(y), "PC_Score": pc_scores})
                        class_means = df_scores.groupby("Label", observed=True)["PC_Score"].mean().sort_values(
                            ascending=False)
                        class_pos = class_means.index[0] if len(class_means) else None
                        class_neg = class_means.index[-1] if len(class_means) else None
                    except Exception:
                        class_pos = class_neg = None

                    # ---------- 3) Create numeric wavenumber axis & focus range ----------
                    try:
                        wn_labels = list(loadings.index)
                    except Exception:
                        wn_labels = list(selected_columns)  # fallback
                    wn_numeric = np.array([_num(v) for v in wn_labels], dtype=float)

                    try:
                        lo, hi = float(selected_range[0]), float(selected_range[1])
                        focus_mask = (wn_numeric >= min(lo, hi)) & (wn_numeric <= max(lo, hi))
                        if not np.any(focus_mask):
                            focus_mask = np.isfinite(wn_numeric)
                    except Exception:
                        focus_mask = np.isfinite(wn_numeric)

                    s_full = pd.to_numeric(loadings[selected_pc_loading], errors="coerce").values
                    s_focus = s_full[focus_mask]
                    wn_focus = wn_numeric[focus_mask]
                    lbl_focus = np.array(wn_labels, dtype=object)[focus_mask]

                    if len(s_focus) < 5 or np.all(~np.isfinite(s_focus)):
                        st.warning("Not enough points in the current range to analyze loadings.")
                        st.stop()

                    # ---------- 4) Controls for detection (defaults) ----------
                    top_k = min(12, len(s_focus))
                    smooth_win = 7
                    half_width_cm = 25

                    y_sm = _smooth(s_focus, win=smooth_win)

                    # ---------- 5) Peak & dip detection ----------
                    order = max(2, smooth_win // 2)
                    pos_idx = _local_extrema(y_sm, order=order, mode="max")
                    neg_idx = _local_extrema(y_sm, order=order, mode="min")

                    cand = [(i, y_sm[i]) for i in pos_idx] + [(i, y_sm[i]) for i in neg_idx]
                    if not cand:
                        cand = list(enumerate(y_sm))
                    cand_sorted = sorted(cand, key=lambda t: abs(t[1]), reverse=True)[:top_k]

                    # ---------- 6) Build "Top drivers" table ----------
                    rows = []
                    for i, amp in cand_sorted:
                        wn = wn_focus[i]
                        label = lbl_focus[i]
                        try:
                            band = get_bond_annotation(float(wn)) if np.isfinite(wn) else ""
                        except Exception:
                            band = ""
                        rows.append({
                            "Wavenumber (cm⁻¹)": str(label),
                            "Numeric (cm⁻¹)": wn if np.isfinite(wn) else np.nan,
                            "Loading": float(s_focus[i]),
                            "|Loading|": float(abs(s_focus[i])),
                            "Band Annotation": band,
                            "Interpretation": _favor_class(s_focus[i], class_pos, class_neg)
                        })

                    drivers_df = (pd.DataFrame(rows)
                                  .sort_values("|Loading|", ascending=False)
                                  .reset_index(drop=True))

                    # ===== 📊 Variable Importance with Bond Annotations + Download (one line) =====
                    col1, col2 = st.columns([4, 1])

                    with col1:
                        st.markdown("### 📊 Variable Importance with Bond Annotations")

                    with col2:
                        var_importance_csv = importance_df.to_csv(index=False).encode("utf-8")
                        st.download_button(
                            "📥 Download CSV",
                            data=var_importance_csv,
                            file_name="variable_importance_with_bands.csv",
                            mime="text/csv"

                        )

                    # ===== Display table =====
                    st.dataframe(importance_df)

                    ########################## Most influenced wavenumbers ###############################
                    col1, col2 = st.columns([4, 1])
                    with col1:
                        st.markdown("### ⭐ Most influential wavenumbers (by |loading|)")
                    with col2:
                        st.download_button(
                            "📥 Download CSV",
                            data=drivers_df[
                                ["Wavenumber (cm⁻¹)", "Loading", "|Loading|", "Band Annotation", "Interpretation"]
                            ].to_csv(index=False).encode("utf-8"),
                            file_name=f"most_influential_wavenumbers_{selected_pc_loading}.csv",
                            mime="text/csv"

                        )

                    st.dataframe(
                        drivers_df[
                            ["Wavenumber (cm⁻¹)", "Loading", "|Loading|", "Band Annotation", "Interpretation"]
                        ].style.format({"Loading": "{:.4f}", "|Loading|": "{:.4f}"})

                    )

                    ################################################# [  Auto-detected regions ] ###############################################################################

                    # ---------- 7) Auto-detected regions around each peak ----------
                    def _region_for(center_cm1):
                        lo = center_cm1 - half_width_cm
                        hi = center_cm1 + half_width_cm
                        return lo, hi


                    regs = []
                    for _, r in drivers_df.dropna(subset=["Numeric (cm⁻¹)"]).iterrows():
                        c = float(r["Numeric (cm⁻¹)"])
                        lo, hi = _region_for(c)
                        regs.append([lo, hi, r["Wavenumber (cm⁻¹)"], r["Band Annotation"]])

                    regs.sort(key=lambda x: x[0])
                    merged = []
                    for lo, hi, lbl, ann in regs:
                        if not merged or lo > merged[-1][1]:
                            merged.append([lo, hi, [lbl], [ann]])
                        else:
                            merged[-1][1] = max(merged[-1][1], hi)
                            merged[-1][2].append(lbl)
                            merged[-1][3].append(ann)

                    reg_rows = []
                    for lo, hi, lbls, anns in merged:
                        m = (wn_focus >= lo) & (wn_focus <= hi)
                        if not np.any(m):
                            continue
                        net = float(np.nansum(s_focus[m]))
                        mass = float(np.nansum(np.abs(s_focus[m])))
                        lbl_text = ", ".join(map(str, lbls[:3])) + ("..." if len(lbls) > 3 else "")
                        ann_text = ", ".join(sorted(set(a for a in anns if a)))
                        reg_rows.append({
                            "Region (cm⁻¹)": f"{int(lo)}–{int(hi)}",
                            "|Σ loading|": mass,
                            "Net sign": "Positive" if net > 0 else ("Negative" if net < 0 else "≈0"),
                            "Peak labels": lbl_text,
                            "Likely bands": ann_text or "—",
                        })

                    contrib_df = (pd.DataFrame(reg_rows)
                                  .sort_values("|Σ loading|", ascending=False)
                                  .reset_index(drop=True))

                    ###########################################################################################################################################
                    # --- Ensure lv_idx / var_pct / class_pos / class_neg available for the summary ---
                    try:
                        lv_idx
                    except NameError:
                        import re

                        m = re.search(r"\d+", str(selected_pc_loading))
                        lv_idx = max(int(m.group()) - 1, 0) if m else 0

                    try:
                        var_pct
                    except NameError:
                        # explained_x should be a normalized ratio array for PLS LVs (sum≈1). If missing, compute from scores.
                        explained_x_attr = getattr(pls, "x_explained_variance_ratio_", None)
                        if explained_x_attr is not None:
                            explained_x = np.asarray(explained_x_attr, dtype=float)
                        else:
                            T_all = getattr(pls, "x_scores_", None)
                            if T_all is not None and np.sum(np.var(T_all, axis=0)) > 0:
                                raw_var = np.var(T_all, axis=0)
                                explained_x = raw_var / np.sum(raw_var)
                            else:
                                explained_x = np.ones(pls.x_loadings_.shape[1], dtype=float) / pls.x_loadings_.shape[1]
                        var_pct = float(explained_x[min(lv_idx, len(explained_x) - 1)] * 100.0)

                    # If class context wasn’t computed earlier, build it from PLS scores
                    try:
                        class_pos, class_neg
                    except NameError:
                        try:
                            lv_scores = pls.x_scores_[:, lv_idx]
                            y_labels = y.values if hasattr(y, "values") else np.asarray(y)
                            _df_tmp = pd.DataFrame({"Label": y_labels, "LV_Score": lv_scores})
                            class_means = _df_tmp.groupby("Label", observed=True)["LV_Score"].mean().sort_values(
                                ascending=False)
                            class_pos = class_means.index[0] if len(class_means) else None
                            class_neg = class_means.index[-1] if len(class_means) else None
                        except Exception:
                            class_pos = class_neg = None

                    # ---------- Narrative summary (same structure as PCA, PLS wording/vars) ----------
                    pos_mass = float(np.nansum(np.abs(s_focus[s_focus > 0])))
                    neg_mass = float(np.nansum(np.abs(s_focus[s_focus < 0])))
                    contrast = ("Positive-weighted regions dominate"
                                if pos_mass > neg_mass else
                                (
                                    "Negative-weighted regions dominate" if neg_mass > pos_mass else "Balanced positive/negative weights"))

                    summary = []
                    summary.append(f"- **{selected_pc_loading}** explains **{var_pct:.2f}%** of total variance.")
                    summary.append(f"- **Contrast:** {contrast} in the current range.")
                    if class_pos and class_neg:
                        summary.append(
                            f"- **Class context:** Higher LV scores (positive side) align with **{class_pos}**; "
                            f"lower scores (negative side) align with **{class_neg}**."
                        )
                    if len(contrib_df):
                        top_reg = contrib_df.iloc[0]
                        summary.append(
                            f"- **Strongest region:** {top_reg['Region (cm⁻¹)']} "
                            f"({top_reg['Net sign']}, |Σloading|={top_reg['|Σ loading|']:.4f}); "
                            f"likely bands: {top_reg['Likely bands'] or '—'}."
                        )

                    head_n = min(6, len(drivers_df))
                    bullets = []
                    for _, r in drivers_df.head(head_n).iterrows():
                        side = "positive" if r["Loading"] > 0 else "negative"
                        interp = f" — {r['Interpretation']}" if r["Interpretation"] else ""
                        bullets.append(
                            f"  • {r['Wavenumber (cm⁻¹)']}: **{side}** loading ({r['Band Annotation']}){interp}"
                        )

                    st.markdown("### 📝 Auto-summary")
                    st.markdown("\n".join(summary + ["- **Key drivers:**"] + bullets))

                    st.caption(
                        "Notes: (1) PLS LV axis sign is arbitrary; interpretation uses relative +/− sides and class means. "
                        "(2) Smoothing only aids peak detection; it doesn’t change the underlying loadings. "
                        "(3) Preprocessing (e.g., autoscale) affects which bands dominate."
                    )

                    # (Optional correlation block can be added below, identical to your PCA version but using lv_scores if needed.)
                    ###########################################################################################################################################

                    # ---------- 9) Optional correlation ----------
                    try:
                        pos_first = drivers_df.loc[drivers_df["Loading"] > 0].iloc[0]
                        neg_first = drivers_df.loc[drivers_df["Loading"] < 0].iloc[0]
                        pos_col = pos_first["Wavenumber (cm⁻¹)"]
                        neg_col = neg_first["Wavenumber (cm⁻¹)"]

                        all_idx = pd.Series(wn_focus, index=lbl_focus)
                        pos_label = str(pos_col)
                        neg_label = str(neg_col)
                        if hasattr(X, "columns"):
                            if pos_label in X.columns:
                                x_pos = pd.to_numeric(X[pos_label], errors="coerce")
                            else:
                                pn = _num(pos_label)
                                diffs = pd.Series([abs(_num(c) - pn) for c in X.columns], index=X.columns)
                                x_pos = pd.to_numeric(X[diffs.idxmin()], errors="coerce")

                            if neg_label in X.columns:
                                x_neg = pd.to_numeric(X[neg_label], errors="coerce")
                            else:
                                nn = _num(neg_label)
                                diffs = pd.Series([abs(_num(c) - nn) for c in X.columns], index=X.columns)
                                x_neg = pd.to_numeric(X[diffs.idxmin()], errors="coerce")

                            ratio = (x_pos / x_neg).replace([np.inf, -np.inf], np.nan)
                            valid = ratio.notna() & np.isfinite(pc_scores)
                            if valid.any():
                                corr = float(np.corrcoef(ratio[valid], pd.Series(pc_scores)[valid])[0, 1])
                                st.caption(
                                    f"Correlation proxy (PC score vs I({pos_label})/I({neg_label})): **{corr:.3f}** "
                                    f"(expect positive if the positive-weighted region truly drives PC).")
                    except Exception:
                        pass

                    # # ---------- 10) Downloads ----------
                    out_txt = []
                    out_txt.append(f"Selected: {selected_pc_loading}")
                    out_txt.append(f"Explained variance: {var_pct:.2f}%")
                    out_txt.append(f"Class (+): {class_pos} | Class (−): {class_neg}")
                    out_txt.append("Top drivers (by |loading|):")
                    out_txt.append(drivers_df.to_string(index=False))
                    if len(contrib_df):
                        out_txt.append("\nAuto-detected regions:")
                        out_txt.append(contrib_df.to_string(index=False))


                    ############################################[  model performance summary ]#########################################################################
                    st.markdown("### 📊 Model Performance Summary")

                    run_pc_clf = st.checkbox("📊 Model Performance Summary for selected PC", value=True, key="pca_run_selected_pc")

                    # ============================ Run Models on the selected PC ============================
                    if run_pc_clf:
                        # --- choose 1+ models (all selected by default) ---
                        model_choices_pc = ["SVM", "Logistic Regression", "KNN", "Random Forest", "XGBoost"]
                        selected_models_pc = st.multiselect(
                            "Select classification models (trained on the selected PC score)",
                            options=model_choices_pc,
                            default=model_choices_pc,
                            key="pca_selected_pc_models"
                        )

                        # Parse selected PC index (e.g., "PC3" -> 2)
                        sel_pc_idx = int(str(selected_pc_loading).replace("PC", "")) - 1

                        # --- NEW: single vs cumulative toggle ---
                        mode = st.radio(
                            "PC feature mode",
                            ["Single (selected only)", "Cumulative (PC1..selected)"],
                            index=1,
                            key="pc_feature_mode"
                        )

                        # Build feature matrix X_pc based on mode
                        if mode == "Single (selected only)":
                            X_pc = X_pca_all[:, sel_pc_idx].reshape(-1, 1)
                            pc_label = f"{selected_pc_loading}"
                        else:
                            X_pc = X_pca_all[:, :sel_pc_idx + 1]
                            pc_label = f"PC1–PC{sel_pc_idx + 1}"

                        # Encode labels locally to keep ordering consistent for the report & matrix
                        from sklearn.preprocessing import LabelEncoder

                        le_local = LabelEncoder()
                        y_codes_local = le_local.fit_transform(y)

                        # Train/test split (stratified)
                        from sklearn.model_selection import train_test_split

                        X_train_pc, X_test_pc, y_train_pc, y_test_pc = train_test_split(
                            X_pc, y_codes_local, test_size=0.2, stratify=y_codes_local, random_state=42
                        )

                        # Helper to build a model by name
                        from sklearn.svm import SVC
                        from sklearn.linear_model import LogisticRegression
                        from sklearn.neighbors import KNeighborsClassifier
                        from sklearn.ensemble import RandomForestClassifier
                        from xgboost import XGBClassifier
                        from sklearn.metrics import precision_score, recall_score, f1_score, classification_report, \
                            confusion_matrix
                        import numpy as np
                        import pandas as pd
                        import seaborn as sns
                        import matplotlib.pyplot as plt


                        def _make_model_pc(name):
                            if name == "SVM":
                                return SVC()
                            if name == "Logistic Regression":
                                return LogisticRegression(max_iter=1000)
                            if name == "KNN":
                                return KNeighborsClassifier(n_neighbors=3)
                            if name == "Random Forest":
                                return RandomForestClassifier(n_estimators=100)
                            # XGBoost
                            return XGBClassifier(eval_metric="logloss")



                        # Run all selected models
                        results_pc = []
                        for mdl_name in selected_models_pc:
                            clf_pc = _make_model_pc(mdl_name)
                            clf_pc.fit(X_train_pc, y_train_pc)
                            y_pred_pc = clf_pc.predict(X_test_pc)

                            # Metrics (weighted for multi-class safety)
                            acc_pc = clf_pc.score(X_test_pc, y_test_pc)
                            prec_pc = precision_score(y_test_pc, y_pred_pc, average="weighted", zero_division=0)
                            rec_pc = recall_score(y_test_pc, y_pred_pc, average="weighted", zero_division=0)
                            f1_pc = f1_score(y_test_pc, y_pred_pc, average="weighted", zero_division=0)

                            results_pc.append({
                                "Model": mdl_name,
                                "PC used": pc_label,  # <-- updated label to reflect mode
                                "Accuracy": acc_pc,
                                "Precision (weighted)": prec_pc,
                                "Recall (weighted)": rec_pc,
                                "F1-Score (weighted)": f1_pc
                            })

                        # ===== 📊 Model Performance Summary (selected PC) + Download (one line) =====
                        col1, col2 = st.columns([4, 1])

                        with col1:
                            st.markdown("### 📊 Model Performance Summary (selected PC)")

                        with col2:
                            st.download_button(
                                "📥 Download CSV",
                                data=pd.DataFrame(results_pc).to_csv(index=False).encode("utf-8"),
                                file_name="pca_selected_pc_model_metrics.csv",
                                mime="text/csv"

                            )

                        results_pc_df = pd.DataFrame(results_pc)

                        st.dataframe(
                            results_pc_df.style
                            .background_gradient(cmap="YlGnBu")
                            .format({
                                "Accuracy": "{:.3f}",
                                "Precision (weighted)": "{:.3f}",
                                "Recall (weighted)": "{:.3f}",
                                "F1-Score (weighted)": "{:.3f}"
                            })

                        )

                        # ---- Detailed results per model ----
                        for mdl_name in selected_models_pc:
                            st.markdown(f"### 🔎 Details: {mdl_name} on {pc_label}")  # <-- label reflects mode
                            report_col_pc, matrix_col_pc = st.columns([2, 3])

                            # Rebuild + predict
                            clf_pc = _make_model_pc(mdl_name)
                            clf_pc.fit(X_train_pc, y_train_pc)
                            y_pred_pc = clf_pc.predict(X_test_pc)

                            with report_col_pc:
                                st.markdown("**Classification Report**")
                                report_dict_pc = classification_report(
                                    y_test_pc, y_pred_pc, target_names=list(le_local.classes_), output_dict=True,
                                    zero_division=0
                                )
                                report_df_pc = pd.DataFrame(report_dict_pc).transpose()
                                styled_report_pc = report_df_pc.style.set_table_styles([
                                    {'selector': 'th', 'props': [('border', '1px solid black'), ('padding', '8px'),
                                                                 ('background', '#f2f2f2')]},
                                    {'selector': 'td', 'props': [('border', '1px solid black'), ('padding', '8px')]},
                                    {'selector': 'tr', 'props': [('border', '1px solid black')]}
                                ]).format("{:.2f}")
                                st.write(styled_report_pc)

                            with matrix_col_pc:
                                # Row-normalized confusion matrix -> percentages
                                cm_pc = confusion_matrix(y_test_pc, y_pred_pc, labels=np.arange(len(le_local.classes_)))
                                with np.errstate(divide="ignore", invalid="ignore"):
                                    cm_pct_pc = cm_pc.astype(float) / cm_pc.sum(axis=1, keepdims=True) * 100
                                cm_pct_pc = np.nan_to_num(cm_pct_pc)

                                annot_labels_pc = np.array([[f"{v:.1f}%" for v in row] for row in cm_pct_pc])

                                fig_cm_pc, ax_pc = plt.subplots()
                                sns.heatmap(
                                    cm_pct_pc,
                                    annot=annot_labels_pc,
                                    fmt="",
                                    cmap="Blues",
                                    vmin=0, vmax=100,
                                    xticklabels=list(le_local.classes_),
                                    yticklabels=list(le_local.classes_),
                                    cbar_kws={"label": "% of true class"},
                                    ax=ax_pc
                                )
                                ax_pc.set_xlabel("Predicted")
                                ax_pc.set_ylabel("True")
                                ax_pc.set_title(
                                    f"{mdl_name} — Confusion Matrix (%) on {pc_label}")  # <-- label reflects mode
                                st.pyplot(fig_cm_pc)
                                plt.close(fig_cm_pc)

                    # ============================== One-file export for the entire "🔬 PCA: Principal Component Projection" ==============================

                    import io, numpy as np, pandas as pd, matplotlib.pyplot as plt, seaborn as sns
                    from sklearn.preprocessing import LabelEncoder
                    from sklearn.model_selection import train_test_split
                    from sklearn.metrics import classification_report, confusion_matrix
                    from sklearn.svm import SVC
                    from sklearn.linear_model import LogisticRegression
                    from sklearn.neighbors import KNeighborsClassifier
                    from sklearn.ensemble import RandomForestClassifier

                    try:
                        from xgboost import XGBClassifier
                        _has_xgb = True
                    except Exception:
                        _has_xgb = False


                    # ---------- helper: auto-fit Excel columns ----------
                    def _autosize_columns(ws, df, extra_pad=2, max_width=60):
                        for i, col in enumerate(df.columns):
                            series = df[col].astype(str)
                            max_len = max([len(str(col))] + [len(s) for s in series.tolist()])
                            ws.set_column(i, i, min(max_len + extra_pad, max_width))


                    def _autosize_range(ws, first_col, last_col, rows, headers=None, extra_pad=2, max_width=60):
                        # Auto-fit arbitrary cell blocks (for Model_Metrics where we write cells)
                        for c in range(first_col, last_col + 1):
                            values = []
                            if headers and c - first_col < len(headers):
                                values.append(str(headers[c - first_col]))
                            for r in rows:
                                v = r.get(c, "")
                                values.append("" if v is None else str(v))
                            width = min(max(len(x) for x in values) + extra_pad, max_width) if values else 10
                            ws.set_column(c, c, width)


                    # ---------- (A) Rebuild model results (AUTO-RUN; independent of any checkbox) ----------
                    _sel_pc_idx = int(str(selected_pc_loading).replace("PC", "")) - 1
                    X_pc_export = X_pca_all[:, :_sel_pc_idx + 1]  # cumulative PC1..selected
                    pc_label_export = f"PC1–PC{_sel_pc_idx + 1}"

                    le_export = LabelEncoder()
                    y_codes_export = le_export.fit_transform(y)

                    X_train_e, X_test_e, y_train_e, y_test_e = train_test_split(
                        X_pc_export, y_codes_export, test_size=0.2, stratify=y_codes_export, random_state=42
                    )

                    models = {
                        "SVM": SVC(),
                        "Logistic Regression": LogisticRegression(max_iter=1000),
                        "KNN": KNeighborsClassifier(n_neighbors=3),
                        "Random Forest": RandomForestClassifier(n_estimators=100),
                    }
                    if _has_xgb:
                        models["XGBoost"] = XGBClassifier(eval_metric="logloss")


                    results_rows = []
                    confmats_counts = {}
                    confmats_pct = {}
                    confmat_heatmap_png = {}

                    for name, clf in models.items():
                        clf.fit(X_train_e, y_train_e)
                        y_pred = clf.predict(X_test_e)

                        acc = clf.score(X_test_e, y_test_e)
                        rep = classification_report(
                            y_test_e, y_pred, target_names=list(le_export.classes_), output_dict=True, zero_division=0
                        )
                        prec = rep["weighted avg"]["precision"]
                        rec = rep["weighted avg"]["recall"]
                        f1 = rep["weighted avg"]["f1-score"]

                        results_rows.append({
                            "Model": name,
                            "PC used": pc_label_export,
                            "Accuracy": acc,
                            "Precision (weighted)": prec,
                            "Recall (weighted)": rec,
                            "F1-Score (weighted)": f1
                        })

                        # Confusion matrices
                        cm = confusion_matrix(y_test_e, y_pred, labels=np.arange(len(le_export.classes_)))
                        confmats_counts[name] = pd.DataFrame(cm, index=list(le_export.classes_),
                                                             columns=list(le_export.classes_))

                        with np.errstate(divide="ignore", invalid="ignore"):
                            cm_pct = cm.astype(float) / cm.sum(axis=1, keepdims=True) * 100.0
                        cm_pct = np.nan_to_num(cm_pct)
                        confmats_pct[name] = pd.DataFrame(cm_pct, index=list(le_export.classes_),
                                                          columns=list(le_export.classes_))

                        # Heatmap PNG (in-memory)
                        fig, ax = plt.subplots(figsize=(5, 4))
                        sns.heatmap(cm_pct, annot=True, fmt=".1f", cmap="Blues",
                                    xticklabels=list(le_export.classes_), yticklabels=list(le_export.classes_),
                                    cbar=True, ax=ax, vmin=0, vmax=100)
                        ax.set_xlabel("Predicted");
                        ax.set_ylabel("True")
                        ax.set_title(f"{name} — Confusion Matrix (%) on {pc_label_export}")
                        _png = io.BytesIO()
                        fig.tight_layout()
                        fig.savefig(_png, format="png", dpi=180)
                        plt.close(fig)
                        _png.seek(0)
                        confmat_heatmap_png[name] = _png

                    results_export_df = pd.DataFrame(results_rows)

                    # ---------- (B) Loadings PNG for 'Loadings Graph' (cleaner visual) ----------
                    try:
                        _pc_num = int(str(selected_pc_loading).replace("PC", ""))
                        _var_pct = float(explained_var[_pc_num - 1]) * 100.0
                    except Exception:
                        _var_pct = np.nan

                    _series_vals = pd.to_numeric(loadings[selected_pc_loading], errors="coerce").values
                    _x_idx = np.arange(len(_series_vals))
                    figL, axL = plt.subplots(figsize=(11, 3.2))
                    axL.plot(_x_idx, _series_vals, linewidth=2.8)
                    axL.axhline(0.0, linestyle="--", color="gray", linewidth=1)
                    axL.set_title(f"{selected_pc_loading} Loadings ({_var_pct:.2f}% Variance)")
                    axL.set_xlabel("Index (wavenumber order)")
                    axL.set_ylabel("Loading")
                    axL.grid(alpha=0.15)
                    _loadings_png = io.BytesIO()
                    figL.tight_layout()
                    figL.savefig(_loadings_png, format="png", dpi=180)
                    plt.close(figL)
                    _loadings_png.seek(0)

                    # ---------- (C) Data for 2D/3D projections ----------
                    proj2d_df = df_pca_all[[pc_x, pc_y, "Label"]].copy()
                    if pc_z in df_pca_all.columns:
                        proj3d_df = df_pca_all[[pc_x, pc_y, pc_z, "Label"]].copy()
                    else:
                        proj3d_df = pd.DataFrame(columns=[pc_x, pc_y, "Z", "Label"])

                    # Optional PNG for 3D visualization (Excel has no native 3D scatter)
                    try:
                        from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

                        fig3, ax3 = plt.subplots(subplot_kw={"projection": "3d"}, figsize=(6, 5))
                        for lab in proj3d_df["Label"].unique():
                            d = proj3d_df[proj3d_df["Label"] == lab]
                            ax3.scatter(d[pc_x], d[pc_y], d[pc_z], label=str(lab), s=16)
                        ax3.set_xlabel(pc_x);
                        ax3.set_ylabel(pc_y);
                        ax3.set_zlabel(pc_z)
                        ax3.set_title("PCA 3D Visualization")
                        ax3.legend(loc="best", fontsize=8)
                        _pca3d_png = io.BytesIO()
                        fig3.tight_layout()
                        fig3.savefig(_pca3d_png, format="png", dpi=180)
                        plt.close(fig3)
                        _pca3d_png.seek(0)
                    except Exception:
                        _pca3d_png = None

                    # ---------- (D) Build Excel with all pages & editable charts ----------
                    buffer_full = io.BytesIO()
                    with pd.ExcelWriter(buffer_full, engine="xlsxwriter") as writer:
                        wb = writer.book

                        # --- PCA Summary ---
                        try:
                            pca_summary.to_excel(writer, sheet_name="PCA Summary", index=False)
                            _autosize_columns(writer.sheets["PCA Summary"], pca_summary)
                        except Exception:
                            pass

                        # --- Variable Importance / Drivers / Regions ---
                        try:
                            importance_df.to_excel(writer, sheet_name="Variable Importance", index=False)
                            _autosize_columns(writer.sheets["Variable Importance"], importance_df)
                        except Exception:
                            pass

                        try:
                            _drivers_view = drivers_df[
                                ["Wavenumber (cm⁻¹)", "Loading", "|Loading|", "Band Annotation", "Interpretation"]]
                            _drivers_view.to_excel(writer, sheet_name="Most Influential", index=False)
                            _autosize_columns(writer.sheets["Most Influential"], _drivers_view)
                        except Exception:
                            pass

                        try:
                            contrib_df.to_excel(writer, sheet_name="Regions", index=False)
                            _autosize_columns(writer.sheets["Regions"], contrib_df)
                        except Exception:
                            pass

                        # --- Loadings (table + Excel-editable chart) ---
                        try:
                            def _num_first(val):
                                s = str(val).strip()
                                if "/" in s:
                                    s = s.split("/")[0].strip()
                                try:
                                    return float(s)
                                except Exception:
                                    return np.nan


                            load_tbl = pd.DataFrame({
                                "Wavenumber (cm⁻¹)": list(loadings.index),
                                "Numeric Wavenumber": [_num_first(v) for v in loadings.index],
                                "Index": np.arange(len(loadings.index)),
                                "Loading": pd.to_numeric(loadings[selected_pc_loading], errors="coerce").values
                            })

                            load_tbl.to_excel(writer, sheet_name="Loadings", index=False)
                            ws_load = writer.sheets["Loadings"]
                            _autosize_columns(ws_load, load_tbl)

                            # Editable line chart (cleaner): thicker stroke, legend bottom
                            chart = wb.add_chart({"type": "line"})
                            n_rows = len(load_tbl)
                            chart.add_series({
                                "name": f"{selected_pc_loading}",
                                "categories": ["Loadings", 1, 1, n_rows, 1],  # Numeric Wavenumber
                                "values": ["Loadings", 1, 3, n_rows, 3],  # Loading
                                "line": {"width": 2.25},
                            })
                            chart.set_title({"name": f"Loadings — {selected_pc_loading} ({_var_pct:.2f}% var)"})
                            chart.set_x_axis({"name": "Wavenumber (cm⁻¹)"})
                            chart.set_y_axis({"name": "Loading"})
                            chart.set_legend({"position": "bottom"})
                            ws_load.insert_chart(n_rows + 3, 0, chart)
                        except Exception:
                            pass

                        # --- 📝 Auto-summary ---
                        try:
                            auto_lines = []
                            auto_lines.extend(summary)
                            auto_lines.append("Key drivers:")
                            auto_lines.extend(bullets)
                            auto_df = pd.DataFrame({"Summary": auto_lines})
                            auto_df.to_excel(writer, sheet_name="📝 Auto-summary", index=False)
                            _autosize_columns(writer.sheets["📝 Auto-summary"], auto_df)
                        except Exception:
                            pass

                        # --- 📉 PCA 2D Projection (data + editable scatter chart; markers only, no lines) ---
                        try:
                            sheet2d = "📉 PCA 2D Projection"
                            proj2d_df.to_excel(writer, sheet_name=sheet2d, index=False)
                            ws2d = writer.sheets[sheet2d]
                            _autosize_columns(ws2d, proj2d_df)

                            chart2d = wb.add_chart({"type": "scatter"})
                            cur_row = len(proj2d_df) + 3
                            for lab in proj2d_df["Label"].unique():
                                sub = proj2d_df[proj2d_df["Label"] == lab][[pc_x, pc_y]].reset_index(drop=True)
                                if sub.empty:
                                    continue
                                ws2d.write(cur_row, 0, f"{lab} — {pc_x}")
                                ws2d.write(cur_row, 1, f"{lab} — {pc_y}")
                                ws2d.write_column(cur_row + 1, 0, sub[pc_x].tolist())
                                ws2d.write_column(cur_row + 1, 1, sub[pc_y].tolist())

                                s_start, s_end = cur_row + 1, cur_row + len(sub)
                                chart2d.add_series({
                                    "name": str(lab),
                                    "categories": [sheet2d, s_start, 0, s_end, 0],
                                    "values": [sheet2d, s_start, 1, s_end, 1],
                                    "marker": {"type": "circle", "size": 5},
                                    "line": {"none": True},
                                })
                                cur_row = s_end + 2

                            chart2d.set_title({"name": f"PCA 2D: {pc_x} vs {pc_y}"})
                            chart2d.set_x_axis({"name": pc_x})
                            chart2d.set_y_axis({"name": pc_y})
                            chart2d.set_legend({"position": "bottom"})
                            ws2d.insert_chart(1, 4, chart2d)
                        except Exception:
                            pass

                        # --- 🧭 PCA 3D Visualization (data + PNG) ---
                        try:
                            sheet3d = "🧭 PCA 3D Visualization"
                            proj3d_df.to_excel(writer, sheet_name=sheet3d, index=False)
                            ws3d = writer.sheets[sheet3d]
                            _autosize_columns(ws3d, proj3d_df)
                            if _pca3d_png is not None:
                                ws3d.insert_image(1, 5, "pca3d.png", {"image_data": _pca3d_png})
                        except Exception:
                            pass

                        # --- Model_Metrics (summary + all confusion matrices + heatmaps on one sheet) ---
                        wsM = wb.add_worksheet("Model_Metrics")

                        headers = list(results_export_df.columns)
                        for c, h in enumerate(headers):
                            wsM.write(0, c, h)
                        for r, row in results_export_df.iterrows():
                            for c, h in enumerate(headers):
                                wsM.write(r + 1, c, row[h])
                        _autosize_columns(wsM, results_export_df)

                        start_row = len(results_export_df) + 3

                        for name in models.keys():
                            wsM.write(start_row, 0, f"{name} — Confusion Matrix (counts)")
                            cnt = confmats_counts[name]
                            for j, col in enumerate([""] + cnt.columns.tolist()):
                                wsM.write(start_row + 1, j, col)
                            for i, idx in enumerate(cnt.index.tolist()):
                                wsM.write(start_row + 2 + i, 0, idx)
                                for j, col in enumerate(cnt.columns.tolist()):
                                    wsM.write(start_row + 2 + i, 1 + j, int(cnt.loc[idx, col]))
                            _autosize_range(
                                wsM,
                                first_col=0,
                                last_col=len(cnt.columns),
                                rows=[{k: None for k in range(len(cnt.columns) + 1)} for _ in range(len(cnt.index) + 2)],
                                headers=[""] + cnt.columns.tolist()
                            )
                            start_row += len(cnt) + 3

                            wsM.write(start_row, 0, f"{name} — Confusion Matrix (%)")
                            pct = confmats_pct[name]
                            for j, col in enumerate([""] + pct.columns.tolist()):
                                wsM.write(start_row + 1, j, col)
                            for i, idx in enumerate(pct.index.tolist()):
                                wsM.write(start_row + 2 + i, 0, idx)
                                for j, col in enumerate(pct.columns.tolist()):
                                    wsM.write_number(start_row + 2 + i, 1 + j, float(pct.loc[idx, col]))
                            img_col = pct.shape[1] + 3
                            wsM.insert_image(start_row + 1, img_col, f"{name}_cm.png",
                                             {"image_data": confmat_heatmap_png[name], "x_scale": 1.0, "y_scale": 1.0})
                            start_row += len(pct) + 10

                    buffer_full.seek(0)

                    # ================= Download Complete PCA Analysis =================
                    # st.markdown("### 📥 Download Complete PCA Analysis")

                    colA, colB = st.columns([4, 1])  # adjust ratio if you want button size balance

                    with colA:
                        st.download_button(
                            "📥 Download PCA Full Workbook (editable loadings chart, 2D chart, 3D, summaries, all confusion matrices)",
                            data=buffer_full.getvalue(),
                            file_name="pca_full_export.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                        )

                    with colB:
                        st.empty()  # placeholder in case you want another button/space

                    # =====================================================================================================================================

            ###############    📊 PLS: Partial Least Squares Regression
            if show_pls:

                col1, col2 = st.columns([4, 1])  # 4:1 width ratio (adjust as needed)

                with col1:
                    st.subheader("📊 PLS: Partial Least Squares Regression")


                with st.expander(
                    "PLS is a supervised projection method. This panel shows a summary table, interactive 2D/3D projections with selectable axes, loadings for a chosen component, and variable-importance with bond annotations.",
                    expanded=False
                ):

                    ############################################ [ PLS Analysis ] ##############################################################################

                    import numpy as np
                    import pandas as pd
                    import streamlit as st
                    from sklearn.cross_decomposition import PLSRegression
                    from sklearn.model_selection import StratifiedKFold, KFold
                    from sklearn.metrics import mean_squared_error, r2_score
                    from scipy.signal import savgol_filter
                    from scipy.sparse import diags, eye
                    from scipy.sparse.linalg import spsolve

                    # ---------------- Preprocessing selection (combined) ----------------
                    PREPROCESSING_OPTIONS = ["Mean Center", "Autoscale", "Smoothing", "Normalization", "Baseline",
                                             "Second Derivative"]
                    x_preproc_pls = st.multiselect(
                        "Select Preprocessing Techniques for PLS (applied together, in order)",
                        options=PREPROCESSING_OPTIONS,
                        default=["Autoscale"]
                    )

                    # ========== Controls ==========
                    max_pls = int(min(10, X.shape[1], len(selected_columns)))
                    n_comp = st.slider("Number of PLS components (LVs)", 2, max_pls, value=min(5, max_pls), step=1)


                    # ===================== helpers: Savitzky–Golay =====================
                    def _savgol_params(n_feat: int, window: int, poly: int, deriv: int):
                        w_max_odd = n_feat if (n_feat % 2 == 1) else (n_feat - 1)
                        if w_max_odd < 3:
                            return None
                        w = int(window)
                        if w % 2 == 0: w -= 1
                        if w < 3: w = 3
                        if w > w_max_odd: w = w_max_odd
                        p = int(poly)
                        p = max(p, deriv, 1)
                        if p >= w: p = w - 1
                        return w, p

                    def smooth_savgol(X, window=7, poly=3):
                        Xn = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                        Xn = Xn.copy()
                        params = _savgol_params(Xn.shape[1], window, poly, deriv=0)
                        if params is None: return Xn
                        w, p = params
                        return savgol_filter(Xn, window_length=w, polyorder=p, deriv=0, axis=1, mode="interp")

                    def second_derivative_savgol(X, window=7, poly=3):
                        Xn = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                        Xn = Xn.copy()
                        params = _savgol_params(Xn.shape[1], window, poly, deriv=2)
                        if params is None: return Xn
                        w, p = params
                        return savgol_filter(Xn, window_length=w, polyorder=p, deriv=2, axis=1, mode="interp")


                    # ===================== helpers: Baseline (ALS, sparse + fast) =====================
                    def als_baseline_sparse(y, lam=1e5, p=0.001, niter=10, eps=1e-12):
                        y = np.asarray(y, dtype=float).ravel()
                        L = y.size
                        if L < 3:
                            return np.zeros_like(y)
                        D = diags([1, -2, 1], [0, 1, 2], shape=(L - 2, L))
                        DTD = lam * (D.T @ D)
                        w = np.ones(L)
                        I = eye(L) * eps
                        for _ in range(niter):
                            W = diags(w, 0, shape=(L, L))
                            z = spsolve(W + DTD + I, w * y)
                            w = p * (y > z) + (1 - p) * (y < z)
                        return z


                    # ===================== helpers: leakage-safe MC/Autoscale =====================
                    def fit_center_scale_stats(X_arr, need_mean: bool, need_std: bool):
                        mu = sd = None
                        if need_mean or need_std:
                            mu = X_arr.mean(axis=0, keepdims=True)
                        if need_std:
                            sd = X_arr.std(axis=0, keepdims=True)
                            sd = sd.copy();
                            sd[sd == 0] = 1.0
                        return mu, sd


                    def apply_preprocessing_chain(
                            X,
                            ops, *,
                            smooth_window=7, smooth_poly=3,
                            normalization_mode="l2",
                            baseline_lam=100_000.0, baseline_p=0.001,
                            mean_=None, std_=None
                    ):
                        Xn = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                        Xn = Xn.copy()

                        # 1) Baseline
                        if "Baseline" in ops:
                            Xcorr = np.empty_like(Xn)
                            for i in range(Xn.shape[0]):
                                bl = als_baseline_sparse(Xn[i, :], lam=baseline_lam, p=baseline_p)
                                Xcorr[i, :] = Xn[i, :] - bl
                            Xn = Xcorr

                        # 2) Smoothing / Second Derivative
                        if "Smoothing" in ops and "Second Derivative" in ops:
                            # If both checked, prefer derivative (common chemometrics convention)
                            Xn = second_derivative_savgol(Xn, window=smooth_window, poly=smooth_poly)
                        elif "Smoothing" in ops:
                            Xn = smooth_savgol(Xn, window=smooth_window, poly=smooth_poly)
                        elif "Second Derivative" in ops:
                            Xn = second_derivative_savgol(Xn, window=smooth_window, poly=smooth_poly)

                        # 3) Mean Center / Autoscale
                        if "Mean Center" in ops or "Autoscale" in ops:
                            mu = mean_ if mean_ is not None else Xn.mean(axis=0, keepdims=True)
                            Xn = Xn - mu
                        if "Autoscale" in ops:
                            sd = std_ if std_ is not None else Xn.std(axis=0, keepdims=True)
                            sd = sd.copy();
                            sd[sd == 0] = 1.0
                            Xn = Xn / sd

                        # 4) Normalization (row-wise)
                        if "Normalization" in ops:
                            mode = normalization_mode.lower()
                            if mode == "l2":
                                denom = np.linalg.norm(Xn, axis=1, keepdims=True)
                            elif mode == "l1":
                                denom = np.sum(np.abs(Xn), axis=1, keepdims=True)
                            else:
                                denom = np.max(np.abs(Xn), axis=1, keepdims=True)
                            denom[denom == 0] = 1.0
                            Xn = Xn / denom

                        return Xn


                    # ===================== helpers: Y shape & metrics =====================
                    def _get_y_matrix():
                        """Fetch y from globals and return as (n_samples, n_targets) float array."""
                        y_name = None
                        if "y_encoded" in globals():
                            y_val = globals()["y_encoded"];
                            y_name = "y_encoded"
                        elif "y" in globals():
                            y_val = globals()["y"];
                            y_name = "y"
                        else:
                            st.error("PLS requires a target vector `y` or `y_encoded` to be defined.")
                            raise RuntimeError("Missing y")

                        y_arr = y_val.values if hasattr(y_val, "values") else y_val
                        y_arr = np.asarray(y_arr)
                        if y_arr.ndim == 1:
                            y_arr = y_arr.reshape(-1, 1)
                        return y_arr, y_name


                    def _rmse_r2_bias(y_true, y_pred):
                        """RMSE / R2 / Bias averaged across targets if multi-output."""
                        y_true = np.asarray(y_true, dtype=float)
                        y_pred = np.asarray(y_pred, dtype=float)
                        if y_true.ndim == 1: y_true = y_true.reshape(-1, 1)
                        if y_pred.ndim == 1: y_pred = y_pred.reshape(-1, 1)
                        # RMSE averaged over targets
                        rmse = float(np.sqrt(np.mean((y_true - y_pred) ** 2)))
                        # R2 averaged over targets
                        try:
                            r2_vals = []
                            for j in range(y_true.shape[1]):
                                r2_vals.append(r2_score(y_true[:, j], y_pred[:, j]))
                            r2 = float(np.mean(r2_vals))
                        except Exception:
                            r2 = float("nan")
                        bias = float(np.mean(y_pred - y_true))
                        return rmse, r2, bias


                    # ===================== Controls (PLS) =====================
                    max_pls = max(2, min(10, X.shape[1]))

                    cv_splits_pls = st.slider("CV Splits (PLS)", 3, 15, value=10, step=1)
                    conf_limit_pls = st.selectbox("Confidence Limit (PLS)", [0.90, 0.95, 0.99], index=1)
                    lv_list_pls = st.multiselect(
                        "Ncomp/LVs to evaluate (PLS)",
                        options=list(range(2, max_pls + 1)),
                        default=list(range(2, min(3, max_pls) + 1))
                    )

                    # ===================== CAL & CV with the combined preprocessing =====================
                    try:
                        y_mat, y_label_used = _get_y_matrix()
                    except RuntimeError:
                        st.stop()

                    rows_pls = []
                    model_id = 1

                    # Label for the chosen combination
                    preproc_label_pls = ", ".join(x_preproc_pls) if x_preproc_pls else "None"

                    # ---- Calibration data: apply full chain to ALL X ----
                    X_all = apply_preprocessing_chain(
                        X, x_preproc_pls,
                        smooth_window=7, smooth_poly=3,
                        baseline_lam=100_000.0, baseline_p=0.001
                    )

                    X_size_pls = f"{X.shape[0]} x {X.shape[1]}"
                    Y_size_pls = f"{y_mat.shape[0]} x {y_mat.shape[1]}"

                    for ncomp in lv_list_pls:
                        # ===== Calibration on ALL data =====
                        pls_all = PLSRegression(n_components=ncomp)
                        pls_all.fit(X_all, y_mat)
                        y_hat_all = pls_all.predict(X_all)

                        rmsec_pls, r2c_pls, bias_c_pls = _rmse_r2_bias(y_mat, y_hat_all)

                        # ===== CV =====
                        # Prefer StratifiedKFold when y is 1D discrete-like, otherwise KFold
                        y_for_split_1d = y_mat.ravel()  # flatten for stratification attempt
                        try:
                            folds = StratifiedKFold(n_splits=cv_splits_pls, shuffle=True, random_state=42).split(X,
                                                                                                                 y_for_split_1d)
                        except Exception:
                            folds = KFold(n_splits=cv_splits_pls, shuffle=True, random_state=42).split(X)

                        y_cv_pred = np.zeros_like(y_mat, dtype=float)

                        for tr_idx, te_idx in folds:
                            X_tr = X.iloc[tr_idx].values.astype(float)
                            X_te = X.iloc[te_idx].values.astype(float)
                            y_tr = y_mat[tr_idx, :]

                            # Train-only stats for MC/Autoscale if selected
                            need_mean = ("Mean Center" in x_preproc_pls) or ("Autoscale" in x_preproc_pls)
                            need_std = ("Autoscale" in x_preproc_pls)
                            mu_tr, sd_tr = fit_center_scale_stats(X_tr, need_mean, need_std)

                            X_tr_s = apply_preprocessing_chain(
                                X_tr, x_preproc_pls,
                                smooth_window=7, smooth_poly=3,
                                baseline_lam=100_000.0, baseline_p=0.001,
                                mean_=mu_tr, std_=sd_tr
                            )
                            X_te_s = apply_preprocessing_chain(
                                X_te, x_preproc_pls,
                                smooth_window=7, smooth_poly=3,
                                baseline_lam=100_000.0, baseline_p=0.001,
                                mean_=mu_tr, std_=sd_tr
                            )

                            pls_cv = PLSRegression(n_components=ncomp)
                            pls_cv.fit(X_tr_s, y_tr)
                            y_cv_pred[te_idx, :] = pls_cv.predict(X_te_s)

                        rmsecv_pls, r2cv_pls, bias_cv_pls = _rmse_r2_bias(y_mat, y_cv_pred)
                        rmse_ratio_pls = rmsecv_pls / rmsec_pls if rmsec_pls > 0 else float("nan")

                        rows_pls.append({
                            "CV Splits": cv_splits_pls,
                            "Confidence Limit": conf_limit_pls,
                            "RMSE Ratio (RMSECV/RMSEC)": rmse_ratio_pls,
                            "RMSEC (Cal)": rmsec_pls,
                            "RMSECV (CV)": rmsecv_pls,
                            "Model Name": f"Model {model_id}",
                            "Ncomp/LVs": ncomp,
                            "X-Preprocessing": preproc_label_pls,
                            "X Include Size": X_size_pls,
                            "Y Include Size": Y_size_pls,
                            "R2C (Cal)": r2c_pls,
                            "R2CV (CV)": r2cv_pls,
                            "Bias": bias_c_pls,
                            "Bias (CV)": bias_cv_pls
                        })
                        model_id += 1

                    # ---------------- Table + download ----------------
                    pls_calcv_df = pd.DataFrame(rows_pls, columns=[
                        "CV Splits", "Confidence Limit", "RMSE Ratio (RMSECV/RMSEC)",
                        "RMSEC (Cal)", "RMSECV (CV)", "Model Name", "Ncomp/LVs",
                        "X-Preprocessing", "X Include Size", "Y Include Size",
                        "R2C (Cal)", "R2CV (CV)", "Bias", "Bias (CV)"
                    ])

                    fmt_cols_pls = ["RMSE Ratio (RMSECV/RMSEC)", "RMSEC (Cal)", "RMSECV (CV)", "R2C (Cal)", "R2CV (CV)",
                                    "Bias", "Bias (CV)"]

                    st.markdown("### 📋 PLS Analysis Result Summary")
                    st.dataframe(
                        pls_calcv_df.style
                        .format({c: "{:.4f}" for c in fmt_cols_pls})
                        .background_gradient(subset=["R2CV (CV)", "R2C (Cal)"], cmap="YlGn")
                        .background_gradient(subset=["RMSECV (CV)", "RMSEC (Cal)"], cmap="YlOrRd_r")

                    )

                    st.download_button(
                        "📥 Download PLS Summary (CSV)",
                        data=pls_calcv_df.to_csv(index=False).encode("utf-8"),
                        file_name="pls_summary_table.csv",
                        mime="text/csv"
                    )

                    ##########################################################################################################################

                    ##### Display plots

                    # Scale X (recommended for PLS on spectroscopy)
                    scaler = StandardScaler()
                    X_scaled = scaler.fit_transform(X)

                    # y must be continuous for PLSRegression; for classification we use the encoded labels
                    pls = PLSRegression(n_components=n_comp)
                    pls.fit(X_scaled, y_encoded)

                    # Scores / Loadings / Weights
                    T = pls.x_scores_                    # (n_samples, n_comp)
                    U = pls.y_scores_                    # (n_samples, n_comp)
                    P = pls.x_loadings_                  # (n_features, n_comp)
                    W = pls.x_weights_                   # (n_features, n_comp)
                    Q = pls.y_loadings_                  # (1 or n_classes, n_comp) depending on encoding

                    # ========== Axis Selection for 2D/3D ==========
                    all_pls_labels = [f"PLS{i+1}" for i in range(n_comp)]
                    col_ax1, col_ax2, col_ax3 = st.columns(3)
                    with col_ax1:
                        pls_x_label = st.selectbox("X-axis", all_pls_labels, index=0, key="pls_x_axis")
                    with col_ax2:
                        pls_y_label = st.selectbox("Y-axis (2D/3D)", all_pls_labels, index=min(1, n_comp-1), key="pls_y_axis")
                    with col_ax3:
                        pls_z_label = st.selectbox("Z-axis (3D)", all_pls_labels, index=min(2, n_comp-1), key="pls_z_axis")

                    x_idx = int(pls_x_label.replace("PLS","")) - 1
                    y_idx = int(pls_y_label.replace("PLS","")) - 1
                    z_idx = int(pls_z_label.replace("PLS","")) - 1

                    # Build a DataFrame of selected scores
                    df_pls = pd.DataFrame({
                        "Label": y.values,
                        "PLS1": T[:,0] if n_comp >= 1 else np.nan,
                        "PLS2": T[:,1] if n_comp >= 2 else np.nan,
                        "PLS3": T[:,2] if n_comp >= 3 else np.nan
                    })
                    # Include arbitrary components by their chosen indexes
                    for i in range(n_comp):
                        df_pls[f"PLS{i+1}"] = T[:, i]

                    # 2D Plot
                    st.subheader("📉 PLS 2D Projection")
                    xcol = f"PLS{x_idx + 1}"
                    ycol = f"PLS{y_idx + 1}"

                    fig_pls_2d = px.scatter(
                        df_pls,
                        x=xcol,
                        y=ycol,
                        color="Label",
                        title=f"PLS Projection ({pls_x_label} vs {pls_y_label})",
                        labels={xcol: pls_x_label, ycol: pls_y_label}
                    )

                    # ---- Add class ellipses (same approach as PCA) ----
                    chi2_map = {0.90: 4.605, 0.95: 5.991, 0.99: 9.210}
                    chi2_val = chi2_map.get(conf_limit_pls, 5.991)  # default 95%


                    def _ellipse_xy(points_2d, chi2_q, npts=200):
                        pts = np.asarray(points_2d)
                        if pts.shape[0] < 3:
                            return None, None
                        mu = pts.mean(axis=0)
                        S = np.cov(pts.T)
                        vals, vecs = np.linalg.eigh(S)
                        order = np.argsort(vals)[::-1]
                        vals, vecs = vals[order], vecs[:, order]
                        radii = np.sqrt(np.maximum(vals, 1e-12) * chi2_q)
                        t = np.linspace(0, 2 * np.pi, npts)
                        circ = np.stack([np.cos(t) * radii[0], np.sin(t) * radii[1]], axis=0)
                        ell = (vecs @ circ).T + mu
                        return ell[:, 0], ell[:, 1]


                    for lab in df_pls["Label"].unique():
                        d = df_pls[df_pls["Label"] == lab][[xcol, ycol]].values
                        ex, ey = _ellipse_xy(d, chi2_val)
                        if ex is not None:
                            fig_pls_2d.add_trace(
                                go.Scatter(
                                    x=ex, y=ey, mode="lines",
                                    name=f"{lab} {int(conf_limit_pls * 100)}% ellipse",
                                    line=dict(width=1),
                                    showlegend=False
                                )
                            )

                    st.plotly_chart(fig_pls_2d)

                    # 3D Plot (only if at least 3 comps)
                    if n_comp >= 3:
                        st.subheader("🧭 PLS 3D Visualization")
                        fig_pls_3d = px.scatter_3d(
                            df_pls,
                            x=f"PLS{x_idx+1}",
                            y=f"PLS{y_idx+1}",
                            z=f"PLS{z_idx+1}",
                            color="Label",
                            title=f"3D PLS: {pls_x_label} vs {pls_y_label} vs {pls_z_label}",
                            height=520, width=800
                        )
                        st.plotly_chart(fig_pls_3d)
                    else:
                        st.info("Add ≥3 components to view the 3D projection.")

                    #### ======================= PLS Loadings with Bond Annotation =======================

                    # -------------------- Helpers --------------------
                    def _num_first(val):
                        s = str(val).strip()
                        if "/" in s:
                            s = s.split("/")[0].strip()
                        try:
                            return float(s)
                        except Exception:
                            return np.nan


                    # -------------------- Prepare X (numeric) --------------------
                    if isinstance(X, pd.DataFrame):
                        X_num = X.apply(pd.to_numeric, errors="coerce").values
                        selected_columns = list(X.columns)
                    else:
                        X_num = np.asarray(X, dtype=float)
                        selected_columns = [str(c) for c in range(X_num.shape[1])]

                    # fill NaNs if any
                    if np.isnan(X_num).any():
                        col_means = np.nanmean(X_num, axis=0)
                        ii = np.where(np.isnan(X_num))
                        X_num[ii] = np.take(col_means, ii[1])

                    # -------------------- Prepare Y (numeric / PLS-DA) --------------------
                    if isinstance(y, pd.Series):
                        y_arr = y.values
                    elif isinstance(y, pd.DataFrame):
                        y_arr = y.values
                    else:
                        y_arr = np.asarray(y)

                    if y_arr.ndim == 1 and not np.issubdtype(y_arr.dtype, np.number):
                        _le = LabelEncoder()
                        y_codes = _le.fit_transform(y_arr)
                        if len(np.unique(y_codes)) == 2:
                            Y_matrix = y_codes.reshape(-1, 1).astype(float)
                        else:
                            Y_matrix = np.eye(len(np.unique(y_codes)), dtype=float)[y_codes]
                    elif y_arr.ndim == 1:
                        Y_matrix = y_arr.reshape(-1, 1).astype(float)
                    else:
                        Y_matrix = y_arr.astype(float)

                    # -------------------- Fit PLS (if not already) & explained X-variance --------------------
                    n_samples, p = X_num.shape
                    q = 1 if Y_matrix.ndim == 1 else Y_matrix.shape[1]
                    max_lv_safe = max(1, min(n_samples - 1, p, max(1, q)))
                    ncomp_pls = min(5, max_lv_safe)

                    try:
                        pls  # reuse if already fitted above
                        T_all = pls.x_scores_
                    except Exception:
                        pls = PLSRegression(n_components=ncomp_pls, scale=False).fit(X_num, Y_matrix)
                        T_all = pls.x_scores_

                    explained_x = getattr(pls, "x_explained_variance_ratio_", None)
                    if explained_x is None:
                        X_tot = float(np.var(X_num, axis=0).sum()) or 1.0
                        explained_x = np.var(T_all, axis=0) / X_tot
                    explained_x = np.asarray(explained_x, dtype=float)

                    # -------------------- Compute PLS Loadings (X-loadings) --------------------
                    loadings = pd.DataFrame(
                        pls.x_loadings_,
                        index=selected_columns,
                        columns=[f"PLS{i + 1}" for i in range(pls.x_loadings_.shape[1])]
                    )

                    # -------------------- Normalize explained X-variance for nice percentages --------------------
                    # Prefer sklearn's attribute if present; otherwise compute from scores and normalize to sum=1
                    T_all = getattr(pls, "x_scores_", None)
                    explained_x_attr = getattr(pls, "x_explained_variance_ratio_", None)
                    if explained_x_attr is not None:
                        explained_x = np.asarray(explained_x_attr, dtype=float)
                    elif T_all is not None:
                        raw_var = np.var(T_all, axis=0)
                        explained_x = raw_var / np.sum(raw_var) if np.sum(raw_var) > 0 else np.zeros_like(raw_var)
                    else:
                        # Fallback: equal share to avoid crashes (should rarely happen)
                        explained_x = np.ones(pls.x_loadings_.shape[1], dtype=float) / pls.x_loadings_.shape[1]

                    def _lv_index(name: str) -> int:
                        # Parse "PLS3" -> 2; be defensive
                        import re
                        m = re.search(r"\d+", str(name))
                        idx = int(m.group()) - 1 if m else 0
                        return max(0, min(idx, len(explained_x) - 1))

                    # -------------------- User Selects LV to View Loadings --------------------
                    st.subheader("📊 PLS Loadings")
                    selected_lv_loading = st.selectbox("Select PLS to view Loadings:", loadings.columns, key="pls_loadings_select")

                    lv_idx = _lv_index(selected_lv_loading)
                    var_pct = float(explained_x[lv_idx] * 100.0)

                    # Clean, consistent title line (like PCA)
                    st.markdown(
                        f"### 🔑 Loadings Plot for {selected_lv_loading} &nbsp;&nbsp;—&nbsp;&nbsp;**{var_pct:.2f}%** of total variance")

                    # ===== line plot of loading (with quick band annotations like your PCA block) =====
                    fig_load1 = go.Figure()
                    fig_load1.add_trace(go.Scatter(
                        x=selected_columns,
                        y=loadings[selected_lv_loading],
                        mode="lines",
                        line=dict(color="blue"),
                        name=f"{selected_lv_loading} Loadings"
                    ))
                    fig_load1.add_hline(y=0, line=dict(color="gray", dash="dash"))  # Zero reference line

                    # ✅ Add Band Annotations (same set as PCA)
                    band_annotations1 = {
                        1740: "C=O Stretch (Lipid)",
                        1650: "Amide I (Protein)",
                        1550: "Amide II (Protein)",
                        1338: "CH₂ (Collagen)",
                        1200: "C–O Stretch (Matrix)",
                        1115: "HPO₄²⁻ (Mineral)",
                        1060: "Sugar Ring C–O",
                        1030: "PO₄³⁻ (Mineral)",
                        875: "CO₃²⁻ Bending",
                        856: "C–S Proteoglycan",
                        7000: "O–H Stretch",
                        6688: "N–H Stretch",
                        5800: "CH₂ Lipid",
                        5200: "O–H Water",
                    }
                    try:
                        _lv_idx_for_title = int(str(selected_lv_loading).replace("LV", "")) - 1
                    except Exception:
                        _lv_idx_for_title = 0

                    try:
                        # use selected_range if present (from your global UI)
                        _lo, _hi = float(selected_range[0]), float(selected_range[1])
                    except Exception:
                        # infer a full range from numeric column names
                        wn_numeric_all = np.array([_num_first(c) for c in selected_columns], dtype=float)
                        mask_valid = np.isfinite(wn_numeric_all)
                        if mask_valid.any():
                            _lo, _hi = float(np.nanmin(wn_numeric_all[mask_valid])), float(
                                np.nanmax(wn_numeric_all[mask_valid]))
                        else:
                            _lo, _hi = -np.inf, np.inf

                    for wn, label in band_annotations1.items():
                        if _lo <= wn <= _hi:
                            fig_load1.add_vline(x=wn, line=dict(color="gray", width=1, dash="dash"))
                            fig_load1.add_annotation(
                                x=wn, y=loadings[selected_lv_loading].max(), text=label,
                                showarrow=True, arrowhead=1, yshift=10, font=dict(size=9)
                            )

                    fig_load1.update_layout(
                        title=f"PLS Loadings for {selected_lv_loading} ({(explained_x[_lv_idx_for_title] * 100.0 if _lv_idx_for_title < len(explained_x) else float('nan')):.2f}% X-Variance)",
                        xaxis_title="Wavenumber (cm⁻¹)",
                        yaxis_title="Loading Score",
                        height=400
                    )
                    st.plotly_chart(fig_load1)

                    ############################## [  📊 Variable Importance with Bond Annotations (PLS) ] ############################################################

                    import numpy as np
                    import pandas as pd
                    import streamlit as st

                    # --- Bond Annotation Mapping Table (same as your PCA section) ---
                    bond_annotations_table = [
                        (8500, "O–H Stretching and Bending (Water)"),
                        (7000, "O–H Stretching (Water)"),
                        (6688, "N–H Stretching (Protein/Collagen)"),
                        (5800, "CH₂ Stretching (Lipid)"),
                        (5200, "O–H Stretching and Bending (Water)"),
                        (4890, "N–H Bending (Protein/Collagen)"),
                        (4610, "C–H Stretching & Deformation (Protein/Collagen)"),
                        (4310, "Sugar Ring Vibrations (Proteoglycan)"),
                        ((3600, 3200), "O–H Stretching (Water/Hydroxyl)"),
                        ((3500, 3300), "N–H Stretching (Proteins)"),
                        ((3000, 2800), "C–H Stretching (Lipids, CH₂)"),
                        ((1750, 1650), "C=O Stretching (Proteins/Lipids)"),
                        (1550, "Amide II (Proteins)"),
                        (1338, "CH₂ Side Chain Bending (Collagen)"),
                        ((1100, 900), "PO₄³⁻ Stretching (Bone Mineral)"),
                        ((890, 850), "CO₃²⁻ Bending (Carbonate)"),
                        (1740, "C=O Stretching (Ester, Lipids)"),
                        (1650, "Amide I (Proteins)"),
                        (1630, "Water O–H Bending (Water)"),
                        ((1200, 1000), "C–O Stretching (Alcohols/Ethers)"),
                        (1115, "HPO₄²⁻ Stretching (Bone mineral)"),
                        (1060, "Sugar Ring C–O Stretch (Carbohydrates)"),
                        (1030, "PO₄³⁻ Stretching (Bone Mineral)"),
                        (875, "CO₃²⁻ Bending (Carbonates)"),
                        (856, "C–S Bending (Proteoglycans)"),
                        ((900, 800), "Aromatic C–H Bending (Fingerprint region)")
                    ]


                    def get_bond_annotation(wavenumber: float) -> str:
                        # Check ranges first
                        for entry in bond_annotations_table:
                            if isinstance(entry[0], tuple):
                                low, high = entry[0]
                                if high <= wavenumber <= low:
                                    return entry[1]
                        # Exact match
                        for entry in bond_annotations_table:
                            if not isinstance(entry[0], tuple) and entry[0] == int(wavenumber):
                                return entry[1]
                        # Nearest match if no exact/range match
                        all_wavenumbers = [val if isinstance(val, int) else (val[0] + val[1]) // 2 for val, _ in
                                           bond_annotations_table]
                        closest_idx = int(np.argmin([abs(wavenumber - wn) for wn in all_wavenumbers]))
                        return bond_annotations_table[closest_idx][1]


                    # --- Variable Importance by |loadings| (top 20) ---
                    variable_importance = loadings.abs()[selected_lv_loading].sort_values(ascending=False)
                    importance_df = pd.DataFrame(variable_importance.head(20)).reset_index()
                    importance_df.columns = ["Wavenumber (cm⁻¹)", "Importance"]


                    # Safely extract numeric value (handles labels like '1030/1240')
                    def extract_first_numeric(value):
                        s = str(value).strip()
                        if "/" in s:
                            s = s.split("/")[0].strip()
                        try:
                            return float(s)
                        except ValueError:
                            return None


                    importance_df["Numeric Wavenumber"] = importance_df["Wavenumber (cm⁻¹)"].apply(
                        extract_first_numeric)
                    importance_df = importance_df.dropna(subset=["Numeric Wavenumber"])
                    importance_df["Band Annotation"] = importance_df["Numeric Wavenumber"].apply(get_bond_annotation)

                    ############################## [ 📊 Variable Importance with Bond Annotations - PLS ] ############################################################

                    col1, col2 = st.columns([4, 1])
                    with col1:
                        st.markdown("### 📊 Variable Importance with Bond Annotations")
                    with col2:
                        var_importance_csv = importance_df.to_csv(index=False).encode("utf-8")
                        st.download_button(
                            "📥 Download CSV",
                            data=var_importance_csv,
                            file_name="variable_importance_with_bands.csv",
                            mime="text/csv",

                            key = "varimp_csv_btn"
                        )

                    # ===== Display table =====
                    st.dataframe(importance_df)

                    ##########################################[  ⭐ Most influential wavenumbers (by |loading|)]]################################################
                    # ========================= ⭐ Most influential wavenumbers (by |loading|) — PLS =========================

                    import numpy as np
                    import pandas as pd
                    import streamlit as st

                    # --- helpers (only defined if missing) ---
                    if "_num_first" not in globals():
                        def _num_first(val):
                            s = str(val).strip()
                            if "/" in s:
                                s = s.split("/")[0].strip()
                            try:
                                return float(s)
                            except Exception:
                                return np.nan

                    if "get_bond_annotation" not in globals():
                        bond_annotations_table = [
                            (8500, "O–H Stretching and Bending (Water)"),
                            (7000, "O–H Stretching (Water)"),
                            (6688, "N–H Stretching (Protein/Collagen)"),
                            (5800, "CH₂ Stretching (Lipid)"),
                            (5200, "O–H Stretching and Bending (Water)"),
                            (4890, "N–H Bending (Protein/Collagen)"),
                            (4610, "C–H Stretching & Deformation (Protein/Collagen)"),
                            (4310, "Sugar Ring Vibrations (Proteoglycan)"),
                            ((3600, 3200), "O–H Stretching (Water/Hydroxyl)"),
                            ((3500, 3300), "N–H Stretching (Proteins)"),
                            ((3000, 2800), "C–H Stretching (Lipids, CH₂)"),
                            ((1750, 1650), "C=O Stretching (Proteins/Lipids)"),
                            (1550, "Amide II (Proteins)"),
                            (1338, "CH₂ Side Chain Bending (Collagen)"),
                            ((1100, 900), "PO₄³⁻ Stretching (Bone Mineral)"),
                            ((890, 850), "CO₃²⁻ Bending (Carbonate)"),
                            (1740, "C=O Stretching (Ester, Lipids)"),
                            (1650, "Amide I (Proteins)"),
                            (1630, "Water O–H Bending (Water)"),
                            ((1200, 1000), "C–O Stretching (Alcohols/Ethers)"),
                            (1115, "HPO₄²⁻ Stretching (Bone mineral)"),
                            (1060, "Sugar Ring C–O Stretch (Carbohydrates)"),
                            (1030, "PO₄³⁻ Stretching (Bone Mineral)"),
                            (875, "CO₃²⁻ Bending (Carbonates)"),
                            (856, "C–S Bending (Proteoglycans)"),
                            ((900, 800), "Aromatic C–H Bending (Fingerprint region)")
                        ]


                        def get_bond_annotation(wavenumber: float) -> str:
                            for entry in bond_annotations_table:
                                if isinstance(entry[0], tuple):
                                    low, high = entry[0]
                                    if high <= wavenumber <= low:
                                        return entry[1]
                            for entry in bond_annotations_table:
                                if not isinstance(entry[0], tuple) and entry[0] == int(wavenumber):
                                    return entry[1]
                            mids = [v if isinstance(v, int) else (v[0] + v[1]) // 2 for v, _ in bond_annotations_table]
                            idx = int(np.argmin([abs(wavenumber - m) for m in mids]))
                            return bond_annotations_table[idx][1]

                    if "_favor_class" not in globals():
                        def _favor_class(value, class_pos, class_neg):
                            if class_pos is None or class_neg is None:
                                return ""
                            return f"favors **{class_pos}**" if value > 0 else f"favors **{class_neg}**"

                    # --- compute LV index + X-variance explained for title & summary ---
                    try:
                        lv_idx = int(str(selected_lv_loading).replace("PLS", "")) - 1
                    except Exception:
                        lv_idx = 0
                    try:
                        var_pct = float(explained_x[lv_idx] * 100.0)
                    except Exception:
                        var_pct = float("nan")

                    # --- numeric axis from feature labels ---
                    wn_labels = list(loadings.index)
                    wn_numeric = np.array([_num_first(v) for v in wn_labels], dtype=float)

                    # --- focus window from selected_range if present ---
                    try:
                        lo, hi = float(selected_range[0]), float(selected_range[1])
                        focus_mask = (wn_numeric >= min(lo, hi)) & (wn_numeric <= max(lo, hi))
                        if not np.any(focus_mask):
                            focus_mask = np.isfinite(wn_numeric)
                    except Exception:
                        focus_mask = np.isfinite(wn_numeric)

                    # --- loadings for selected LV ---
                    s_full = pd.to_numeric(loadings[selected_lv_loading], errors="coerce").values
                    s_focus = s_full[focus_mask]
                    wn_focus = wn_numeric[focus_mask]
                    lbl_focus = np.array(wn_labels, dtype=object)[focus_mask]

                    # --- determine class orientation (+/− side) from LV scores (if available) ---
                    class_pos = class_neg = None
                    try:
                        df_scores_pls = pd.DataFrame({"Label": np.asarray(y), "LV_Score": T[:, lv_idx]})
                        class_means_pls = df_scores_pls.groupby("Label", observed=True)["LV_Score"].mean().sort_values(
                            ascending=False)
                        if len(class_means_pls):
                            class_pos = class_means_pls.index[0]
                            class_neg = class_means_pls.index[-1]
                    except Exception:
                        pass

                    # --- build drivers_df (with Band Annotation + Interpretation filled) ---
                    rows = []
                    for lbl, wn, val in zip(lbl_focus, wn_focus, s_focus):
                        band = get_bond_annotation(float(wn)) if np.isfinite(wn) else ""
                        interp = _favor_class(float(val), class_pos, class_neg) if np.isfinite(val) else ""
                        rows.append({
                            "Wavenumber (cm⁻¹)": str(lbl),
                            "Numeric (cm⁻¹)": wn if np.isfinite(wn) else np.nan,
                            "Loading": float(val) if np.isfinite(val) else np.nan,
                            "|Loading|": float(abs(val)) if np.isfinite(val) else np.nan,
                            "Band Annotation": band,
                            "Interpretation": interp
                        })

                    drivers_df = (pd.DataFrame(rows)
                                  .dropna(subset=["Numeric (cm⁻¹)", "Loading"])
                                  .sort_values("|Loading|", ascending=False)
                                  .reset_index(drop=True))

                    # --- UI: Most influential table + CSV ---
                    col1, col2 = st.columns([4, 1])
                    with col1:
                        st.markdown("### ⭐ Most influential wavenumbers (by |loading|)")
                    with col2:
                        st.download_button(
                            "📥 Download CSV",
                            data=drivers_df[
                                ["Wavenumber (cm⁻¹)", "Loading", "|Loading|", "Band Annotation", "Interpretation"]
                            ].to_csv(index=False).encode("utf-8"),
                            file_name=f"most_influential_wavenumbers_{selected_lv_loading}.csv",
                            mime="text/csv"

                        )

                    st.dataframe(
                        drivers_df[
                            ["Wavenumber (cm⁻¹)", "Loading", "|Loading|", "Band Annotation", "Interpretation"]
                        ].style.format({"Loading": "{:.4f}", "|Loading|": "{:.4f}"})

                    )

                    ############################## [ 🧩 Auto-detected regions (aggregated contribution) — PLS ] ############################################################

                    # Window half-width for region detection (cm⁻¹)
                    half_width_cm = 25

                    def _region_for(center_cm1):
                        lo = center_cm1 - half_width_cm
                        hi = center_cm1 + half_width_cm
                        return lo, hi

                    regs = []
                    for _, r in drivers_df.dropna(subset=["Numeric (cm⁻¹)"]).iterrows():
                        c = float(r["Numeric (cm⁻¹)"])
                        lo_r, hi_r = _region_for(c)
                        regs.append([lo_r, hi_r, r["Wavenumber (cm⁻¹)"], r["Band Annotation"]])

                    regs.sort(key=lambda x: x[0])
                    merged = []
                    for lo_r, hi_r, lbl, ann in regs:
                        if not merged or lo_r > merged[-1][1]:
                            merged.append([lo_r, hi_r, [lbl], [ann]])
                        else:
                            merged[-1][1] = max(merged[-1][1], hi_r)
                            merged[-1][2].append(lbl)
                            merged[-1][3].append(ann)

                    # Compute region contributions using the focused loadings vector
                    reg_rows = []
                    for lo_r, hi_r, lbls, anns in merged:
                        m = (wn_focus >= lo_r) & (wn_focus <= hi_r)
                        if not np.any(m):
                            continue
                        net = float(np.nansum(s_focus[m]))
                        mass = float(np.nansum(np.abs(s_focus[m])))
                        lbl_text = ", ".join(map(str, lbls[:3])) + ("..." if len(lbls) > 3 else "")
                        ann_text = ", ".join(sorted(set(a for a in anns if a)))
                        reg_rows.append({
                            "Region (cm⁻¹)": f"{int(lo_r)}–{int(hi_r)}",
                            "|Σ loading|": mass,
                            "Net sign": "Positive" if net > 0 else ("Negative" if net < 0 else "≈0"),
                            "Peak labels": lbl_text,
                            "Likely bands": ann_text or "—",
                        })

                    contrib_df = (pd.DataFrame(reg_rows)
                                  .sort_values("|Σ loading|", ascending=False)
                                  .reset_index(drop=True))

                    ################## [ 🧩 Auto-detected regions (aggregated contribution) ] ###############
                    ###### removed script from here


                    ############################## [📝 Auto-summary - PLS ] ############################################################

                    pos_mass = float(np.nansum(np.abs(s_focus[s_focus > 0])))
                    neg_mass = float(np.nansum(np.abs(s_focus[s_focus < 0])))
                    contrast = ("Positive-weighted regions dominate"
                                if pos_mass > neg_mass else
                                ("Negative-weighted regions dominate" if neg_mass > pos_mass else "Balanced positive/negative weights"))

                    summary_lines = []
                    summary_lines.append(f"- **{selected_lv_loading}** explains **{var_pct:.2f}%** of X-variance.")
                    summary_lines.append(f"- **Contrast:** {contrast} in the current range.")
                    if class_pos and class_neg:
                        summary_lines.append(
                            f"- **Class context:** Higher LV scores (positive side) align with **{class_pos}**; "
                            f"lower scores (negative side) align with **{class_neg}**."
                        )
                    if len(contrib_df):
                        top_reg = contrib_df.iloc[0]
                        summary_lines.append(
                            f"- **Strongest region:** {top_reg['Region (cm⁻¹)']} "
                            f"({top_reg['Net sign']}, |Σloading|={top_reg['|Σ loading|']:.4f}); "
                            f"likely bands: {top_reg['Likely bands'] or '—'}."
                        )

                    head_n = min(6, len(drivers_df))
                    bullets = []
                    for _, r in drivers_df.head(head_n).iterrows():
                        side = "positive" if r["Loading"] > 0 else "negative"
                        interp = f" — {r['Interpretation']}" if r["Interpretation"] else ""
                        bullets.append(
                            f"  • {r['Wavenumber (cm⁻¹)']}: **{side}** loading ({r['Band Annotation']}){interp}")

                    st.markdown("### 📝 Auto-summary")

                    st.markdown("\n".join(summary_lines + ["- **Key drivers:**"] + bullets))
                    st.caption(
                        "Notes: (1) PLS loading sign is relative to LV orientation; interpretation uses class-wise mean scores. "
                        "(2) Region aggregation uses |loading| within ±half-width around top wavenumbers. "
                        "(3) Preprocessing (e.g., autoscale/derivative) affects which bands dominate."
                    )
                    ############################## [ 📊 Model Performance Summary (selected PLS) ] ############################################################
                    ############################################[  model performance summary (PLS) ]#########################################################################
                    st.markdown("### 📊 Model Performance Summary")

                    run_pls_clf = st.checkbox("📊 Model Performance Summary for selected LV", value=True,key="pls_run_selected_lv")

                    if run_pls_clf:
                        # --- choose 1+ models (all selected by default) ---
                        model_choices_pls = ["SVM", "Logistic Regression", "KNN", "Random Forest", "XGBoost"]
                        selected_models_pls = st.multiselect(
                            "Select classification models (trained on the selected LV score)",
                            options=model_choices_pls,
                            default=model_choices_pls,
                            key="pls_selected_lv_models"
                        )

                        # Parse selected LV index (e.g., "PLS3" -> 2)
                        try:
                            sel_lv_idx = int(str(selected_lv_loading).replace("PLS", "")) - 1
                        except Exception:
                            sel_lv_idx = 0

                        # --- single vs cumulative toggle (PLS) ---
                        mode_pls = st.radio(
                            "LV feature mode",
                            ["Single (selected only)", "Cumulative (PLS1..selected)"],
                            index=1,
                            key="pls_feature_mode"
                        )

                        # --- Build feature matrix from PLS scores ---
                        # Ensure we have PLS scores matrix T (n_samples x n_comp)
                        try:
                            T_scores = pls.x_scores_
                        except Exception:
                            # fallback: recompute from fitted pls
                            T_scores = pls.transform(X)  # assumes `pls` already fitted

                        if mode_pls == "Single (selected only)":
                            X_pls_feat = T_scores[:, sel_lv_idx].reshape(-1, 1)
                            lv_label = f"{selected_lv_loading}"
                        else:
                            X_pls_feat = T_scores[:, :sel_lv_idx + 1]
                            lv_label = f"PLS1–PLS{sel_lv_idx + 1}"

                        # --- Encode labels locally (use raw y; fallback to y_encoded if needed) ---
                        from sklearn.preprocessing import LabelEncoder

                        if "y" in globals():
                            y_labels_src = y
                        elif "y_encoded" in globals():
                            y_labels_src = y_encoded
                        else:
                            st.error("Targets `y` (labels) or `y_encoded` not found for classification.")
                            st.stop()

                        le_local_pls = LabelEncoder()
                        y_codes_pls = le_local_pls.fit_transform(
                            y_labels_src.values if hasattr(y_labels_src, "values") else y_labels_src
                        )

                        # --- Train/test split (stratified) ---
                        from sklearn.model_selection import train_test_split

                        X_train_pls, X_test_pls, y_train_pls, y_test_pls = train_test_split(
                            X_pls_feat, y_codes_pls, test_size=0.2, stratify=y_codes_pls, random_state=42
                        )

                        # --- Helper to build a model by name (same as PCA block) ---
                        from sklearn.svm import SVC
                        from sklearn.linear_model import LogisticRegression
                        from sklearn.neighbors import KNeighborsClassifier
                        from sklearn.ensemble import RandomForestClassifier
                        from sklearn.metrics import precision_score, recall_score, f1_score, classification_report, \
                            confusion_matrix
                        import numpy as np
                        import pandas as pd
                        import seaborn as sns
                        import matplotlib.pyplot as plt

                        try:
                            from xgboost import XGBClassifier

                            _has_xgb_pls = True
                        except Exception:
                            _has_xgb_pls = False


                        def _make_model_pls(name):
                            if name == "SVM":
                                return SVC()
                            if name == "Logistic Regression":
                                return LogisticRegression(max_iter=1000)
                            if name == "KNN":
                                return KNeighborsClassifier(n_neighbors=3)
                            if name == "Random Forest":
                                return RandomForestClassifier(n_estimators=100)
                            # XGBoost
                            if _has_xgb_pls:
                                return XGBClassifier(eval_metric="logloss")

                            # graceful fallback if xgboost not available
                            return RandomForestClassifier(n_estimators=100)


                        # --- Run all selected models ---
                        results_pls = []
                        for mdl_name in selected_models_pls:
                            clf_pls = _make_model_pls(mdl_name)
                            clf_pls.fit(X_train_pls, y_train_pls)
                            y_pred_pls = clf_pls.predict(X_test_pls)

                            # Metrics (weighted for multi-class safety)
                            acc_pls = clf_pls.score(X_test_pls, y_test_pls)
                            prec_pls = precision_score(y_test_pls, y_pred_pls, average="weighted", zero_division=0)
                            rec_pls = recall_score(y_test_pls, y_pred_pls, average="weighted", zero_division=0)
                            f1_pls = f1_score(y_test_pls, y_pred_pls, average="weighted", zero_division=0)

                            results_pls.append({
                                "Model": mdl_name,
                                "LV used": lv_label,  # mirrors PCA's "PC used"
                                "Accuracy": acc_pls,
                                "Precision (weighted)": prec_pls,
                                "Recall (weighted)": rec_pls,
                                "F1-Score (weighted)": f1_pls
                            })

                        # ===== 📊 Model Performance Summary (selected LV) + Download (one line) =====
                        col1, col2 = st.columns([4, 1])

                        with col1:
                            st.markdown("### 📊 Model Performance Summary (selected PLS)")

                        with col2:
                            st.download_button(
                                "📥 Download CSV",
                                data=pd.DataFrame(results_pls).to_csv(index=False).encode("utf-8"),
                                file_name="pls_selected_lv_model_metrics.csv",
                                mime="text/csv"

                            )

                        results_pls_df = pd.DataFrame(results_pls)
                        st.dataframe(
                            results_pls_df.style
                            .background_gradient(cmap="YlGnBu")
                            .format({
                                "Accuracy": "{:.3f}",
                                "Precision (weighted)": "{:.3f}",
                                "Recall (weighted)": "{:.3f}",
                                "F1-Score (weighted)": "{:.3f}"
                            })
                        )

                        # ---- Detailed results per model ----
                        for mdl_name in selected_models_pls:
                            st.markdown(f"### 🔎 Details: {mdl_name} on {lv_label}")
                            report_col_pls, matrix_col_pls = st.columns([2, 3])

                            clf_pls = _make_model_pls(mdl_name)
                            clf_pls.fit(X_train_pls, y_train_pls)
                            y_pred_pls = clf_pls.predict(X_test_pls)

                            with report_col_pls:
                                st.markdown("**Classification Report**")
                                report_dict_pls = classification_report(
                                    y_test_pls, y_pred_pls, target_names=list(le_local_pls.classes_), output_dict=True,
                                    zero_division=0
                                )
                                report_df_pls = pd.DataFrame(report_dict_pls).transpose()
                                styled_report_pls = report_df_pls.style.set_table_styles([
                                    {'selector': 'th', 'props': [('border', '1px solid black'), ('padding', '8px'),
                                                                 ('background', '#f2f2f2')]},
                                    {'selector': 'td', 'props': [('border', '1px solid black'), ('padding', '8px')]},
                                    {'selector': 'tr', 'props': [('border', '1px solid black')]}
                                ]).format("{:.2f}")
                                st.write(styled_report_pls)

                            with matrix_col_pls:
                                # Row-normalized confusion matrix -> percentages
                                cm_pls = confusion_matrix(y_test_pls, y_pred_pls,
                                                          labels=np.arange(len(le_local_pls.classes_)))
                                with np.errstate(divide="ignore", invalid="ignore"):
                                    cm_pct_pls = cm_pls.astype(float) / cm_pls.sum(axis=1, keepdims=True) * 100
                                cm_pct_pls = np.nan_to_num(cm_pct_pls)

                                annot_labels_pls = np.array([[f"{v:.1f}%" for v in row] for row in cm_pct_pls])

                                fig_cm_pls, ax_pls = plt.subplots()
                                sns.heatmap(
                                    cm_pct_pls,
                                    annot=annot_labels_pls,
                                    fmt="",
                                    cmap="Blues",
                                    vmin=0, vmax=100,
                                    xticklabels=list(le_local_pls.classes_),
                                    yticklabels=list(le_local_pls.classes_),
                                    cbar_kws={"label": "% of true class"},
                                    ax=ax_pls
                                )
                                ax_pls.set_xlabel("Predicted")
                                ax_pls.set_ylabel("True")
                                ax_pls.set_title(f"{mdl_name} — Confusion Matrix (%) on {lv_label}")
                                st.pyplot(fig_cm_pls)
                                plt.close(fig_cm_pls)

                    ############################## [  ] ############################################################
                    # ============================== 📥 Download PLS Full Workbook (editable loadings chart, 2D chart, 3D, summaries, all confusion matrices) ==============================

                    import io, numpy as np, pandas as pd, matplotlib.pyplot as plt, seaborn as sns
                    # ============================== One-file export for the entire "🔬 PLS: Partial Least Squares" ==============================

                    import io, numpy as np, pandas as pd, matplotlib.pyplot as plt, seaborn as sns
                    from sklearn.preprocessing import LabelEncoder
                    from sklearn.model_selection import train_test_split
                    from sklearn.metrics import classification_report, confusion_matrix
                    from sklearn.svm import SVC
                    from sklearn.linear_model import LogisticRegression
                    from sklearn.neighbors import KNeighborsClassifier
                    from sklearn.ensemble import RandomForestClassifier

                    try:
                        from xgboost import XGBClassifier

                        _has_xgb = True
                    except Exception:
                        _has_xgb = False


                    # ---------- helper: auto-fit Excel columns ----------
                    def _autosize_columns(ws, df, extra_pad=2, max_width=60):
                        for i, col in enumerate(df.columns):
                            series = df[col].astype(str)
                            max_len = max([len(str(col))] + [len(s) for s in series.tolist()])
                            ws.set_column(i, i, min(max_len + extra_pad, max_width))


                    def _autosize_range(ws, first_col, last_col, rows, headers=None, extra_pad=2, max_width=60):
                        # Auto-fit arbitrary cell blocks (for Model_Metrics where we write cells)
                        for c in range(first_col, last_col + 1):
                            values = []
                            if headers and c - first_col < len(headers):
                                values.append(str(headers[c - first_col]))
                            for r in rows:
                                v = r.get(c, "")
                                values.append("" if v is None else str(v))
                            width = min(max(len(x) for x in values) + extra_pad, max_width) if values else 10
                            ws.set_column(c, c, width)


                    # ---------- (A) Rebuild model results (AUTO-RUN; independent of any checkbox) ----------
                    # Use cumulative PLS1..selected for export
                    _sel_lv_idx = int(str(selected_lv_loading).replace("PLS", "")) - 1

                    # Get PLS scores (T); if not present, compute via pls.transform(X_num)
                    try:
                        T_all = pls.x_scores_
                    except Exception:
                        # Fallback if scores not cached; assumes X_num (numeric X) exists, else try X
                        X_for_tr = X_num if 'X_num' in globals() else (
                            X.values if hasattr(X, "values") else np.asarray(X, dtype=float))
                        T_all = pls.transform(X_for_tr)

                    X_pls_export = T_all[:, :_sel_lv_idx + 1]  # cumulative PLS1..selected
                    lv_label_export = f"PLS1–PLS{_sel_lv_idx + 1}"

                    # Labels
                    le_export = LabelEncoder()
                    y_codes_export = le_export.fit_transform(y if not hasattr(y, "values") else y.values)

                    # Split
                    X_train_e, X_test_e, y_train_e, y_test_e = train_test_split(
                        X_pls_export, y_codes_export, test_size=0.2, stratify=y_codes_export, random_state=42
                    )

                    # Models
                    models = {
                        "SVM": SVC(),
                        "Logistic Regression": LogisticRegression(max_iter=1000),
                        "KNN": KNeighborsClassifier(n_neighbors=3),
                        "Random Forest": RandomForestClassifier(n_estimators=100),
                    }
                    if _has_xgb:
                        models["XGBoost"] = XGBClassifier(eval_metric="logloss")


                    results_rows = []
                    confmats_counts = {}
                    confmats_pct = {}
                    confmat_heatmap_png = {}

                    for name, clf in models.items():
                        clf.fit(X_train_e, y_train_e)
                        y_pred = clf.predict(X_test_e)

                        acc = clf.score(X_test_e, y_test_e)
                        rep = classification_report(
                            y_test_e, y_pred, target_names=list(le_export.classes_), output_dict=True, zero_division=0
                        )
                        prec = rep["weighted avg"]["precision"]
                        rec = rep["weighted avg"]["recall"]
                        f1 = rep["weighted avg"]["f1-score"]

                        results_rows.append({
                            "Model": name,
                            "LV used": lv_label_export,
                            "Accuracy": acc,
                            "Precision (weighted)": prec,
                            "Recall (weighted)": rec,
                            "F1-Score (weighted)": f1
                        })

                        # Confusion matrices
                        cm = confusion_matrix(y_test_e, y_pred, labels=np.arange(len(le_export.classes_)))
                        confmats_counts[name] = pd.DataFrame(cm, index=list(le_export.classes_),
                                                             columns=list(le_export.classes_))

                        with np.errstate(divide="ignore", invalid="ignore"):
                            cm_pct = cm.astype(float) / cm.sum(axis=1, keepdims=True) * 100.0
                        cm_pct = np.nan_to_num(cm_pct)
                        confmats_pct[name] = pd.DataFrame(cm_pct, index=list(le_export.classes_),
                                                          columns=list(le_export.classes_))

                        # Heatmap PNG (in-memory)
                        fig, ax = plt.subplots(figsize=(5, 4))
                        sns.heatmap(cm_pct, annot=True, fmt=".1f", cmap="Blues",
                                    xticklabels=list(le_export.classes_), yticklabels=list(le_export.classes_),
                                    cbar=True, ax=ax, vmin=0, vmax=100)
                        ax.set_xlabel("Predicted");
                        ax.set_ylabel("True")
                        ax.set_title(f"{name} — Confusion Matrix (%) on {lv_label_export}")
                        _png = io.BytesIO()
                        fig.tight_layout()
                        fig.savefig(_png, format="png", dpi=180)
                        plt.close(fig)
                        _png.seek(0)
                        confmat_heatmap_png[name] = _png

                    results_export_df = pd.DataFrame(results_rows)

                    # ---------- (B) Loadings PNG for 'Loadings Graph' (cleaner visual) ----------
                    try:
                        _lv_num = int(str(selected_lv_loading).replace("PLS", ""))
                        _var_pct = float(
                            explained_x[_lv_num - 1]) * 100.0  # explained_x should be PLS X-variance ratio per LV
                    except Exception:
                        _var_pct = np.nan

                    _series_vals = pd.to_numeric(loadings[selected_lv_loading], errors="coerce").values
                    _x_idx = np.arange(len(_series_vals))
                    figL, axL = plt.subplots(figsize=(11, 3.2))
                    axL.plot(_x_idx, _series_vals, linewidth=2.8)
                    axL.axhline(0.0, linestyle="--", color="gray", linewidth=1)
                    axL.set_title(f"{selected_lv_loading} Loadings ({_var_pct:.2f}% X-Variance)")
                    axL.set_xlabel("Index (wavenumber order)")
                    axL.set_ylabel("Loading")
                    axL.grid(alpha=0.15)
                    _loadings_png = io.BytesIO()
                    figL.tight_layout()
                    figL.savefig(_loadings_png, format="png", dpi=180)
                    plt.close(figL)
                    _loadings_png.seek(0)

                    # ---------- (C) Data for 2D/3D projections ----------
                    # Build a scores DataFrame with labels for export
                    labels_arr = y if not hasattr(y, "values") else y.values
                    pls_cols = [f"PLS{i + 1}" for i in range(T_all.shape[1])]
                    df_pls_all = pd.DataFrame(T_all, columns=pls_cols)
                    df_pls_all["Label"] = labels_arr

                    # Default axis names (you can pass your UI selections if you have them)
                    lv_x = pls_cols[0] if len(pls_cols) > 0 else "PLS1"
                    lv_y = pls_cols[1] if len(pls_cols) > 1 else lv_x
                    lv_z = pls_cols[2] if len(pls_cols) > 2 else None

                    proj2d_df = df_pls_all[[lv_x, lv_y, "Label"]].copy()
                    if lv_z and lv_z in df_pls_all.columns:
                        proj3d_df = df_pls_all[[lv_x, lv_y, lv_z, "Label"]].copy()
                    else:
                        proj3d_df = pd.DataFrame(columns=[lv_x, lv_y, "Z", "Label"])

                    # Optional PNG for 3D visualization (Excel has no native 3D scatter)
                    try:
                        from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

                        if lv_z and lv_z in df_pls_all.columns:
                            fig3, ax3 = plt.subplots(subplot_kw={"projection": "3d"}, figsize=(6, 5))
                            for lab in proj3d_df["Label"].unique():
                                d = proj3d_df[proj3d_df["Label"] == lab]
                                ax3.scatter(d[lv_x], d[lv_y], d[lv_z], label=str(lab), s=16)
                            ax3.set_xlabel(lv_x);
                            ax3.set_ylabel(lv_y);
                            ax3.set_zlabel(lv_z)
                            ax3.set_title("PLS 3D Visualization")
                            ax3.legend(loc="best", fontsize=8)
                            _pls3d_png = io.BytesIO()
                            fig3.tight_layout()
                            fig3.savefig(_pls3d_png, format="png", dpi=180)
                            plt.close(fig3)
                            _pls3d_png.seek(0)
                        else:
                            _pls3d_png = None
                    except Exception:
                        _pls3d_png = None

                    # ---------- (D) Build Excel with all pages & editable charts ----------
                    buffer_full = io.BytesIO()
                    with pd.ExcelWriter(buffer_full, engine="xlsxwriter") as writer:
                        wb = writer.book

                        # --- PLS Summary (if you built one) ---
                        try:
                            pls_calcv_df.to_excel(writer, sheet_name="PLS Summary", index=False)
                            _autosize_columns(writer.sheets["PLS Summary"], pls_calcv_df)
                        except Exception:
                            pass

                        # --- Variable Importance / Drivers / Regions ---
                        try:
                            importance_df.to_excel(writer, sheet_name="Variable Importance", index=False)
                            _autosize_columns(writer.sheets["Variable Importance"], importance_df)
                        except Exception:
                            pass

                        try:
                            _drivers_view = drivers_df[
                                ["Wavenumber (cm⁻¹)", "Loading", "|Loading|", "Band Annotation", "Interpretation"]]
                            _drivers_view.to_excel(writer, sheet_name="Most Influential", index=False)
                            _autosize_columns(writer.sheets["Most Influential"], _drivers_view)
                        except Exception:
                            pass

                        # try:
                        #     contrib_df.to_excel(writer, sheet_name="Regions", index=False)
                        #     _autosize_columns(writer.sheets["Regions"], contrib_df)
                        # except Exception:
                        #     pass

                        # --- 📝 Auto-summary (PLS) -> Excel sheet ---
                        try:
                            # Fallbacks so the sheet is always populated
                            if "summary_lines" not in globals() or not isinstance(summary_lines, list) or len(
                                    summary_lines) == 0:
                                try:
                                    _lv_num = int(str(selected_lv_loading).replace("PLS", ""))
                                except Exception:
                                    _lv_num = 1
                                try:
                                    _var_pct_safe = float(explained_x[_lv_num - 1]) * 100.0
                                except Exception:
                                    _var_pct_safe = float("nan")
                                summary_lines = [
                                    f"- **{selected_lv_loading}** explains **{_var_pct_safe:.2f}%** of X-variance."]

                            if "bullets" not in globals() or not isinstance(bullets, list):
                                bullets = []

                            auto_df_pls = pd.DataFrame({"Summary": summary_lines + ["- **Key drivers:**"] + bullets})
                            auto_df_pls.to_excel(writer, sheet_name="📝 Auto-summary", index=False)

                            # Auto-fit columns if helper is available
                            try:
                                _autosize_columns(writer.sheets["📝 Auto-summary"], auto_df_pls)
                            except Exception:
                                pass
                        except Exception:
                            # Keep exporter resilient
                            pass

                        # --- Loadings (table + Excel-editable chart) ---
                        try:
                            def _num_first(val):
                                s = str(val).strip()
                                if "/" in s:
                                    s = s.split("/")[0].strip()
                                try:
                                    return float(s)
                                except Exception:
                                    return np.nan


                            load_tbl = pd.DataFrame({
                                "Wavenumber (cm⁻¹)": list(loadings.index),
                                "Numeric Wavenumber": [_num_first(v) for v in loadings.index],
                                "Index": np.arange(len(loadings.index)),
                                "Loading": pd.to_numeric(loadings[selected_lv_loading], errors="coerce").values
                            })

                            load_tbl.to_excel(writer, sheet_name="Loadings", index=False)
                            ws_load = writer.sheets["Loadings"]
                            _autosize_columns(ws_load, load_tbl)

                            # Editable line chart (cleaner): thicker stroke, legend bottom
                            chart = wb.add_chart({"type": "line"})
                            n_rows = len(load_tbl)
                            chart.add_series({
                                "name": f"{selected_lv_loading}",
                                "categories": ["Loadings", 1, 1, n_rows, 1],  # Numeric Wavenumber
                                "values": ["Loadings", 1, 3, n_rows, 3],  # Loading
                                "line": {"width": 2.25},
                            })
                            chart.set_title({"name": f"Loadings — {selected_lv_loading} ({_var_pct:.2f}% X-var)"})
                            chart.set_x_axis({"name": "Wavenumber (cm⁻¹)"})
                            chart.set_y_axis({"name": "Loading"})
                            chart.set_legend({"position": "bottom"})
                            ws_load.insert_chart(n_rows + 3, 0, chart)
                        except Exception:
                            pass

                        # --- 📝 Auto-summary (if available) ---
                        try:
                            auto_lines = []
                            auto_lines.extend(summary)
                            auto_lines.append("Key drivers:")
                            auto_lines.extend(bullets)
                            auto_df = pd.DataFrame({"Summary": auto_lines})
                            auto_df.to_excel(writer, sheet_name="📝 Auto-summary", index=False)
                            _autosize_columns(writer.sheets["📝 Auto-summary"], auto_df)
                        except Exception:
                            pass

                        # --- 📉 PLS 2D Projection (data + editable scatter chart; markers only, no lines) ---
                        try:
                            sheet2d = "📉 PLS 2D Projection"
                            proj2d_df.to_excel(writer, sheet_name=sheet2d, index=False)
                            ws2d = writer.sheets[sheet2d]
                            _autosize_columns(ws2d, proj2d_df)

                            chart2d = wb.add_chart({"type": "scatter"})
                            cur_row = len(proj2d_df) + 3
                            for lab in proj2d_df["Label"].unique():
                                sub = proj2d_df[proj2d_df["Label"] == lab][[lv_x, lv_y]].reset_index(drop=True)
                                if sub.empty:
                                    continue
                                ws2d.write(cur_row, 0, f"{lab} — {lv_x}")
                                ws2d.write(cur_row, 1, f"{lab} — {lv_y}")
                                ws2d.write_column(cur_row + 1, 0, sub[lv_x].tolist())
                                ws2d.write_column(cur_row + 1, 1, sub[lv_y].tolist())

                                s_start, s_end = cur_row + 1, cur_row + len(sub)
                                chart2d.add_series({
                                    "name": str(lab),
                                    "categories": [sheet2d, s_start, 0, s_end, 0],
                                    "values": [sheet2d, s_start, 1, s_end, 1],
                                    "marker": {"type": "circle", "size": 5},
                                    "line": {"none": True},
                                })
                                cur_row = s_end + 2

                            chart2d.set_title({"name": f"PLS 2D: {lv_x} vs {lv_y}"})
                            chart2d.set_x_axis({"name": lv_x})
                            chart2d.set_y_axis({"name": lv_y})
                            chart2d.set_legend({"position": "bottom"})
                            ws2d.insert_chart(1, 4, chart2d)
                        except Exception:
                            pass

                        # --- 🧭 PLS 3D Visualization (data + PNG) ---
                        try:
                            sheet3d = "🧭 PLS 3D Visualization"
                            proj3d_df.to_excel(writer, sheet_name=sheet3d, index=False)
                            ws3d = writer.sheets[sheet3d]
                            _autosize_columns(ws3d, proj3d_df)
                            if _pls3d_png is not None:
                                ws3d.insert_image(1, 5, "pls3d.png", {"image_data": _pls3d_png})
                        except Exception:
                            pass

                        # --- Model_Metrics (summary + all confusion matrices + heatmaps on one sheet) ---
                        wsM = wb.add_worksheet("Model_Metrics")

                        headers = list(results_export_df.columns)
                        for c, h in enumerate(headers):
                            wsM.write(0, c, h)
                        for r, row in results_export_df.iterrows():
                            for c, h in enumerate(headers):
                                wsM.write(r + 1, c, row[h])
                        _autosize_columns(wsM, results_export_df)

                        start_row = len(results_export_df) + 3

                        for name in models.keys():
                            wsM.write(start_row, 0, f"{name} — Confusion Matrix (counts)")
                            cnt = confmats_counts[name]
                            for j, col in enumerate([""] + cnt.columns.tolist()):
                                wsM.write(start_row + 1, j, col)
                            for i, idx in enumerate(cnt.index.tolist()):
                                wsM.write(start_row + 2 + i, 0, idx)
                                for j, col in enumerate(cnt.columns.tolist()):
                                    wsM.write(start_row + 2 + i, 1 + j, int(cnt.loc[idx, col]))
                            _autosize_range(
                                wsM,
                                first_col=0,
                                last_col=len(cnt.columns),
                                rows=[{k: None for k in range(len(cnt.columns) + 1)} for _ in
                                      range(len(cnt.index) + 2)],
                                headers=[""] + cnt.columns.tolist()
                            )
                            start_row += len(cnt) + 3

                            wsM.write(start_row, 0, f"{name} — Confusion Matrix (%)")
                            pct = confmats_pct[name]
                            for j, col in enumerate([""] + pct.columns.tolist()):
                                wsM.write(start_row + 1, j, col)
                            for i, idx in enumerate(pct.index.tolist()):
                                wsM.write(start_row + 2 + i, 0, idx)
                                for j, col in enumerate(pct.columns.tolist()):
                                    wsM.write_number(start_row + 2 + i, 1 + j, float(pct.loc[idx, col]))
                            img_col = pct.shape[1] + 3
                            wsM.insert_image(start_row + 1, img_col, f"{name}_cm.png",
                                             {"image_data": confmat_heatmap_png[name], "x_scale": 1.0, "y_scale": 1.0})
                            start_row += len(pct) + 10

                    buffer_full.seek(0)

                    # ================= Download Complete PLS Analysis =================
                    colA, colB = st.columns([4, 1])

                    with colA:
                        st.download_button(
                            "📥 Download PLS Full Workbook (editable loadings chart, 2D chart, 3D, summaries, all confusion matrices)",
                            data=buffer_full.getvalue(),
                            file_name="pls_full_export.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                    with colB:
                        st.empty()

            #####################################################################################################################
            ##############    📊 PLS-DA: Partial Least Squares Discriminant Analysis

            if show_plsda:

                # ============================
                # 📊 PLS-DA: Partial Least Squares Discriminant Analysis
                # ============================

                col1, col2 = st.columns([4, 1])  # 4:1 width ratio (adjust as needed)

                with col1:
                    st.subheader("📊 PLS-DA: Partial Least Squares Discriminant Analysis")

                with st.expander(
                        "PLS-DA is a supervised projection method that maximizes separation across labeled groups. This panel provides a summary table, selectable 2D/3D projections, component-wise loadings with bond annotations, and VIP with bond annotations.",
                        expanded=True
                ):

                    ############################################ [ PLS-DA Analysis ] ##############################################################################
                    import numpy as np
                    import pandas as pd
                    import streamlit as st
                    from sklearn.cross_decomposition import PLSRegression
                    from sklearn.model_selection import StratifiedKFold, KFold
                    from sklearn.preprocessing import LabelEncoder
                    from sklearn.metrics import r2_score

                    # ---------- Preprocessing selection (combined) ----------
                    PREPROCESSING_OPTIONS = ["Mean Center", "Autoscale", "Smoothing", "Normalization", "Baseline",
                                             "Second Derivative"]
                    x_preproc_plsda = st.multiselect(
                        "Select Preprocessing Techniques for PLS-DA (applied together, in order)",
                        options=PREPROCESSING_OPTIONS,
                        default=["Autoscale"],
                        key="plsda_preproc"
                    )

                    # ========== Controls (give every widget a unique key) ==========
                    max_plsda = max(2, min(10, X.shape[1]))

                    # Slider: number of PLS-DA components
                    n_comp_da = st.slider(
                        "Number of PLS-DA components (LVs)",
                        2,
                        max_plsda,
                        value=min(5, max_plsda),
                        step=1,
                        key="plsda_n_comp"
                    )

                    cv_splits_plsda = st.slider("CV Splits (PLS-DA)", 3, 15, value=10, step=1, key="plsda_cv_splits")
                    conf_limit_plsda = st.selectbox("Confidence Limit (PLS-DA)", [0.90, 0.95, 0.99], index=1,
                                                    key="plsda_conf_limit")
                    lv_list_plsda = st.multiselect(
                        "Ncomp/LVs to evaluate (PLS-DA)",
                        options=list(range(2, max_plsda + 1)),
                        default=[2, 3],
                        key="plsda_lv_list"
                    )

                    # ===================== helpers: Savitzky–Golay =====================
                    from scipy.signal import savgol_filter


                    def _savgol_params(n_feat: int, window: int, poly: int, deriv: int):
                        w_max_odd = n_feat if (n_feat % 2 == 1) else (n_feat - 1)
                        if w_max_odd < 3:
                            return None
                        w = int(window)
                        if w % 2 == 0: w -= 1
                        if w < 3: w = 3
                        if w > w_max_odd: w = w_max_odd
                        p = int(poly)
                        p = max(p, deriv, 1)
                        if p >= w: p = w - 1
                        return w, p


                    def smooth_savgol(X, window=7, poly=3):
                        Xn = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                        Xn = Xn.copy()
                        params = _savgol_params(Xn.shape[1], window, poly, deriv=0)
                        if params is None: return Xn
                        w, p = params
                        return savgol_filter(Xn, window_length=w, polyorder=p, deriv=0, axis=1, mode="interp")

                    def second_derivative_savgol(X, window=7, poly=3):
                        Xn = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                        Xn = Xn.copy()
                        params = _savgol_params(Xn.shape[1], window, poly, deriv=2)
                        if params is None: return Xn
                        w, p = params
                        return savgol_filter(Xn, window_length=w, polyorder=p, deriv=2, axis=1, mode="interp")

                    # ===================== helpers: Baseline (ALS, sparse + fast) =====================
                    from scipy.sparse import diags, eye
                    from scipy.sparse.linalg import spsolve


                    def als_baseline_sparse(y, lam=1e5, p=0.001, niter=10, eps=1e-12):
                        y = np.asarray(y, dtype=float).ravel()
                        L = y.size
                        if L < 3:
                            return np.zeros_like(y)
                        D = diags([1, -2, 1], [0, 1, 2], shape=(L - 2, L))
                        DTD = lam * (D.T @ D)
                        w = np.ones(L)
                        I = eye(L) * eps
                        for _ in range(niter):
                            W = diags(w, 0, shape=(L, L))
                            z = spsolve(W + DTD + I, w * y)
                            w = p * (y > z) + (1 - p) * (y < z)
                        return z

                    # ===================== helpers: leakage-safe MC/Autoscale =====================
                    def fit_center_scale_stats(X_arr, need_mean: bool, need_std: bool):
                        mu = sd = None
                        if need_mean or need_std:
                            mu = X_arr.mean(axis=0, keepdims=True)
                        if need_std:
                            sd = X_arr.std(axis=0, keepdims=True)
                            sd = sd.copy();
                            sd[sd == 0] = 1.0
                        return mu, sd

                    def apply_preprocessing_chain(
                            X,
                            ops, *,
                            smooth_window=7, smooth_poly=3,
                            normalization_mode="l2",
                            baseline_lam=100_000.0, baseline_p=0.001,
                            mean_=None, std_=None
                    ):
                        Xn = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                        Xn = Xn.copy()

                        # 1) Baseline
                        if "Baseline" in ops:
                            Xcorr = np.empty_like(Xn)
                            for i in range(Xn.shape[0]):
                                bl = als_baseline_sparse(Xn[i, :], lam=baseline_lam, p=baseline_p)
                                Xcorr[i, :] = Xn[i, :] - bl
                            Xn = Xcorr

                        # 2) Smoothing / Second Derivative
                        if "Smoothing" in ops and "Second Derivative" in ops:
                            Xn = second_derivative_savgol(Xn, window=smooth_window, poly=smooth_poly)
                        elif "Smoothing" in ops:
                            Xn = smooth_savgol(Xn, window=smooth_window, poly=smooth_poly)
                        elif "Second Derivative" in ops:
                            Xn = second_derivative_savgol(Xn, window=smooth_window, poly=smooth_poly)

                        # 3) Mean Center / Autoscale
                        if "Mean Center" in ops or "Autoscale" in ops:
                            mu = mean_ if mean_ is not None else Xn.mean(axis=0, keepdims=True)
                            Xn = Xn - mu
                        if "Autoscale" in ops:
                            sd = std_ if std_ is not None else Xn.std(axis=0, keepdims=True)
                            sd = sd.copy();
                            sd[sd == 0] = 1.0
                            Xn = Xn / sd

                        # 4) Normalization (row-wise)
                        if "Normalization" in ops:
                            mode = normalization_mode.lower()
                            if mode == "l2":
                                denom = np.linalg.norm(Xn, axis=1, keepdims=True)
                            elif mode == "l1":
                                denom = np.sum(np.abs(Xn), axis=1, keepdims=True)
                            else:
                                denom = np.max(np.abs(Xn), axis=1, keepdims=True)
                            denom[denom == 0] = 1.0
                            Xn = Xn / denom

                        return Xn

                    # ===================== helpers: labels → one-hot & metrics =====================
                    def _labels_and_onehot():
                        if "y_encoded" in globals():
                            y_raw = globals()["y_encoded"]
                        elif "y" in globals():
                            y_raw = globals()["y"]
                        else:
                            st.error("PLS-DA requires a target vector `y` or `y_encoded` to be defined.")
                            raise RuntimeError("Missing y")

                        y_arr = y_raw.values if hasattr(y_raw, "values") else y_raw
                        y_arr = np.asarray(y_arr)
                        # integer-encode if needed
                        if np.issubdtype(y_arr.dtype, np.integer):
                            y_labels = y_arr.astype(int)
                        else:
                            y_labels = LabelEncoder().fit_transform(y_arr)
                        n_classes = int(np.max(y_labels)) + 1
                        Y_onehot = np.eye(n_classes)[y_labels]
                        return y_labels, Y_onehot

                    def _rmse_r2_bias_matrix(Y_true, Y_pred):
                        Y_true = np.asarray(Y_true, dtype=float)
                        Y_pred = np.asarray(Y_pred, dtype=float)
                        if Y_true.ndim == 1: Y_true = Y_true.reshape(-1, 1)
                        if Y_pred.ndim == 1: Y_pred = Y_pred.reshape(-1, 1)
                        rmse = float(np.sqrt(np.mean((Y_true - Y_pred) ** 2)))
                        r2_vals = []
                        for j in range(Y_true.shape[1]):
                            try:
                                r2_vals.append(r2_score(Y_true[:, j], Y_pred[:, j]))
                            except Exception:
                                r2_vals.append(np.nan)
                        r2_mean = float(np.nanmean(r2_vals))
                        bias = float(np.mean(Y_pred - Y_true))
                        return rmse, r2_mean, bias

                    # ===================== CAL & CV (PLS-DA with PLS-style metrics) =====================
                    try:
                        y_labels, Y_onehot = _labels_and_onehot()
                    except RuntimeError:
                        st.stop()

                    rows_plsda = []
                    model_id = 1
                    preproc_label_plsda = ", ".join(x_preproc_plsda) if x_preproc_plsda else "None"
                    X_size_da = f"{X.shape[0]} x {X.shape[1]}"
                    Y_size_da = f"{Y_onehot.shape[0]} x {Y_onehot.shape[1]}"

                    # Preprocess ALL data for CAL pass
                    X_all = apply_preprocessing_chain(
                        X, x_preproc_plsda,
                        smooth_window=7, smooth_poly=3,
                        baseline_lam=100_000.0, baseline_p=0.001
                    )

                    for ncomp in lv_list_plsda:
                        # ===== CAL =====
                        pls_all = PLSRegression(n_components=ncomp)
                        pls_all.fit(X_all, Y_onehot)
                        Yhat_all = pls_all.predict(X_all)
                        rmsec, r2c, bias_c = _rmse_r2_bias_matrix(Y_onehot, Yhat_all)

                        # ===== CV (leakage-safe preprocessing per fold) =====
                        try:
                            fold_iter = StratifiedKFold(n_splits=cv_splits_plsda, shuffle=True, random_state=42).split(
                                X, y_labels)
                        except Exception:
                            fold_iter = KFold(n_splits=cv_splits_plsda, shuffle=True, random_state=42).split(X)

                        Y_pred_cv = np.zeros_like(Y_onehot, dtype=float)

                        for tr_idx, te_idx in fold_iter:
                            X_tr = X.iloc[tr_idx].values.astype(float)
                            X_te = X.iloc[te_idx].values.astype(float)
                            Y_tr = Y_onehot[tr_idx, :]

                            need_mean = ("Mean Center" in x_preproc_plsda) or ("Autoscale" in x_preproc_plsda)
                            need_std = ("Autoscale" in x_preproc_plsda)
                            mu_tr, sd_tr = fit_center_scale_stats(X_tr, need_mean, need_std)

                            X_tr_s = apply_preprocessing_chain(
                                X_tr, x_preproc_plsda,
                                smooth_window=7, smooth_poly=3,
                                baseline_lam=100_000.0, baseline_p=0.001,
                                mean_=mu_tr, std_=sd_tr
                            )
                            X_te_s = apply_preprocessing_chain(
                                X_te, x_preproc_plsda,
                                smooth_window=7, smooth_poly=3,
                                baseline_lam=100_000.0, baseline_p=0.001,
                                mean_=mu_tr, std_=sd_tr
                            )

                            pls_cv = PLSRegression(n_components=ncomp)
                            pls_cv.fit(X_tr_s, Y_tr)
                            Y_pred_cv[te_idx, :] = pls_cv.predict(X_te_s)

                        rmsecv, r2cv, bias_cv = _rmse_r2_bias_matrix(Y_onehot, Y_pred_cv)
                        rmse_ratio = rmsecv / rmsec if rmsec > 0 else float("nan")

                        rows_plsda.append({
                            "CV Splits": cv_splits_plsda,
                            "Confidence Limit": conf_limit_plsda,
                            "RMSE Ratio (RMSECV/RMSEC)": rmse_ratio,
                            "RMSEC (Cal)": rmsec,
                            "RMSECV (CV)": rmsecv,
                            "Model Name": f"Model {model_id}",
                            "Ncomp/LVs": ncomp,
                            "X-Preprocessing": preproc_label_plsda,
                            "X Include Size": X_size_da,
                            "Y Include Size": Y_size_da,
                            "R2C (Cal)": r2c,
                            "R2CV (CV)": r2cv,
                            "Bias": bias_c,
                            "Bias (CV)": bias_cv
                        })
                        model_id += 1

                    # ---------------- Table + download ----------------
                    plsda_calcv_df = pd.DataFrame(rows_plsda, columns=[
                        "CV Splits", "Confidence Limit", "RMSE Ratio (RMSECV/RMSEC)",
                        "RMSEC (Cal)", "RMSECV (CV)", "Model Name", "Ncomp/LVs",
                        "X-Preprocessing", "X Include Size", "Y Include Size",
                        "R2C (Cal)", "R2CV (CV)", "Bias", "Bias (CV)"
                    ])

                    fmt_cols_da = ["RMSE Ratio (RMSECV/RMSEC)", "RMSEC (Cal)", "RMSECV (CV)", "R2C (Cal)", "R2CV (CV)",
                                   "Bias", "Bias (CV)"]

                    st.markdown("### 📋 PLS-DA Analysis Result Summary (PLS-style metrics)")
                    st.dataframe(
                        plsda_calcv_df.style
                        .format({c: "{:.4f}" for c in fmt_cols_da if c in plsda_calcv_df.columns})
                        .background_gradient(
                            subset=[c for c in ["R2CV (CV)", "R2C (Cal)"] if c in plsda_calcv_df.columns], cmap="YlGn")
                        .background_gradient(
                            subset=[c for c in ["RMSECV (CV)", "RMSEC (Cal)"] if c in plsda_calcv_df.columns],
                            cmap="YlOrRd_r")

                    )

                    st.download_button(
                        "📥 Download PLS-DA Summary (CSV)",
                        data=plsda_calcv_df.to_csv(index=False).encode("utf-8"),
                        file_name="plsda_summary_table.csv",
                        mime="text/csv",
                        key="plsda_download"
                    )

                    ################################################################################################################################################

                    # ===== Prepare X, y (one-hot) =====
                    scaler_da = StandardScaler()
                    X_scaled_da = scaler_da.fit_transform(X)

                    le_da = LabelEncoder()
                    y_codes_all = le_da.fit_transform(y)
                    y_onehot = pd.get_dummies(y_codes_all)  # (n_samples, n_classes)

                    # ===== Fit PLS-DA (PLSRegression with one-hot Y) =====
                    pls_da = PLSRegression(n_components=n_comp_da)
                    pls_da.fit(X_scaled_da, y_onehot.values)

                    # Matrices
                    T_da = pls_da.x_scores_  # (n_samples, n_comp_da)
                    U_da = pls_da.y_scores_  # (n_samples, n_comp_da)
                    P_da = pls_da.x_loadings_  # (n_features, n_comp_da)
                    W_da = pls_da.x_weights_  # (n_features, n_comp_da)
                    Q_da = pls_da.y_loadings_  # (n_classes, n_comp_da)

                    # ===== Axis Selection for 2D/3D =====
                    all_da_labels = [f"PLSDA{i + 1}" for i in range(n_comp_da)]
                    colx, coly, colz = st.columns(3)
                    with colx:
                        da_x_label = st.selectbox("X-axis", all_da_labels, index=0, key="plsda_x_axis")
                    with coly:
                        da_y_label = st.selectbox("Y-axis (2D/3D)", all_da_labels, index=min(1, n_comp_da - 1),
                                                  key="plsda_y_axis")
                    with colz:
                        da_z_label = st.selectbox("Z-axis (3D)", all_da_labels, index=min(2, n_comp_da - 1),
                                                  key="plsda_z_axis")

                    da_x_idx = int(da_x_label.replace("PLSDA", "")) - 1
                    da_y_idx = int(da_y_label.replace("PLSDA", "")) - 1
                    da_z_idx = int(da_z_label.replace("PLSDA", "")) - 1

                    # Build projection DF (include all comps so selections work)
                    df_plsda = pd.DataFrame({"Label": y.values})
                    for i in range(n_comp_da):
                        df_plsda[f"PLSDA{i + 1}"] = T_da[:, i]

                    # # ===== 2D Projection =====

                    st.subheader("📉 PLS-DA 2D Projection")
                    xcol = f"PLSDA{da_x_idx + 1}"
                    ycol = f"PLSDA{da_y_idx + 1}"

                    fig_plsda_2d = px.scatter(
                        df_plsda,
                        x=xcol,
                        y=ycol,
                        color="Label",
                        title=f"PLS-DA Projection ({da_x_label} vs {da_y_label})",
                        labels={xcol: da_x_label, ycol: da_y_label}
                    )

                    # ---- Add class ellipses (same approach as PCA/PLS) ----
                    chi2_map = {0.90: 4.605, 0.95: 5.991, 0.99: 9.210}
                    chi2_val = chi2_map.get(conf_limit_plsda, 5.991)  # default 95%


                    def _ellipse_xy(points_2d, chi2_q, npts=200):
                        pts = np.asarray(points_2d)
                        if pts.shape[0] < 3:
                            return None, None
                        mu = pts.mean(axis=0)
                        S = np.cov(pts.T)
                        vals, vecs = np.linalg.eigh(S)
                        order = np.argsort(vals)[::-1]
                        vals, vecs = vals[order], vecs[:, order]
                        radii = np.sqrt(np.maximum(vals, 1e-12) * chi2_q)
                        t = np.linspace(0, 2 * np.pi, npts)
                        circ = np.stack([np.cos(t) * radii[0], np.sin(t) * radii[1]], axis=0)
                        ell = (vecs @ circ).T + mu
                        return ell[:, 0], ell[:, 1]


                    for lab in df_plsda["Label"].unique():
                        d = df_plsda[df_plsda["Label"] == lab][[xcol, ycol]].values
                        ex, ey = _ellipse_xy(d, chi2_val)
                        if ex is not None:
                            fig_plsda_2d.add_trace(
                                go.Scatter(
                                    x=ex, y=ey, mode="lines",
                                    name=f"{lab} {int(conf_limit_plsda * 100)}% ellipse",
                                    line=dict(width=1),
                                    showlegend=False
                                )
                            )

                    st.plotly_chart(fig_plsda_2d)

                    # ===== 3D Projection =====
                    if n_comp_da >= 3:
                        st.subheader("🧭 PLS-DA 3D Visualization")
                        fig_plsda_3d = px.scatter_3d(
                            df_plsda,
                            x=f"PLSDA{da_x_idx + 1}",
                            y=f"PLSDA{da_y_idx + 1}",
                            z=f"PLSDA{da_z_idx + 1}",
                            color="Label",
                            title=f"3D PLS-DA: {da_x_label} vs {da_y_label} vs {da_z_label}",
                            height=520, width=800
                        )
                        st.plotly_chart(fig_plsda_3d)
                    else:
                        st.info("Add ≥3 components to view the 3D projection.")

                    # ===== Loadings for Selected PLS-DA Component =====
                    st.subheader("📊 PLS-DA Loadings")
                    selected_da_loading = st.selectbox(
                        "Select a component to view its loadings:",
                        all_da_labels, index=0, key="plsda_loading_select"
                    )
                    da_load_idx = int(selected_da_loading.replace("PLSDA", "")) - 1

                    #####################################################################################################

                    fig_da_load = go.Figure()
                    fig_da_load.add_trace(go.Scatter(
                        x=selected_columns,
                        y=P_da[:, da_load_idx],
                        mode="lines",
                        name=f"{selected_da_loading} Loadings"
                    ))
                    fig_da_load.add_hline(y=0, line=dict(dash="dash"))

                    # Bond annotations (filtered by selected range)
                    band_annotations_da = {
                        1740: "C=O Stretch (Lipid)",
                        1650: "Amide I (Protein)",
                        1550: "Amide II (Protein)",
                        1338: "CH₂ (Collagen)",
                        1200: "C–O Stretch (Matrix)",
                        1115: "HPO₄²⁻ (Mineral)",
                        1060: "Sugar Ring C–O",
                        1030: "PO₄³⁻ (Mineral)",
                        875: "CO₃²⁻ Bending",
                        856: "C–S Proteoglycan",
                        7000: "O–H Stretch",
                        6688: "N–H Stretch",
                        5800: "CH₂ Lipid",
                        5200: "O–H Water",
                    }
                    y_top_da = float(np.nanmax(np.abs(P_da[:, da_load_idx])))
                    for wn, lbl in band_annotations_da.items():
                        if selected_range[0] <= wn <= selected_range[1]:
                            fig_da_load.add_vline(x=wn, line=dict(width=1, dash="dash"))
                            fig_da_load.add_annotation(x=wn, y=y_top_da, text=lbl, showarrow=True, yshift=10,
                                                       font=dict(size=9))

                    fig_da_load.update_layout(
                        title=f"Loadings for {selected_da_loading}",
                        xaxis_title="Wavenumber (cm⁻¹)",
                        yaxis_title="Loading",
                        height=400
                    )
                    st.plotly_chart(fig_da_load )

                    # ===== VIP (Variable Importance in Projection) with Bond Annotations =====

                    def compute_vip_multiresponse(pls_model):
                        T = pls_model.x_scores_  # (n_samples, A)
                        W = pls_model.x_weights_  # (p, A)
                        Q = pls_model.y_loadings_  # (m, A) for m classes
                        p, A = W.shape

                        ssq_y_comp = []
                        for a in range(A):
                            ss_t = np.sum(T[:, a] ** 2)
                            ss_q = np.sum(Q[:, a] ** 2)
                            ssq_y_comp.append(ss_t * ss_q)
                        ssq_y_comp = np.array(ssq_y_comp)
                        total_ssy = np.sum(ssq_y_comp)

                        vip = np.zeros(p)
                        denom_w = np.sum(W ** 2, axis=0)  # (A,)
                        for j in range(p):
                            weight = (W[j, :] ** 2) / denom_w  # (A,)
                            vip[j] = np.sqrt(p * np.sum(ssq_y_comp * weight) / total_ssy)
                        return vip


                    vip_da = compute_vip_multiresponse(pls_da)

                    vip_df_da = pd.DataFrame({
                        "Wavenumber (cm⁻¹)": selected_columns,
                        "VIP": vip_da,
                        "Abs Loading (selected)": np.abs(P_da[:, da_load_idx])
                    })

                    # ---- Bond annotation utilities ----
                    bond_annotations_table = [
                        (8500, "O–H Stretching and Bending (Water)"),
                        (7000, "O–H Stretching (Water)"),
                        (6688, "N–H Stretching (Protein/Collagen)"),
                        (5800, "CH₂ Stretching (Lipid)"),
                        (5200, "O–H Stretching and Bending (Water)"),
                        (4890, "N–H Bending (Protein/Collagen)"),
                        (4610, "C–H Stretching & Deformation (Protein/Collagen)"),
                        (4310, "Sugar Ring Vibrations (Proteoglycan)"),
                        ((3600, 3200), "O–H Stretching (Water/Hydroxyl)"),
                        ((3500, 3300), "N–H Stretching (Proteins)"),
                        ((3000, 2800), "C–H Stretching (Lipids, CH₂)"),
                        ((1750, 1650), "C=O Stretching (Proteins/Lipids)"),
                        (1550, "Amide II (Proteins)"),
                        (1338, "CH₂ Side Chain Bending (Collagen)"),
                        ((1100, 900), "PO₄³⁻ Stretching (Bone Mineral)"),
                        ((890, 850), "CO₃²⁻ Bending (Carbonate)"),
                        (1740, "C=O Stretching (Ester, Lipids)"),
                        (1650, "Amide I (Proteins)"),
                        (1630, "Water O–H Bending (Water)"),
                        ((1200, 1000), "C–O Stretching (Alcohols/Ethers)"),
                        (1115, "HPO₄²⁻ Stretching (Bone mineral)"),
                        (1060, "Sugar Ring C–O Stretch (Carbohydrates)"),
                        (1030, "PO₄³⁻ Stretching (Bone Mineral)"),
                        (875, "CO₃²⁻ Bending (Carbonates)"),
                        (856, "C–S Bending (Proteoglycans)"),
                        ((900, 800), "Aromatic C–H Bending (Fingerprint region)")
                    ]


                    def get_bond_annotation(wavenumber):
                        for entry in bond_annotations_table:
                            if isinstance(entry[0], tuple):
                                low, high = entry[0]
                                if high <= wavenumber <= low:
                                    return entry[1]
                        for entry in bond_annotations_table:
                            if not isinstance(entry[0], tuple) and entry[0] == int(wavenumber):
                                return entry[1]
                        all_wavenumbers = [val if isinstance(val, int) else (val[0] + val[1]) // 2 for val, _ in
                                           bond_annotations_table]
                        closest_idx = int(np.argmin([abs(wavenumber - wn) for wn in all_wavenumbers]))
                        return bond_annotations_table[closest_idx][1]


                    def extract_first_numeric(value):
                        value_str = str(value).strip()
                        if "/" in value_str:
                            value_str = value_str.split("/")[0].strip()
                        try:
                            return float(value_str)
                        except ValueError:
                            return None

                    vip_df_da["Numeric Wavenumber"] = vip_df_da["Wavenumber (cm⁻¹)"].apply(extract_first_numeric)
                    vip_df_da = vip_df_da.dropna(subset=["Numeric Wavenumber"])
                    vip_df_da["Band Annotation"] = vip_df_da["Numeric Wavenumber"].apply(get_bond_annotation)

                    vip_top_da = vip_df_da.sort_values("VIP", ascending=False).head(20).reset_index(drop=True)

                    # ===== 📊 Variable Importance with Bond Annotations + Download (one line) PLSDA =====
                    # ===== Header + Download =====
                    col1, col2 = st.columns([4, 1])
                    with col1:
                        st.markdown("### ⭐ Variable Importance with Bond Annotations")
                    with col2:
                        st.download_button(
                            "📥 Download CSV",
                            data=vip_top_da.to_csv(index=False).encode("utf-8"),
                            file_name="plsda_vip_with_bands.csv",
                            mime="text/csv"
                        )
                    # ===== 📊 Display Table =====
                    st.dataframe(
                        vip_top_da.style.format({
                            "VIP": "{:.3f}",
                            "Abs Loading (selected)": "{:.3f}"
                        })
                    )

                    ############### PLSDA Most Influential wavenumber
                    # ========================= ⭐ Most influential wavenumbers (by |loading|) — PLS-DA =========================

                    # --- tiny helpers (use your existing ones if already defined) ---
                    def _extract_first_numeric(v):
                        s = str(v).strip()
                        if "/" in s:
                            s = s.split("/")[0].strip()
                        try:
                            return float(s)
                        except Exception:
                            return np.nan


                    # If you already defined `get_bond_annotation` earlier in PLS-DA, you can remove this block
                    bond_annotations_table_da = [
                        (8500, "O–H Stretching and Bending (Water)"),
                        (7000, "O–H Stretching (Water)"),
                        (6688, "N–H Stretching (Protein/Collagen)"),
                        (5800, "CH₂ Stretching (Lipid)"),
                        (5200, "O–H Stretching and Bending (Water)"),
                        (4890, "N–H Bending (Protein/Collagen)"),
                        (4610, "C–H Stretching & Deformation (Protein/Collagen)"),
                        (4310, "Sugar Ring Vibrations (Proteoglycan)"),
                        ((3600, 3200), "O–H Stretching (Water/Hydroxyl)"),
                        ((3500, 3300), "N–H Stretching (Proteins)"),
                        ((3000, 2800), "C–H Stretching (Lipids, CH₂)"),
                        ((1750, 1650), "C=O Stretching (Proteins/Lipids)"),
                        (1550, "Amide II (Proteins)"),
                        (1338, "CH₂ Side Chain Bending (Collagen)"),
                        ((1100, 900), "PO₄³⁻ Stretching (Bone Mineral)"),
                        ((890, 850), "CO₃²⁻ Bending (Carbonate)"),
                        (1740, "C=O Stretching (Ester, Lipids)"),
                        (1650, "Amide I (Proteins)"),
                        (1630, "Water O–H Bending (Water)"),
                        ((1200, 1000), "C–O Stretching (Alcohols/Ethers)"),
                        (1115, "HPO₄²⁻ Stretching (Bone mineral)"),
                        (1060, "Sugar Ring C–O Stretch (Carbohydrates)"),
                        (1030, "PO₄³⁻ Stretching (Bone Mineral)"),
                        (875, "CO₃²⁻ Bending (Carbonates)"),
                        (856, "C–S Bending (Proteoglycans)"),
                        ((900, 800), "Aromatic C–H Bending (Fingerprint region)")
                    ]


                    def get_bond_annotation_da(wavenumber: float) -> str:
                        # ranges first
                        for entry in bond_annotations_table_da:
                            if isinstance(entry[0], tuple):
                                lo, hi = entry[0]
                                if hi <= wavenumber <= lo:
                                    return entry[1]
                        # exact
                        for entry in bond_annotations_table_da:
                            if not isinstance(entry[0], tuple) and entry[0] == int(wavenumber):
                                return entry[1]
                        # nearest
                        centers = [v if isinstance(v, int) else (v[0] + v[1]) / 2 for v, _ in bond_annotations_table_da]
                        idx = int(np.argmin([abs(wavenumber - c) for c in centers]))
                        return bond_annotations_table_da[idx][1]


                    # --- compute class context on the selected LV scores (T_da) ---
                    try:
                        df_scores_da = pd.DataFrame({
                            "Label": np.asarray(y),
                            "LV_Score": T_da[:, da_load_idx]
                        })
                        class_means_da = df_scores_da.groupby("Label", observed=True)["LV_Score"].mean().sort_values(
                            ascending=False)
                        class_pos_da = class_means_da.index[0] if len(class_means_da) else None
                        class_neg_da = class_means_da.index[-1] if len(class_means_da) else None
                    except Exception:
                        class_pos_da = class_neg_da = None


                    def _favor_class_da(value):
                        if class_pos_da is None or class_neg_da is None:
                            return ""
                        return f"favors **{class_pos_da}**" if value > 0 else f"favors **{class_neg_da}**"


                    # --- build drivers_df for the selected PLS-DA loading vector ---
                    load_vec_da = P_da[:, da_load_idx]  # shape: (n_features,)
                    drivers_rows_da = []
                    for label, val in zip(selected_columns, load_vec_da):
                        num = _extract_first_numeric(label)
                        band = get_bond_annotation_da(num) if np.isfinite(num) else ""
                        interp = _favor_class_da(val)
                        drivers_rows_da.append({
                            "Wavenumber (cm⁻¹)": str(label),
                            "Numeric (cm⁻¹)": float(num) if np.isfinite(num) else np.nan,
                            "Loading": float(val) if np.isfinite(val) else np.nan,
                            "|Loading|": float(abs(val)) if np.isfinite(val) else np.nan,
                            "Band Annotation": band,
                            "Interpretation": interp
                        })

                    drivers_df_da = (pd.DataFrame(drivers_rows_da)
                                     .sort_values("|Loading|", ascending=False)
                                     .reset_index(drop=True))

                    # For downstream code that expects `drivers_df`, mirror the variable name (safe alias).
                    drivers_df = drivers_df_da.copy()

                    # ===== Header + Download =====
                    col1, col2 = st.columns([4, 1])
                    with col1:
                        st.markdown("### ⭐ Most influential wavenumbers (by |loading|)")
                    with col2:
                        st.download_button(
                            "📥 Download CSV",
                            data=drivers_df_da[
                                ["Wavenumber (cm⁻¹)", "Loading", "|Loading|", "Band Annotation", "Interpretation"]]
                            .to_csv(index=False).encode("utf-8"),
                            file_name=f"most_influential_wavenumbers_{selected_da_loading}.csv",
                            mime="text/csv"

                        )

                    # ===== Table =====
                    st.dataframe(
                        drivers_df_da[
                            ["Wavenumber (cm⁻¹)", "Loading", "|Loading|", "Band Annotation", "Interpretation"]]
                        .style.format({"Loading": "{:.4f}", "|Loading|": "{:.4f}"})

                    )

                    ############## new summary table to run all model on selected pls-da###############

                    ############################# [  Narrative summary (PLS-DA style, parallel to PCA/PLS) ] #######################################################

                    ### ---------- Narrative summary (PLS-DA style, parallel to PCA/PLS) ----------
                    try:
                        var_pct_da = float(explained_x[da_load_idx]) * 100.0
                    except Exception:
                        var_pct_da = np.nan

                    # Positive vs negative contribution (aggregated in current focus)
                    pos_mass_da = float(np.nansum(np.abs(s_focus[s_focus > 0]))) if "s_focus" in globals() else np.nan
                    neg_mass_da = float(np.nansum(np.abs(s_focus[s_focus < 0]))) if "s_focus" in globals() else np.nan

                    contrast_da = (
                        "Positive-weighted regions dominate"
                        if pos_mass_da > neg_mass_da else
                        (
                            "Negative-weighted regions dominate" if neg_mass_da > pos_mass_da else "Balanced positive/negative weights")
                    )

                    summary_da = []

                    # summary_da.append(f"- **{selected_da_loading}** explains **{var_pct_da:.2f}%** of X-variance.")
                    # --- Compute explained X-variance ratio for PLS-DA manually ---
                    try:
                        X_residual = X_scaled_da.copy()
                        ss_total = np.sum(np.var(X_residual, axis=0, ddof=1))  # total variance in X
                        ss_comp = []

                        for a in range(pls_da.x_scores_.shape[1]):
                            t = pls_da.x_scores_[:, [a]]  # LV score
                            p = pls_da.x_loadings_[:, [a]].T  # LV loading
                            X_hat = t @ p  # variance explained by this LV
                            ss_comp.append(np.sum(np.var(X_hat, axis=0, ddof=1)))
                            X_residual = X_residual - X_hat  # subtract variance explained

                        explained_x_da = np.array(ss_comp) / ss_total
                    except Exception as e:
                        explained_x_da = np.full(pls_da.x_scores_.shape[1], np.nan)

                    # --- Use this in summary ---
                    try:
                        var_pct_da = float(explained_x_da[da_load_idx]) * 100.0
                    except Exception:
                        var_pct_da = np.nan

                    summary_da.append(f"- **{selected_da_loading}** explains **{var_pct_da:.2f}%** of X-variance.")

                    summary_da.append(f"- **Contrast:** {contrast_da} in the current range.")
                    if class_pos_da and class_neg_da:
                        summary_da.append(
                            f"- **Class context:** Higher LV scores (positive side) align with **{class_pos_da}**; "
                            f"lower scores (negative side) align with **{class_neg_da}**."
                        )
                    if "contrib_df" in globals() and len(contrib_df):
                        top_reg_da = contrib_df.iloc[0]
                        summary_da.append(
                            f"- **Strongest region:** {top_reg_da['Region (cm⁻¹)']} "
                            f"({top_reg_da['Net sign']}, |Σloading|={top_reg_da['|Σ loading|']:.4f}); "
                            f"likely bands: {top_reg_da['Likely bands'] or '—'}."
                        )

                    # Collect top driver bullets
                    head_n_da = min(6, len(drivers_df_da))
                    bullets_da = []
                    for _, r in drivers_df_da.head(head_n_da).iterrows():
                        side = "positive" if r["Loading"] > 0 else "negative"
                        interp = f" — {r['Interpretation']}" if r["Interpretation"] else ""
                        bullets_da.append(
                            f"  • {r['Wavenumber (cm⁻¹)']}: **{side}** loading ({r['Band Annotation']}){interp}"
                        )

                    # ===== Render summary in Streamlit =====
                    st.markdown("### 📝 Auto-summary (PLS-DA)")
                    st.markdown("\n".join(summary_da + ["- **Key drivers:**"] + bullets_da))

                    st.caption(
                        "Notes: (1) PLS-DA LV axis sign is arbitrary; interpretation uses relative +/− sides and class means. "
                        "(2) Region aggregation uses |loading| within ±half-width around top wavenumbers. "
                        "(3) Preprocessing (e.g., autoscale, derivatives) affects which bands dominate."
                    )

                    # ======================= TOGGLE: Run models on selected PLS-DA LV =======================
                    run_plsda_clf = st.checkbox("📊 Model Performance Summary for selected PLS-DA", value=True,key="plsda_run_selected_lv")

                    if run_plsda_clf:
                        # --- choose 1+ models (all selected by default) ---
                        model_choices_da = ["SVM", "Logistic Regression", "KNN", "Random Forest", "XGBoost"]
                        selected_models_da = st.multiselect(
                            "Select classification models (trained on the selected PLS-DA score)",
                            options=model_choices_da,
                            default=model_choices_da,
                            key="plsda_selected_lv_models"
                        )

                        # parse LV index, e.g. "PLSDA3" -> 2
                        sel_da_idx = int(str(selected_da_loading).replace("PLSDA", "")) - 1

                        # 1D feature = selected PLS-DA X-score
                        X_da = T_da[:, sel_da_idx].reshape(-1, 1)

                        # encode labels (keeps names for reports)
                        from sklearn.preprocessing import LabelEncoder

                        le_local_da = LabelEncoder()
                        y_codes_local_da = le_local_da.fit_transform(y)

                        # stratified split
                        from sklearn.model_selection import train_test_split

                        X_train_da, X_test_da, y_train_da, y_test_da = train_test_split(
                            X_da, y_codes_local_da, test_size=0.2, stratify=y_codes_local_da, random_state=42
                        )

                        # helper to build a model by name
                        def _make_model(name):
                            if name == "SVM":
                                return SVC()
                            if name == "Logistic Regression":
                                return LogisticRegression(max_iter=1000)
                            if name == "KNN":
                                return KNeighborsClassifier(n_neighbors=3)
                            if name == "Random Forest":
                                return RandomForestClassifier(n_estimators=100)
                            # XGBoost
                            return XGBClassifier(eval_metric="logloss")


                        # run all selected models
                        results_da = []
                        for mdl_name in selected_models_da:
                            clf_da = _make_model(mdl_name)
                            clf_da.fit(X_train_da, y_train_da)
                            y_pred_da = clf_da.predict(X_test_da)

                            # metrics (weighted for multi-class safety)
                            acc_da = clf_da.score(X_test_da, y_test_da)
                            prec_da = precision_score(y_test_da, y_pred_da, average="weighted", zero_division=0)
                            rec_da = recall_score(y_test_da, y_pred_da, average="weighted", zero_division=0)
                            f1_da = f1_score(y_test_da, y_pred_da, average="weighted", zero_division=0)

                            results_da.append({
                                "Model": mdl_name,
                                "LV used": selected_da_loading,
                                "Accuracy": acc_da,
                                "Precision (weighted)": prec_da,
                                "Recall (weighted)": rec_da,
                                "F1-Score (weighted)": f1_da
                            })

                        # ====== Colorful combined summary table + optional bar chart ======

                        results_da_df = pd.DataFrame(results_da)

                        # ===== 📊 Model Performance Summary (selected PC) + Download (one line) =====
                        col1, col2 = st.columns([4, 1])

                        with col1:
                            st.markdown("### 📊 Model Performance Summary (selected PLS-DA LV)")

                        with col2:
                            st.download_button(
                                "📥 Download CSV ", results_da_df.to_csv(index=False).encode("utf-8"),
                            file_name="plsda_selected_lv_model_metrics.csv",
                            mime="text/csv"
                        )

                        st.dataframe(
                            results_da_df.style
                            .background_gradient(cmap="YlGnBu")
                            .format({
                                "Accuracy": "{:.3f}",
                                "Precision (weighted)": "{:.3f}",
                                "Recall (weighted)": "{:.3f}",
                                "F1-Score (weighted)": "{:.3f}"
                            })

                        )

                        # detailed outputs per model
                        for mdl_name in selected_models_da:
                            st.markdown(f"### 🔎 Details: {mdl_name} on {selected_da_loading}")
                            col_rep, col_cm = st.columns([2, 3])

                            # rebuild + predict (keeps code simple; small 1D fit)
                            clf_da = _make_model(mdl_name)
                            clf_da.fit(X_train_da, y_train_da)
                            y_pred_da = clf_da.predict(X_test_da)

                            with col_rep:
                                st.markdown("**Classification Report**")
                                rep_dict = classification_report(
                                    y_test_da, y_pred_da,
                                    target_names=list(le_local_da.classes_),
                                    output_dict=True, zero_division=0
                                )
                                rep_df = pd.DataFrame(rep_dict).transpose()
                                rep_styled = rep_df.style.set_table_styles([
                                    {'selector': 'th', 'props': [('border', '1px solid black'), ('padding', '8px'),
                                                                 ('background', '#f2f2f2')]},
                                    {'selector': 'td', 'props': [('border', '1px solid black'), ('padding', '8px')]},
                                    {'selector': 'tr', 'props': [('border', '1px solid black')]}
                                ]).format("{:.2f}")
                                st.write(rep_styled)

                            with col_cm:
                                st.markdown("**Confusion Matrix (% by true class)**")
                                cm = confusion_matrix(y_test_da, y_pred_da, labels=np.arange(len(le_local_da.classes_)))
                                with np.errstate(divide="ignore", invalid="ignore"):
                                    cm_pct = cm.astype(float) / cm.sum(axis=1, keepdims=True) * 100
                                cm_pct = np.nan_to_num(cm_pct)
                                annot = np.array([[f"{v:.1f}%" for v in row] for row in cm_pct])

                                fig_cm, ax = plt.subplots()
                                sns.heatmap(
                                    cm_pct, annot=annot, fmt="", cmap="Blues", vmin=0, vmax=100,
                                    xticklabels=list(le_local_da.classes_),
                                    yticklabels=list(le_local_da.classes_),
                                    cbar_kws={"label": "% of true class"},
                                    ax=ax
                                )
                                ax.set_xlabel("Predicted")
                                ax.set_ylabel("True")
                                ax.set_title(f"{mdl_name} — Confusion Matrix (%) on {selected_da_loading}")
                                st.pyplot(fig_cm)
                                plt.close(fig_cm)
                    # ===================================================================================================



                    ############################## [  ] ############################################################
                    # ============================== 📥 Download PLSDA Full Workbook (editable loadings chart, 2D chart, 3D, summaries, all confusion matrices) ==============================

                    import io, numpy as np, pandas as pd, matplotlib.pyplot as plt, seaborn as sns
                    # ============================== One-file export for the entire "🔬 PLS: Partial Least Squares" ==============================
                    # ============================== One-file export for the entire "📈 PLS-DA: Partial Least Squares Discriminant Analysis" ==============================

                    import io, numpy as np, pandas as pd, matplotlib.pyplot as plt, seaborn as sns
                    from sklearn.preprocessing import LabelEncoder
                    from sklearn.model_selection import train_test_split
                    from sklearn.metrics import classification_report, confusion_matrix
                    from sklearn.svm import SVC
                    from sklearn.linear_model import LogisticRegression
                    from sklearn.neighbors import KNeighborsClassifier
                    from sklearn.ensemble import RandomForestClassifier

                    try:
                        from xgboost import XGBClassifier

                        _has_xgb = True
                    except Exception:
                        _has_xgb = False


                    # ---------- helper: auto-fit Excel columns ----------
                    def _autosize_columns(ws, df, extra_pad=2, max_width=60):
                        for i, col in enumerate(df.columns):
                            series = df[col].astype(str)
                            max_len = max([len(str(col))] + [len(s) for s in series.tolist()])
                            ws.set_column(i, i, min(max_len + extra_pad, max_width))


                    def _autosize_range(ws, first_col, last_col, rows, headers=None, extra_pad=2, max_width=60):
                        for c in range(first_col, last_col + 1):
                            values = []
                            if headers and c - first_col < len(headers):
                                values.append(str(headers[c - first_col]))
                            for r in rows:
                                v = r.get(c, "")
                                values.append("" if v is None else str(v))
                            width = min(max(len(x) for x in values) + extra_pad, max_width) if values else 10
                            ws.set_column(c, c, width)


                    # ---------- (A) Rebuild model results (AUTO-RUN; cumulative PLS-DA scores up to selected) ----------
                    # Parse k from "PLSDAk"
                    try:
                        _sel_da_idx = int(str(selected_da_loading).replace("PLSDA", "")) - 1
                    except Exception:
                        _sel_da_idx = 0

                    # Cumulative feature matrix: PLSDA1..selected
                    # (T_da shape: (n_samples, n_comp_da))
                    X_da_export = T_da[:, :_sel_da_idx + 1]
                    da_label_export = f"PLSDA1–PLSDA{_sel_da_idx + 1}"

                    # Encode labels
                    le_export_da = LabelEncoder()
                    y_codes_export_da = le_export_da.fit_transform(
                        df_plsda["Label"] if "Label" in df_plsda.columns else y)

                    # Split
                    X_train_e_da, X_test_e_da, y_train_e_da, y_test_e_da = train_test_split(
                        X_da_export, y_codes_export_da, test_size=0.2, stratify=y_codes_export_da, random_state=42
                    )

                    models_da = {
                        "SVM": SVC(),
                        "Logistic Regression": LogisticRegression(max_iter=1000),
                        "KNN": KNeighborsClassifier(n_neighbors=3),
                        "Random Forest": RandomForestClassifier(n_estimators=100),
                    }
                    if _has_xgb:
                        models_da["XGBoost"] = XGBClassifier(eval_metric="logloss")


                    results_rows_da = []
                    confmats_counts_da = {}
                    confmats_pct_da = {}
                    confmat_heatmap_png_da = {}

                    for name, clf in models_da.items():
                        clf.fit(X_train_e_da, y_train_e_da)
                        y_pred = clf.predict(X_test_e_da)

                        acc = clf.score(X_test_e_da, y_test_e_da)
                        rep = classification_report(
                            y_test_e_da, y_pred, target_names=list(le_export_da.classes_), output_dict=True,
                            zero_division=0
                        )
                        prec = rep["weighted avg"]["precision"]
                        rec = rep["weighted avg"]["recall"]
                        f1 = rep["weighted avg"]["f1-score"]

                        results_rows_da.append({
                            "Model": name,
                            "LVs used": da_label_export,
                            "Accuracy": acc,
                            "Precision (weighted)": prec,
                            "Recall (weighted)": rec,
                            "F1-Score (weighted)": f1
                        })

                        # Confusion matrices
                        cm = confusion_matrix(y_test_e_da, y_pred, labels=np.arange(len(le_export_da.classes_)))
                        confmats_counts_da[name] = pd.DataFrame(cm, index=list(le_export_da.classes_),
                                                                columns=list(le_export_da.classes_))

                        with np.errstate(divide="ignore", invalid="ignore"):
                            cm_pct = cm.astype(float) / cm.sum(axis=1, keepdims=True) * 100.0
                        cm_pct = np.nan_to_num(cm_pct)
                        confmats_pct_da[name] = pd.DataFrame(cm_pct, index=list(le_export_da.classes_),
                                                             columns=list(le_export_da.classes_))

                        # Heatmap PNG (in-memory)
                        fig, ax = plt.subplots(figsize=(5, 4))
                        sns.heatmap(cm_pct, annot=True, fmt=".1f", cmap="Blues",
                                    xticklabels=list(le_export_da.classes_), yticklabels=list(le_export_da.classes_),
                                    cbar=True, ax=ax, vmin=0, vmax=100)
                        ax.set_xlabel("Predicted");
                        ax.set_ylabel("True")
                        ax.set_title(f"{name} — Confusion Matrix (%) on {da_label_export}")
                        _png = io.BytesIO()
                        fig.tight_layout()
                        fig.savefig(_png, format="png", dpi=180)
                        plt.close(fig)
                        _png.seek(0)
                        confmat_heatmap_png_da[name] = _png

                    results_export_da = pd.DataFrame(results_rows_da)

                    # ---------- (B) Loadings PNG for 'Loadings Graph' ----------
                    # Compute %X-variance for selected PLS-DA LV (fallback if attribute missing)
                    try:
                        _da_num = int(str(selected_da_loading).replace("PLSDA", ""))
                    except Exception:
                        _da_num = 1

                    try:
                        _xexp = getattr(pls_da, "x_explained_variance_ratio_", None)
                        if _xexp is not None:
                            _var_pct_da = float(_xexp[_da_num - 1]) * 100.0
                        else:
                            # fallback: variance of scores relative to total X variance
                            X_num_all = X.values.astype(float) if hasattr(X, "values") else np.asarray(X, dtype=float)
                            X_tot = float(np.var(X_num_all, axis=0).sum()) or 1.0
                            _var_pct_da = float(np.var(T_da[:, _da_num - 1])) / X_tot * 100.0
                    except Exception:
                        _var_pct_da = np.nan

                    # Build a simple index-based loadings series for the selected LV
                    _da_load_idx = _da_num - 1
                    _series_vals_da = pd.to_numeric(pd.Series(P_da[:, _da_load_idx]), errors="coerce").values
                    _x_idx_da = np.arange(len(_series_vals_da))
                    figLd, axLd = plt.subplots(figsize=(11, 3.2))
                    axLd.plot(_x_idx_da, _series_vals_da, linewidth=2.8)
                    axLd.axhline(0.0, linestyle="--", color="gray", linewidth=1)
                    axLd.set_title(f"{selected_da_loading} Loadings ({_var_pct_da:.2f}% X-Variance)")
                    axLd.set_xlabel("Index (wavenumber order)")
                    axLd.set_ylabel("Loading")
                    axLd.grid(alpha=0.15)
                    _loadings_png_da = io.BytesIO()
                    figLd.tight_layout()
                    figLd.savefig(_loadings_png_da, format="png", dpi=180)
                    plt.close(figLd)
                    _loadings_png_da.seek(0)

                    # ---------- (C) Data for 2D/3D projections (use df_plsda) ----------
                    # Assume df_plsda has columns "PLSDA1", "PLSDA2", "PLSDA3", ..., and "Label"
                    proj2d_da_df = df_plsda[
                        [c for c in df_plsda.columns if c.startswith("PLSDA")][:2] + ["Label"]].copy()
                    # If 3D available:
                    _plsda_cols = [c for c in df_plsda.columns if c.startswith("PLSDA")]
                    if len(_plsda_cols) >= 3:
                        proj3d_da_df = df_plsda[_plsda_cols[:3] + ["Label"]].copy()
                    else:
                        proj3d_da_df = pd.DataFrame(columns=_plsda_cols[:2] + ["Z", "Label"])

                    # ---------- (D) Build Excel with all pages & editable charts ----------
                    buffer_plsda = io.BytesIO()
                    with pd.ExcelWriter(buffer_plsda, engine="xlsxwriter") as writer:
                        wb = writer.book

                        # --- PLS-DA Summary (if you built a summary table like plsda_calcv_df) ---
                        try:
                            plsda_calcv_df.to_excel(writer, sheet_name="PLS-DA Summary", index=False)
                            _autosize_columns(writer.sheets["PLS-DA Summary"], plsda_calcv_df)
                        except Exception:
                            pass

                        # --- VIP / Drivers / Regions (if available) ---
                        try:
                            vip_top_da.to_excel(writer, sheet_name="VIP (Top N)", index=False)
                            _autosize_columns(writer.sheets["VIP (Top N)"], vip_top_da)
                        except Exception:
                            pass

                        try:
                            _drivers_view_da = drivers_df[
                                ["Wavenumber (cm⁻¹)", "Loading", "|Loading|", "Band Annotation", "Interpretation"]]
                            _drivers_view_da.to_excel(writer, sheet_name="Most Influential", index=False)
                            _autosize_columns(writer.sheets["Most Influential"], _drivers_view_da)
                        except Exception:
                            pass

                        try:
                            contrib_df.to_excel(writer, sheet_name="Regions", index=False)
                            _autosize_columns(writer.sheets["Regions"], contrib_df)
                        except Exception:
                            pass

                        # --- Loadings (table + Excel-editable chart) ---
                        try:
                            def _num_first(val):
                                s = str(val).strip()
                                if "/" in s:
                                    s = s.split("/")[0].strip()
                                try:
                                    return float(s)
                                except Exception:
                                    return np.nan


                            load_tbl_da = pd.DataFrame({
                                "Wavenumber (cm⁻¹)": list(selected_columns),
                                "Numeric Wavenumber": [_num_first(v) for v in selected_columns],
                                "Index": np.arange(len(selected_columns)),
                                "Loading": pd.to_numeric(P_da[:, _da_load_idx], errors="coerce")
                            })

                            load_tbl_da.to_excel(writer, sheet_name="Loadings", index=False)
                            ws_load_da = writer.sheets["Loadings"]
                            _autosize_columns(ws_load_da, load_tbl_da)

                            chart_da = wb.add_chart({"type": "line"})
                            n_rows_da = len(load_tbl_da)
                            chart_da.add_series({
                                "name": f"{selected_da_loading}",
                                "categories": ["Loadings", 1, 1, n_rows_da, 1],  # Numeric Wavenumber
                                "values": ["Loadings", 1, 3, n_rows_da, 3],  # Loading
                                "line": {"width": 2.25},
                            })
                            chart_da.set_title({"name": f"Loadings — {selected_da_loading} ({_var_pct_da:.2f}% X-var)"})
                            chart_da.set_x_axis({"name": "Wavenumber (cm⁻¹)"})
                            chart_da.set_y_axis({"name": "Loading"})
                            chart_da.set_legend({"position": "bottom"})
                            ws_load_da.insert_chart(n_rows_da + 3, 0, chart_da)
                        except Exception:
                            pass

                        # --- 📝 Auto-summary ---
                        try:
                            auto_lines_da = []
                            auto_lines_da.extend(summary)  # list of bullet lines built earlier
                            auto_lines_da.append("Key drivers:")
                            auto_lines_da.extend(bullets)
                            auto_df_da = pd.DataFrame({"Summary": auto_lines_da})
                            auto_df_da.to_excel(writer, sheet_name="📝 Auto-summary", index=False)
                            _autosize_columns(writer.sheets["📝 Auto-summary"], auto_df_da)
                        except Exception:
                            pass

                        # --- 📉 PLS-DA 2D Projection (editable scatter) ---
                        try:
                            sheet2d_da = "📉 PLS-DA 2D Projection"
                            proj2d_da_df.to_excel(writer, sheet_name=sheet2d_da, index=False)
                            ws2d_da = writer.sheets[sheet2d_da]
                            _autosize_columns(ws2d_da, proj2d_da_df)

                            # guess first two PLSDA axes
                            da_x, da_y = [c for c in proj2d_da_df.columns if c.startswith("PLSDA")][:2]

                            chart2d_da = wb.add_chart({"type": "scatter"})
                            cur_row = len(proj2d_da_df) + 3
                            for lab in proj2d_da_df["Label"].unique():
                                sub = proj2d_da_df[proj2d_da_df["Label"] == lab][[da_x, da_y]].reset_index(drop=True)
                                if sub.empty:
                                    continue
                                ws2d_da.write(cur_row, 0, f"{lab} — {da_x}")
                                ws2d_da.write(cur_row, 1, f"{lab} — {da_y}")
                                ws2d_da.write_column(cur_row + 1, 0, sub[da_x].tolist())
                                ws2d_da.write_column(cur_row + 1, 1, sub[da_y].tolist())

                                s_start, s_end = cur_row + 1, cur_row + len(sub)
                                chart2d_da.add_series({
                                    "name": str(lab),
                                    "categories": [sheet2d_da, s_start, 0, s_end, 0],
                                    "values": [sheet2d_da, s_start, 1, s_end, 1],
                                    "marker": {"type": "circle", "size": 5},
                                    "line": {"none": True},
                                })
                                cur_row = s_end + 2

                            chart2d_da.set_title({"name": f"PLS-DA 2D: {da_x} vs {da_y}"})
                            chart2d_da.set_x_axis({"name": da_x})
                            chart2d_da.set_y_axis({"name": da_y})
                            chart2d_da.set_legend({"position": "bottom"})
                            ws2d_da.insert_chart(1, 4, chart2d_da)
                        except Exception:
                            pass

                        # --- 🧭 PLS-DA 3D Visualization (PNG fallback for Excel) ---
                        try:
                            sheet3d_da = "🧭 PLS-DA 3D Visualization"
                            proj3d_da_df.to_excel(writer, sheet_name=sheet3d_da, index=False)
                            ws3d_da = writer.sheets[sheet3d_da]
                            _autosize_columns(ws3d_da, proj3d_da_df)

                            _plsda_cols = [c for c in proj3d_da_df.columns if c.startswith("PLSDA")]
                            if len(_plsda_cols) >= 3:
                                from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

                                fig3d_da = plt.figure(figsize=(6, 5))
                                ax3d_da = fig3d_da.add_subplot(111, projection="3d")
                                for lab in proj3d_da_df["Label"].unique():
                                    d = proj3d_da_df[proj3d_da_df["Label"] == lab]
                                    ax3d_da.scatter(d[_plsda_cols[0]], d[_plsda_cols[1]], d[_plsda_cols[2]],
                                                    label=str(lab), s=16)
                                ax3d_da.set_xlabel(_plsda_cols[0]);
                                ax3d_da.set_ylabel(_plsda_cols[1]);
                                ax3d_da.set_zlabel(_plsda_cols[2])
                                ax3d_da.set_title("PLS-DA 3D Visualization")
                                ax3d_da.legend(loc="best", fontsize=8)
                                _plsda3d_png = io.BytesIO()
                                fig3d_da.tight_layout()
                                fig3d_da.savefig(_plsda3d_png, format="png", dpi=180)
                                plt.close(fig3d_da)
                                _plsda3d_png.seek(0)
                                ws3d_da.insert_image(1, 5, "plsda3d.png", {"image_data": _plsda3d_png})
                        except Exception:
                            pass

                        # --- Model_Metrics (summary + all confusion matrices + heatmaps) ---
                        wsM_da = wb.add_worksheet("Model_Metrics")
                        headers_da = list(results_export_da.columns)
                        for c, h in enumerate(headers_da):
                            wsM_da.write(0, c, h)
                        for r, row in results_export_da.iterrows():
                            for c, h in enumerate(headers_da):
                                wsM_da.write(r + 1, c, row[h])
                        _autosize_columns(wsM_da, results_export_da)

                        start_row = len(results_export_da) + 3
                        for name in models_da.keys():
                            # counts
                            wsM_da.write(start_row, 0, f"{name} — Confusion Matrix (counts)")
                            cnt = confmats_counts_da[name]
                            for j, col in enumerate([""] + cnt.columns.tolist()):
                                wsM_da.write(start_row + 1, j, col)
                            for i, idx in enumerate(cnt.index.tolist()):
                                wsM_da.write(start_row + 2 + i, 0, idx)
                                for j, col in enumerate(cnt.columns.tolist()):
                                    wsM_da.write(start_row + 2 + i, 1 + j, int(cnt.loc[idx, col]))
                            _autosize_range(
                                wsM_da,
                                first_col=0,
                                last_col=len(cnt.columns),
                                rows=[{k: None for k in range(len(cnt.columns) + 1)} for _ in
                                      range(len(cnt.index) + 2)],
                                headers=[""] + cnt.columns.tolist()
                            )
                            start_row += len(cnt) + 3

                            # percentages + heatmap image
                            wsM_da.write(start_row, 0, f"{name} — Confusion Matrix (%)")
                            pct = confmats_pct_da[name]
                            for j, col in enumerate([""] + pct.columns.tolist()):
                                wsM_da.write(start_row + 1, j, col)
                            for i, idx in enumerate(pct.index.tolist()):
                                wsM_da.write(start_row + 2 + i, 0, idx)
                                for j, col in enumerate(pct.columns.tolist()):
                                    wsM_da.write_number(start_row + 2 + i, 1 + j, float(pct.loc[idx, col]))
                            img_col = pct.shape[1] + 3
                            wsM_da.insert_image(start_row + 1, img_col, f"{name}_cm.png",
                                                {"image_data": confmat_heatmap_png_da[name], "x_scale": 1.0,
                                                 "y_scale": 1.0})
                            start_row += len(pct) + 10

                    buffer_plsda.seek(0)

                    # ================= Download Complete PLS-DA Analysis =================
                    colA_da, colB_da = st.columns([4, 1])
                    with colA_da:
                        st.download_button(
                            "📥 Download PLS-DA Full Workbook (editable loadings chart, 2D chart, 3D, summaries, all confusion matrices)",
                            data=buffer_plsda.getvalue(),
                            file_name="plsda_full_export.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                        )
                    with colB_da:
                        st.empty()
                    # =====================================================================================================================================

            ###############################################################################################################
            #
            # # ============================================================
            # # 📌 Clustering plots: PCA and PLS-DA
            # # PCA uses BLUE shades; PLS-DA uses PURPLE shades.
            # # ============================================================
            #
            # st.subheader("📌 Clustering plots: PCA and PLS-DA")
            #
            # with st.expander("📌 Clustering plots: PCA and PLS-DA", expanded=True):
            #
            #     import numpy as np
            #     import pandas as pd
            #     import plotly.express as px
            #     import plotly.graph_objects as go
            #     from sklearn.preprocessing import StandardScaler, LabelEncoder
            #     from sklearn.decomposition import PCA
            #     from sklearn.cross_decomposition import PLSRegression
            #
            #     # ---------------- Controls ----------------
            #     view_mode = st.radio(
            #         "Select view",
            #         ["2D", "3D"],
            #         horizontal=True,
            #         key="cluster_view_mode_pca_plsda"
            #     )
            #
            #     if view_mode == "2D":
            #         pcx = st.selectbox("PCA X-axis", ["PC1", "PC2", "PC3"], index=0, key="cluster_pca_x_2d")
            #         pcy = st.selectbox("PCA Y-axis", ["PC1", "PC2", "PC3"], index=1, key="cluster_pca_y_2d")
            #         dax = st.selectbox("PLS-DA X-axis", ["LV1", "LV2", "LV3"], index=0, key="cluster_plsda_x_2d")
            #         day = st.selectbox("PLS-DA Y-axis", ["LV1", "LV2", "LV3"], index=1, key="cluster_plsda_y_2d")
            #         show_ellipses = st.checkbox("Show class ellipses (2σ)", value=True, key="cluster_show_ellipses")
            #     else:
            #         pc_axes = st.multiselect(
            #             "PCA axes (choose 3)",
            #             ["PC1", "PC2", "PC3"],
            #             default=["PC1", "PC2", "PC3"],
            #             key="cluster_pca_axes_3d"
            #         )
            #         da_axes = st.multiselect(
            #             "PLS-DA axes (choose 3)",
            #             ["LV1", "LV2", "LV3"],
            #             default=["LV1", "LV2", "LV3"],
            #             key="cluster_plsda_axes_3d"
            #         )
            #         if len(pc_axes) != 3:
            #             st.warning("Please select exactly 3 PCA axes for 3D (PC1, PC2, PC3).")
            #             pc_axes = ["PC1", "PC2", "PC3"]
            #         if len(da_axes) != 3:
            #             st.warning("Please select exactly 3 PLS-DA axes for 3D (LV1, LV2, LV3).")
            #             da_axes = ["LV1", "LV2", "LV3"]
            #
            #     # ---------------- Prep (scaled X + labels) ----------------
            #     scaler_cp = StandardScaler()
            #     Xz_cp = scaler_cp.fit_transform(X)
            #
            #     y_cp = y.values if hasattr(y, "values") else np.asarray(y)
            #     y_cp = pd.Series(y_cp).astype(str).values  # ensure string labels for Plotly legend
            #
            #     # ---------------- PCA scores ----------------
            #     pca_cp = PCA(n_components=3).fit(Xz_cp)
            #     T_pca_cp = pca_cp.transform(Xz_cp)[:, :3]
            #     df_pca_cp = pd.DataFrame(T_pca_cp, columns=["PC1", "PC2", "PC3"])
            #     df_pca_cp["Label"] = y_cp
            #
            #     # ---------------- PLS-DA scores (PLS on one-hot Y) ----------------
            #     le_cp = LabelEncoder()
            #     y_codes_cp = le_cp.fit_transform(y_cp)
            #     y_onehot_cp = pd.get_dummies(y_codes_cp)
            #
            #     plsda_cp = PLSRegression(n_components=3).fit(Xz_cp, y_onehot_cp.values)
            #     T_da_cp = plsda_cp.x_scores_[:, :3]
            #     df_da_cp = pd.DataFrame(T_da_cp, columns=["LV1", "LV2", "LV3"])
            #     df_da_cp["Label"] = y_cp
            #
            #     # ============================================================
            #     # 🎨 Color palette (PCA = Blue, PLS-DA = Purple)
            #     # ============================================================
            #     df_pca_cp["Label"] = df_pca_cp["Label"].astype(str)
            #     df_da_cp["Label"] = df_da_cp["Label"].astype(str)
            #
            #     # Blue shades for PCA
            #     pca_color_map = {
            #         "WT": "#0E7ACA", #"#2bb5f0", #"#1f77b4",  # deep blue
            #         "OIM": "#7ecbf7"  # "#0469cf"   # "#8ecae6"  # light blue
            #     }
            #
            #     # Purple shades for PLS-DA
            #     plsda_color_map = {
            #         "WT": "#6E46C7", #"#8d65eb" , #"#a692d6",# "#6a0dad",  # deep purple
            #         "OIM":  "#dad7e0" #"#c77dff"  # light purple
            #     }
            #
            #     # ---------------- Optional: ellipses (2D only) ----------------
            #     def _ellipse_xy_cp(xy, nsig=2.0, npts=200):
            #         xy = np.asarray(xy)
            #         if xy.shape[0] < 3:
            #             return None, None
            #         mu = xy.mean(axis=0)
            #         S = np.cov(xy.T)
            #         vals, vecs = np.linalg.eigh(S)
            #         order = np.argsort(vals)[::-1]
            #         vals, vecs = vals[order], vecs[:, order]
            #         radii = nsig * np.sqrt(np.maximum(vals, 1e-12))
            #         ang = np.linspace(0, 2 * np.pi, npts)
            #         circ = np.stack([np.cos(ang) * radii[0], np.sin(ang) * radii[1]], axis=0)
            #         ell = (vecs @ circ).T + mu
            #         return ell[:, 0], ell[:, 1]
            #
            #
            #     def _add_ellipses_to_fig_cp(fig, df, xcol, ycol, label_col="Label", nsig=2.0, color_map=None):
            #         for lab in df[label_col].unique():
            #             pts = df[df[label_col] == lab][[xcol, ycol]].values
            #             ex, ey = _ellipse_xy_cp(pts, nsig=nsig)
            #             if ex is None:
            #                 continue
            #             line_color = None
            #             if isinstance(color_map, dict) and lab in color_map:
            #                 line_color = color_map[lab]
            #             fig.add_trace(
            #                 go.Scatter(
            #                     x=ex, y=ey,
            #                     mode="lines",
            #                     name=f"{lab} ellipse",
            #                     showlegend=False,
            #                     line=dict(width=2, color=line_color) if line_color else dict(width=2)
            #                 )
            #             )
            #
            #
            #     # ---------------- Plotting ----------------
            #     c1, c2 = st.columns(2)
            #
            #     with c1:
            #         st.markdown("### 🔬 PCA Clustering")
            #         if view_mode == "2D":
            #             fig_pca = px.scatter(
            #                 df_pca_cp,
            #                 x=pcx, y=pcy,
            #                 color="Label",
            #                 color_discrete_map=pca_color_map,
            #                 title=f"PCA Scores ({pcx} vs {pcy})",
            #                 hover_data=["Label"]
            #             )
            #             if show_ellipses:
            #                 _add_ellipses_to_fig_cp(fig_pca, df_pca_cp, pcx, pcy, "Label", nsig=2.0,
            #                                         color_map=pca_color_map)
            #             st.plotly_chart(fig_pca, use_container_width=True)
            #         else:
            #             x3, y3, z3 = pc_axes
            #             fig_pca3 = px.scatter_3d(
            #                 df_pca_cp,
            #                 x=x3, y=y3, z=z3,
            #                 color="Label",
            #                 color_discrete_map=pca_color_map,
            #                 title=f"PCA Scores (3D: {x3}, {y3}, {z3})",
            #                 hover_data=["Label"]
            #             )
            #             st.plotly_chart(fig_pca3, use_container_width=True)
            #
            #         st.download_button(
            #             "📥 Download PCA scores (CSV)",
            #             data=df_pca_cp.to_csv(index=False).encode("utf-8"),
            #             file_name="pca_scores.csv",
            #             mime="text/csv",
            #             key="dl_pca_scores_cluster"
            #         )
            #
            #     with c2:
            #         st.markdown("### 📊 PLS-DA Clustering")
            #         if view_mode == "2D":
            #             fig_da = px.scatter(
            #                 df_da_cp,
            #                 x=dax, y=day,
            #                 color="Label",
            #                 color_discrete_map=plsda_color_map,
            #                 title=f"PLS-DA Scores ({dax} vs {day})",
            #                 hover_data=["Label"]
            #             )
            #             if show_ellipses:
            #                 _add_ellipses_to_fig_cp(fig_da, df_da_cp, dax, day, "Label", nsig=2.0,
            #                                         color_map=plsda_color_map)
            #             st.plotly_chart(fig_da, use_container_width=True)
            #         else:
            #             x3, y3, z3 = da_axes
            #             fig_da3 = px.scatter_3d(
            #                 df_da_cp,
            #                 x=x3, y=y3, z=z3,
            #                 color="Label",
            #                 color_discrete_map=plsda_color_map,
            #                 title=f"PLS-DA Scores (3D: {x3}, {y3}, {z3})",
            #                 hover_data=["Label"]
            #             )
            #             st.plotly_chart(fig_da3, use_container_width=True)
            #
            #         st.download_button(
            #             "📥 Download PLS-DA scores (CSV)",
            #             data=df_da_cp.to_csv(index=False).encode("utf-8"),
            #             file_name="plsda_scores.csv",
            #             mime="text/csv",
            #             key="dl_plsda_scores_cluster"
            #         )
            # ============================================================
            # 📌 Clustering plots: PCA and PLS-DA
            # PCA uses BLUE shades; PLS-DA uses PURPLE shades.
            # Ellipse rings: WT = RED, OIM = ORANGE (when checkbox is ON)
            # ============================================================
            #
            # st.subheader("📌 Clustering plots: PCA and PLS-DA")
            #
            # with st.expander("📌 Clustering plots: PCA and PLS-DA", expanded=True):
            #
            #     import numpy as np
            #     import pandas as pd
            #     import plotly.express as px
            #     import plotly.graph_objects as go
            #     from sklearn.preprocessing import StandardScaler, LabelEncoder
            #     from sklearn.decomposition import PCA
            #     from sklearn.cross_decomposition import PLSRegression
            #
            #     # ---------------- Controls ----------------
            #     view_mode = st.radio(
            #         "Select view",
            #         ["2D", "3D"],
            #         horizontal=True,
            #         key="cluster_view_mode_pca_plsda"
            #     )
            #
            #     if view_mode == "2D":
            #         pcx = st.selectbox("PCA X-axis", ["PC1", "PC2", "PC3"], index=0, key="cluster_pca_x_2d")
            #         pcy = st.selectbox("PCA Y-axis", ["PC1", "PC2", "PC3"], index=1, key="cluster_pca_y_2d")
            #         dax = st.selectbox("PLS-DA X-axis", ["LV1", "LV2", "LV3"], index=0, key="cluster_plsda_x_2d")
            #         day = st.selectbox("PLS-DA Y-axis", ["LV1", "LV2", "LV3"], index=1, key="cluster_plsda_y_2d")
            #         show_ellipses = st.checkbox("Show class ellipses (2σ)", value=True, key="cluster_show_ellipses")
            #     else:
            #         pc_axes = st.multiselect(
            #             "PCA axes (choose 3)",
            #             ["PC1", "PC2", "PC3"],
            #             default=["PC1", "PC2", "PC3"],
            #             key="cluster_pca_axes_3d"
            #         )
            #         da_axes = st.multiselect(
            #             "PLS-DA axes (choose 3)",
            #             ["LV1", "LV2", "LV3"],
            #             default=["LV1", "LV2", "LV3"],
            #             key="cluster_plsda_axes_3d"
            #         )
            #         if len(pc_axes) != 3:
            #             st.warning("Please select exactly 3 PCA axes for 3D (PC1, PC2, PC3).")
            #             pc_axes = ["PC1", "PC2", "PC3"]
            #         if len(da_axes) != 3:
            #             st.warning("Please select exactly 3 PLS-DA axes for 3D (LV1, LV2, LV3).")
            #             da_axes = ["LV1", "LV2", "LV3"]
            #
            #     # ---------------- Prep (scaled X + labels) ----------------
            #     scaler_cp = StandardScaler()
            #     Xz_cp = scaler_cp.fit_transform(X)
            #
            #     y_cp = y.values if hasattr(y, "values") else np.asarray(y)
            #     y_cp = pd.Series(y_cp).astype(str).values  # ensure string labels for Plotly legend
            #
            #     # ---------------- PCA scores ----------------
            #     pca_cp = PCA(n_components=3).fit(Xz_cp)
            #     T_pca_cp = pca_cp.transform(Xz_cp)[:, :3]
            #     df_pca_cp = pd.DataFrame(T_pca_cp, columns=["PC1", "PC2", "PC3"])
            #     df_pca_cp["Label"] = y_cp
            #
            #     # ---------------- PLS-DA scores (PLS on one-hot Y) ----------------
            #     le_cp = LabelEncoder()
            #     y_codes_cp = le_cp.fit_transform(y_cp)
            #     y_onehot_cp = pd.get_dummies(y_codes_cp)
            #
            #     plsda_cp = PLSRegression(n_components=3).fit(Xz_cp, y_onehot_cp.values)
            #     T_da_cp = plsda_cp.x_scores_[:, :3]
            #     df_da_cp = pd.DataFrame(T_da_cp, columns=["LV1", "LV2", "LV3"])
            #     df_da_cp["Label"] = y_cp
            #
            #     # ============================================================
            #     # 🎨 Color palette (PCA = Blue, PLS-DA = Purple)
            #     # ============================================================
            #     df_pca_cp["Label"] = df_pca_cp["Label"].astype(str)
            #     df_da_cp["Label"] = df_da_cp["Label"].astype(str)
            #
            #     # Blue shades for PCA
            #     pca_color_map = {
            #         "WT": "#0E7ACA",  # deep blue
            #         "OIM": "#7ecbf7"  # light blue
            #     }
            #
            #     # Purple shades for PLS-DA
            #     plsda_color_map = {
            #         "WT": "#6E46C7",  # deep purple
            #         "OIM": "#dad7e0"  # light purple
            #     }
            #
            #     # ✅ Ellipse ring colors (WT=red, OIM=orange)
            #     ellipse_color_map = {
            #         "WT": "#E63946",  # red
            #         "OIM": "#F4A261"  # orange
            #     }
            #
            #
            #     # ---------------- Optional: ellipses (2D only) ----------------
            #     def _ellipse_xy_cp(xy, nsig=2.0, npts=200):
            #         xy = np.asarray(xy)
            #         if xy.shape[0] < 3:
            #             return None, None
            #         mu = xy.mean(axis=0)
            #         S = np.cov(xy.T)
            #         vals, vecs = np.linalg.eigh(S)
            #         order = np.argsort(vals)[::-1]
            #         vals, vecs = vals[order], vecs[:, order]
            #         radii = nsig * np.sqrt(np.maximum(vals, 1e-12))
            #         ang = np.linspace(0, 2 * np.pi, npts)
            #         circ = np.stack([np.cos(ang) * radii[0], np.sin(ang) * radii[1]], axis=0)
            #         ell = (vecs @ circ).T + mu
            #         return ell[:, 0], ell[:, 1]
            #
            #
            #     # NOTE: Signature unchanged; we still pass a dict via color_map=
            #     # Now it will color the ellipse rings using that dict (red/orange),
            #     # while your points stay blue/purple via px.scatter color_discrete_map.
            #     def _add_ellipses_to_fig_cp(fig, df, xcol, ycol, label_col="Label", nsig=2.0, color_map=None):
            #         for lab in df[label_col].unique():
            #             pts = df[df[label_col] == lab][[xcol, ycol]].values
            #             ex, ey = _ellipse_xy_cp(pts, nsig=nsig)
            #             if ex is None:
            #                 continue
            #             line_color = None
            #             if isinstance(color_map, dict) and lab in color_map:
            #                 line_color = color_map[lab]
            #             fig.add_trace(
            #                 go.Scatter(
            #                     x=ex, y=ey,
            #                     mode="lines",
            #                     name=f"{lab} ellipse",
            #                     showlegend=False,
            #                     line=dict(width=2, color=line_color) if line_color else dict(width=2)
            #                 )
            #             )
            #
            #
            #     # ---------------- Plotting ----------------
            #     c1, c2 = st.columns(2)
            #
            #     with c1:
            #         st.markdown("### 🔬 PCA Clustering")
            #         if view_mode == "2D":
            #             fig_pca = px.scatter(
            #                 df_pca_cp,
            #                 x=pcx, y=pcy,
            #                 color="Label",
            #                 color_discrete_map=pca_color_map,
            #                 title=f"PCA Scores ({pcx} vs {pcy})",
            #                 hover_data=["Label"]
            #             )
            #             if show_ellipses:
            #                 # ✅ Ellipses red/orange
            #                 _add_ellipses_to_fig_cp(fig_pca, df_pca_cp, pcx, pcy, "Label", nsig=2.0,
            #                                         color_map=ellipse_color_map)
            #             st.plotly_chart(fig_pca, use_container_width=True)
            #         else:
            #             x3, y3, z3 = pc_axes
            #             fig_pca3 = px.scatter_3d(
            #                 df_pca_cp,
            #                 x=x3, y=y3, z=z3,
            #                 color="Label",
            #                 color_discrete_map=pca_color_map,
            #                 title=f"PCA Scores (3D: {x3}, {y3}, {z3})",
            #                 hover_data=["Label"]
            #             )
            #             st.plotly_chart(fig_pca3, use_container_width=True)
            #
            #         st.download_button(
            #             "📥 Download PCA scores (CSV)",
            #             data=df_pca_cp.to_csv(index=False).encode("utf-8"),
            #             file_name="pca_scores.csv",
            #             mime="text/csv",
            #             key="dl_pca_scores_cluster"
            #         )
            #
            #     with c2:
            #         st.markdown("### 📊 PLS-DA Clustering")
            #         if view_mode == "2D":
            #             fig_da = px.scatter(
            #                 df_da_cp,
            #                 x=dax, y=day,
            #                 color="Label",
            #                 color_discrete_map=plsda_color_map,
            #                 title=f"PLS-DA Scores ({dax} vs {day})",
            #                 hover_data=["Label"]
            #             )
            #             if show_ellipses:
            #                 # ✅ Ellipses red/orange
            #                 _add_ellipses_to_fig_cp(fig_da, df_da_cp, dax, day, "Label", nsig=2.0,
            #                                         color_map=ellipse_color_map)
            #             st.plotly_chart(fig_da, use_container_width=True)
            #         else:
            #             x3, y3, z3 = da_axes
            #             fig_da3 = px.scatter_3d(
            #                 df_da_cp,
            #                 x=x3, y=y3, z=z3,
            #                 color="Label",
            #                 color_discrete_map=plsda_color_map,
            #                 title=f"PLS-DA Scores (3D: {x3}, {y3}, {z3})",
            #                 hover_data=["Label"]
            #             )
            #             st.plotly_chart(fig_da3, use_container_width=True)
            #
            #         st.download_button(
            #             "📥 Download PLS-DA scores (CSV)",
            #             data=df_da_cp.to_csv(index=False).encode("utf-8"),
            #             file_name="plsda_scores.csv",
            #             mime="text/csv",
            #             key="dl_plsda_scores_cluster"
            #         )

            # ============================
            # 📊 PCA • PLSDA (Side-by-Side) with Class Ellipses
            # (Same style/structure as your original block, just removed PLS)
            # ============================
            st.subheader("📊 PCA • PLSDA")

            with st.expander("📊 PCA • PLSDA — side-by-side comparison", expanded=True):

                # keep the same gating style; just ignore PLS
                if show_plsda and show_pca:

                    from plotly.subplots import make_subplots
                    from sklearn.preprocessing import StandardScaler, LabelEncoder
                    from sklearn.decomposition import PCA
                    from sklearn.cross_decomposition import PLSRegression


                    # ---------- helpers ----------
                    def _class_ellipse_xy(xy, nsig=2.0, npts=200):
                        """Return x,y for an nsig covariance ellipse of 2D points."""
                        xy = np.asarray(xy)
                        if xy.shape[0] < 3:
                            return None, None
                        mu = xy.mean(axis=0)
                        S = np.cov(xy.T)
                        vals, vecs = np.linalg.eigh(S)
                        order = np.argsort(vals)[::-1]
                        vals, vecs = vals[order], vecs[:, order]
                        radii = nsig * np.sqrt(np.maximum(vals, 1e-12))
                        ang = np.linspace(0, 2 * np.pi, npts)
                        circ = np.stack([np.cos(ang) * radii[0], np.sin(ang) * radii[1]], axis=0)
                        ell = (vecs @ circ).T + mu
                        return ell[:, 0], ell[:, 1]


                    def _add_class_ellipses(fig, df, xcol, ycol, label_col, subplot_col, nsig=2.0):
                        for lab in df[label_col].unique():
                            d = df[df[label_col] == lab][[xcol, ycol]].values
                            ex, ey = _class_ellipse_xy(d, nsig=nsig)
                            if ex is None:
                                continue
                            fig.add_trace(
                                go.Scatter(
                                    x=ex, y=ey, mode="lines", name=f"{lab} ellipse",
                                    line=dict(width=1), showlegend=(subplot_col == 1)
                                ),
                                row=1, col=subplot_col
                            )


                    # ---------- prep ----------
                    scaler = StandardScaler()
                    Xz = scaler.fit_transform(X)

                    # ---------- PCA (2D) ----------
                    pca2 = PCA(n_components=2).fit(Xz)
                    T_pca = pca2.transform(Xz)
                    df_pca = pd.DataFrame(T_pca, columns=["PC1", "PC2"])
                    df_pca["Label"] = y.values

                    # ---------- PLS-DA (multiresponse PLS on one-hot Y) ----------
                    y_codes = LabelEncoder().fit_transform(y)
                    y_onehot = pd.get_dummies(y_codes)  # (n_samples, n_classes)
                    plsda2 = PLSRegression(n_components=2).fit(Xz, y_onehot.values)
                    T_da = plsda2.x_scores_[:, :2]
                    df_da = pd.DataFrame(T_da, columns=["PLSDA1", "PLSDA2"])
                    df_da["Label"] = y.values

                    # ---------- figure ----------
                    fig = make_subplots(
                        rows=1, cols=2,
                        subplot_titles=("PCA", "PLS-DA"),
                        horizontal_spacing=0.10
                    )

                    # PCA panel
                    for lab in df_pca["Label"].unique():
                        d = df_pca[df_pca["Label"] == lab]
                        fig.add_trace(
                            go.Scatter(
                                x=d["PC1"], y=d["PC2"], mode="markers",
                                name=str(lab), showlegend=True
                            ),
                            row=1, col=1
                        )
                    _add_class_ellipses(fig, df_pca, "PC1", "PC2", "Label", subplot_col=1)

                    # PLS-DA panel
                    for lab in df_da["Label"].unique():
                        d = df_da[df_da["Label"] == lab]
                        fig.add_trace(
                            go.Scatter(
                                x=d["PLSDA1"], y=d["PLSDA2"], mode="markers",
                                name=str(lab), showlegend=False
                            ),
                            row=1, col=2
                        )
                    _add_class_ellipses(fig, df_da, "PLSDA1", "PLSDA2", "Label", subplot_col=2)

                    # axes + layout
                    fig.update_xaxes(title_text="PC1", row=1, col=1)
                    fig.update_yaxes(title_text="PC2", row=1, col=1)
                    fig.update_xaxes(title_text="PLSDA1", row=1, col=2)
                    fig.update_yaxes(title_text="PLSDA2", row=1, col=2)

                    fig.update_layout(
                        height=420,
                        margin=dict(l=30, r=30, t=60, b=10),
                        legend_title_text="Class"
                    )
                    st.plotly_chart(fig)

            # ============================
            # 📊 PCA • PLS • PLSDA (Side-by-Side) with Class Ellipses
            # ============================
            st.subheader("📊 PCA • PLS • PLSDA")

            with st.expander("📊 PCA • PLS • PLSDA — side-by-side comparison", expanded=True):

                if show_plsda and show_pls and show_pca:

                    from plotly.subplots import make_subplots
                    from sklearn.preprocessing import StandardScaler, LabelEncoder
                    from sklearn.decomposition import PCA
                    from sklearn.cross_decomposition import PLSRegression

                    # ---------- helpers ----------
                    def _class_ellipse_xy(xy, nsig=2.0, npts=200):
                        """Return x,y for an nsig covariance ellipse of 2D points."""
                        xy = np.asarray(xy)
                        if xy.shape[0] < 3:
                            return None, None
                        mu = xy.mean(axis=0)
                        S = np.cov(xy.T)
                        vals, vecs = np.linalg.eigh(S)
                        order = np.argsort(vals)[::-1]
                        vals, vecs = vals[order], vecs[:, order]
                        radii = nsig * np.sqrt(np.maximum(vals, 1e-12))
                        ang = np.linspace(0, 2 * np.pi, npts)
                        circ = np.stack([np.cos(ang) * radii[0], np.sin(ang) * radii[1]], axis=0)
                        ell = (vecs @ circ).T + mu
                        return ell[:, 0], ell[:, 1]

                    def _add_class_ellipses(fig, df, xcol, ycol, label_col, subplot_col, nsig=2.0):
                        for lab in df[label_col].unique():
                            d = df[df[label_col] == lab][[xcol, ycol]].values
                            ex, ey = _class_ellipse_xy(d, nsig=nsig)
                            if ex is None:
                                continue
                            fig.add_trace(
                                go.Scatter(x=ex, y=ey, mode="lines", name=f"{lab} ellipse",
                                           line=dict(width=1), showlegend=(subplot_col == 1)),
                                row=1, col=subplot_col
                            )

                    # ---------- prep ----------
                    scaler = StandardScaler()
                    Xz = scaler.fit_transform(X)

                    # ---------- PCA (2D) ----------
                    pca2 = PCA(n_components=2).fit(Xz)
                    T_pca = pca2.transform(Xz)
                    df_pca = pd.DataFrame(T_pca, columns=["PC1", "PC2"])
                    df_pca["Label"] = y.values

                    # ---------- PLS (2D, regression to encoded labels) ----------
                    y_enc = LabelEncoder().fit_transform(y)
                    pls2 = PLSRegression(n_components=2).fit(Xz, y_enc)
                    T_pls = pls2.x_scores_[:, :2]
                    df_pls = pd.DataFrame(T_pls, columns=["PLS1", "PLS2"])
                    df_pls["Label"] = y.values

                    # ---------- PLS-DA (multiresponse PLS on one-hot Y) ----------
                    y_codes = LabelEncoder().fit_transform(y)
                    y_onehot = pd.get_dummies(y_codes)  # (n_samples, n_classes)
                    plsda2 = PLSRegression(n_components=2).fit(Xz, y_onehot.values)
                    T_da = plsda2.x_scores_[:, :2]
                    df_da = pd.DataFrame(T_da, columns=["PLSDA1", "PLSDA2"])
                    df_da["Label"] = y.values

                    # ---------- figure ----------
                    fig = make_subplots(
                        rows=1, cols=3,
                        subplot_titles=("PCA", "PLS", "PLS-DA"),
                        horizontal_spacing=0.07
                    )

                    # PCA panel
                    for lab in df_pca["Label"].unique():
                        d = df_pca[df_pca["Label"] == lab]
                        fig.add_trace(go.Scatter(x=d["PC1"], y=d["PC2"], mode="markers",
                                                 name=str(lab), showlegend=True),
                                      row=1, col=1)
                    _add_class_ellipses(fig, df_pca, "PC1", "PC2", "Label", subplot_col=1)

                    # PLS panel
                    for lab in df_pls["Label"].unique():
                        d = df_pls[df_pls["Label"] == lab]
                        fig.add_trace(go.Scatter(x=d["PLS1"], y=d["PLS2"], mode="markers",
                                                 name=str(lab), showlegend=False),
                                      row=1, col=2)
                    _add_class_ellipses(fig, df_pls, "PLS1", "PLS2", "Label", subplot_col=2)

                    # PLS-DA panel
                    for lab in df_da["Label"].unique():
                        d = df_da[df_da["Label"] == lab]
                        fig.add_trace(go.Scatter(x=d["PLSDA1"], y=d["PLSDA2"], mode="markers",
                                                 name=str(lab), showlegend=False),
                                      row=1, col=3)
                    _add_class_ellipses(fig, df_da, "PLSDA1", "PLSDA2", "Label", subplot_col=3)

                    # axes + layout
                    fig.update_xaxes(title_text="PC1", row=1, col=1)
                    fig.update_yaxes(title_text="PC2", row=1, col=1)
                    fig.update_xaxes(title_text="PLS1", row=1, col=2)
                    fig.update_yaxes(title_text="PLS2", row=1, col=2)
                    fig.update_xaxes(title_text="PLSDA1", row=1, col=3)
                    fig.update_yaxes(title_text="PLSDA2", row=1, col=3)

                    fig.update_layout(
                        height=420,
                        margin=dict(l=30, r=30, t=60, b=10),
                        legend_title_text="Class"
                    )
                    st.plotly_chart(fig )

            #####################################################################################################################
            #####################################################################################################################

            ##############    📉 Standard Deviation of FTIR Spectra
            if show_stddev:

                st.subheader("📉 Standard Deviation of FTIR Spectra")
                with st.expander(
                        "Standard Deviation of FTIR Spectra reveals variability in absorbance at each wavenumber, helping identify spectral regions with significant differences across sample classes.",
                        expanded=True):
                    # Compute standard deviation for each class dynamically
                    std_df = X.groupby(y).std()

                    fig_std = go.Figure()

                    # Add a line for each class
                    for label in std_df.index:
                        fig_std.add_trace(go.Scatter(
                            x=selected_columns,
                            y=std_df.loc[label],
                            mode='lines',
                            name=f'{label} Std Dev'
                        ))

                    # Define annotation wavenumbers and corresponding labels
                    annotations = {
                        1740: "C=O Stretching (Ester)",
                        1650: "Amide I (C=O Stretch)",
                        1630: "Water O–H Bending",
                        1550: "Amide II (C–N stretch + N–H bend)",
                        1338: "CH₂ Side Chain Vibration",
                        1200: "C–O Stretching",
                        1115: "HPO₄²⁻ Stretching",
                        1060: "Sugar Ring C–O Stretch",
                        1030: "PO₄³⁻ Stretching",
                        875: "CO₃²⁻ Bending",
                        856: "C–S Bending",
                        850: "CO₃²⁻ Bending",
                        7000: "O–H Stretching",
                        6688: "N–H Stretching",
                        5800: "CH₂ Stretching",
                        5200: "O–H Stretching and Bending",
                        4890: "N–H Bending",
                        4610: "C–H Stretching & Deformation",
                        4310: "Sugar Ring Vibrations"
                    }

                    # Add vertical lines and annotations if in selected range
                    for wavenumber, label in annotations.items():
                        if selected_range[0] <= wavenumber <= selected_range[1]:
                            fig_std.add_vline(x=wavenumber, line=dict(color="gray", width=1, dash="dash"))
                            # Compute max Y value from the std_df for proper placement of annotation
                            max_y = std_df.max().max()
                            fig_std.add_annotation(x=wavenumber, y=max_y, text=label,
                                                   showarrow=True, arrowhead=1, yshift=10,
                                                   font=dict(size=9))

                    fig_std.update_layout(title="Standard Deviation of FTIR Spectra with Band Annotations",
                                          xaxis_title="Wavenumber (cm⁻¹)",
                                          yaxis_title="Standard Deviation",
                                          legend_title="Group",
                                          height=400,
                                          margin=dict(l=40, r=40, t=40, b=40))

                    st.plotly_chart(fig_std)

                    st.markdown(
                        "This plot compares the standard deviation of absorbance values across the FTIR spectrum between groups. Higher deviation may indicate greater molecular variability or structural heterogeneity in those regions.")

            #####################################################################################################################
            ##############    📈 Mean of FTIR Spectra

            if show_mean:

                st.subheader("📈 Mean of FTIR Spectra")
                with st.expander(
                        "Average FTIR Spectra with Band Annotations highlights the typical absorbance patterns for each class, with key vibrational bands labeled to aid in molecular interpretation.",
                        expanded=True):
                    # Compute mean for each class dynamically
                    mean_df = X.groupby(y).mean()

                    fig_mean = go.Figure()

                    # Add a line for each class
                    for label in mean_df.index:
                        fig_mean.add_trace(go.Scatter(
                            x=selected_columns,
                            y=mean_df.loc[label],
                            mode='lines',
                            name=f'{label} Mean'
                        ))

                    # Define annotation wavenumbers and corresponding labels
                    annotations = {
                        1740: "C=O Stretching (Ester)",
                        1650: "Amide I (C=O Stretch)",
                        1630: "Water O–H Bending",
                        1550: "Amide II (C–N stretch + N–H bend)",
                        1338: "CH₂ Side Chain Vibration",
                        1200: "C–O Stretching",
                        1115: "HPO₄²⁻ Stretching",
                        1060: "Sugar Ring C–O Stretch",
                        1030: "PO₄³⁻ Stretching",
                        875: "CO₃²⁻ Bending",
                        856: "C–S Bending",
                        850: "CO₃²⁻ Bending",
                        7000: "O–H Stretching",
                        6688: "N–H Stretching",
                        5800: "CH₂ Stretching",
                        5200: "O–H Stretching and Bending",
                        4890: "N–H Bending",
                        4610: "C–H Stretching & Deformation",
                        4310: "Sugar Ring Vibrations"
                    }

                    # Add vertical lines and annotations if in selected range
                    for wavenumber, label in annotations.items():
                        if selected_range[0] <= wavenumber <= selected_range[1]:
                            fig_mean.add_vline(x=wavenumber, line=dict(color="gray", width=1, dash="dash"))
                            # Compute max Y value from the mean_df for proper placement of annotation
                            max_y = mean_df.max().max()
                            fig_mean.add_annotation(x=wavenumber, y=max_y, text=label,
                                                    showarrow=True, arrowhead=1, yshift=10,
                                                    font=dict(size=9))

                    fig_mean.update_layout(title="Mean FTIR Spectra with Band Annotations",
                                           xaxis_title="Wavenumber (cm⁻¹)",
                                           yaxis_title="Mean Absorbance",
                                           legend_title="Group",
                                           height=400,
                                           margin=dict(l=40, r=40, t=40, b=40))

                    st.plotly_chart(fig_mean)

                    st.markdown(
                        "This plot displays the average absorbance spectrum for each class, allowing comparison of characteristic molecular features and structural composition across samples.")

            #####################################################################################################################
            ##############    📊 Mean+StdDev of FTIR Spectra

            if show_mean_stddev:

                st.subheader("📊 Mean + Standard Deviation of FTIR Spectra")
                with st.expander(
                        "Mean+Standard Deviation highlights areas of high absorbance and variability, indicating dynamic molecular activity across classes.",
                        expanded=True):
                    mean_df = X.groupby(y).mean()
                    std_df = X.groupby(y).std()

                    fig_mean_std = go.Figure()

                    for label in mean_df.index:
                        mean_plus_std = mean_df.loc[label] + std_df.loc[label]
                        fig_mean_std.add_trace(go.Scatter(
                            x=selected_columns,
                            y=mean_plus_std,
                            mode='lines',
                            name=f'{label} Mean + StdDev'
                        ))

                    annotations = {
                        1740: "C=O Stretching (Ester)",
                        1650: "Amide I (C=O Stretch)",
                        1630: "Water O–H Bending",
                        1550: "Amide II (C–N stretch + N–H bend)",
                        1338: "CH₂ Side Chain Vibration",
                        1200: "C–O Stretching",
                        1115: "HPO₄²⁻ Stretching",
                        1060: "Sugar Ring C–O Stretch",
                        1030: "PO₄³⁻ Stretching",
                        875: "CO₃²⁻ Bending",
                        856: "C–S Bending",
                        850: "CO₃²⁻ Bending",
                        7000: "O–H Stretching",
                        6688: "N–H Stretching",
                        5800: "CH₂ Stretching",
                        5200: "O–H Stretching and Bending",
                        4890: "N–H Bending",
                        4610: "C–H Stretching & Deformation",
                        4310: "Sugar Ring Vibrations"
                    }

                    for wavenumber, label_text in annotations.items():
                        if selected_range[0] <= wavenumber <= selected_range[1]:
                            fig_mean_std.add_vline(x=wavenumber, line=dict(color="gray", width=1, dash="dash"))
                            max_y = (mean_df + std_df).max().max()
                            fig_mean_std.add_annotation(x=wavenumber, y=max_y, text=label_text,
                                                        showarrow=True, arrowhead=1, yshift=10,
                                                        font=dict(size=9))

                    fig_mean_std.update_layout(title="Mean + StdDev FTIR Spectra with Band Annotations",
                                               xaxis_title="Wavenumber (cm⁻¹)",
                                               yaxis_title="Mean + StdDev Absorbance",
                                               legend_title="Group",
                                               height=400,
                                               margin=dict(l=40, r=40, t=40, b=40))

                    st.plotly_chart(fig_mean_std)

                    st.markdown(
                        "This plot displays the upper envelope (Mean + StdDev) of absorbance values across each FTIR spectrum class, highlighting regions with high activity and variability.")

            #####################################################################################################################
            ##############    📉 Minimum of FTIR Spectra

            if show_minimum:

                st.subheader("📉 Minimum of FTIR Spectra")
                with st.expander(
                        "Minimum absorbance reveals the lowest signal response per wavenumber, indicating consistent absence or weak presence of functional groups across samples.",
                        expanded=True):
                    min_df = X.groupby(y).min()

                    fig_min = go.Figure()

                    for label in min_df.index:
                        fig_min.add_trace(go.Scatter(
                            x=selected_columns,
                            y=min_df.loc[label],
                            mode='lines',
                            name=f'{label} Minimum'
                        ))

                    for wavenumber, label_text in annotations.items():
                        if selected_range[0] <= wavenumber <= selected_range[1]:
                            fig_min.add_vline(x=wavenumber, line=dict(color="gray", width=1, dash="dash"))
                            max_y = min_df.max().max()
                            fig_min.add_annotation(x=wavenumber, y=max_y, text=label_text,
                                                   showarrow=True, arrowhead=1, yshift=10,
                                                   font=dict(size=9))

                    fig_min.update_layout(title="Minimum FTIR Spectra with Band Annotations",
                                          xaxis_title="Wavenumber (cm⁻¹)",
                                          yaxis_title="Minimum Absorbance",
                                          legend_title="Group",
                                          height=400,
                                          margin=dict(l=40, r=40, t=40, b=40))

                    st.plotly_chart(fig_min)

                    st.markdown(
                        "This plot shows the lowest absorbance values per wavenumber for each group, useful for identifying regions with consistent low or absent molecular features.")

            #####################################################################################################################
            ##############    📈 Maximum of FTIR Spectra

            if show_maximum:

                st.subheader("📈 Maximum of FTIR Spectra")
                with st.expander(
                        "Maximum absorbance highlights spectral regions with the highest molecular signal response, often corresponding to dominant functional groups or sample peaks.",
                        expanded=True):
                    max_df = X.groupby(y).max()

                    fig_max = go.Figure()

                    for label in max_df.index:
                        fig_max.add_trace(go.Scatter(
                            x=selected_columns,
                            y=max_df.loc[label],
                            mode='lines',
                            name=f'{label} Maximum'
                        ))

                    for wavenumber, label_text in annotations.items():
                        if selected_range[0] <= wavenumber <= selected_range[1]:
                            fig_max.add_vline(x=wavenumber, line=dict(color="gray", width=1, dash="dash"))
                            max_y = max_df.max().max()
                            fig_max.add_annotation(x=wavenumber, y=max_y, text=label_text,
                                                   showarrow=True, arrowhead=1, yshift=10,
                                                   font=dict(size=9))

                    fig_max.update_layout(title="Maximum FTIR Spectra with Band Annotations",
                                          xaxis_title="Wavenumber (cm⁻¹)",
                                          yaxis_title="Maximum Absorbance",
                                          legend_title="Group",
                                          height=400,
                                          margin=dict(l=40, r=40, t=40, b=40))

                    st.plotly_chart(fig_max)

                    st.markdown(
                        "This plot shows the maximum absorbance values across the FTIR spectrum for each group, identifying dominant peaks and highly active regions.")

            #####################################################################################################################




            # ================================================
            # 📉 Molecular Interpretation from Spectral Bands
            # ================================================

            ##############    🧪 Molecular Interpretation from Spectral Bands

            st.subheader("🧪 Molecular Interpretation from Spectral Bands")
            with st.expander(
                    "Molecular Interpretation from Spectral Bands links specific wavenumber regions to functional group vibrations, enabling identification of molecular structures and chemical compositions in the samples.",
                    expanded=True):
                vib_data = [
                    ["8500", "O–H Stretching and Bending", "Water", "Hydration marker in tissues"],
                    ["7000", "O–H Stretching", "Water", "Water content level (broad absorption)"],
                    ["6688", "N–H Stretching", "Protein / Collagen", "Backbone amide stretching (protein presence)"],
                    ["5800", "CH₂ Stretching", "Lipid", "Lipid content (fatty acid chains)"],
                    ["5200", "O–H Stretching and Bending", "Water", "Hydration-sensitive feature"],
                    ["4890", "N–H Bending", "Protein / Collagen", "Protein conformation changes"],
                    ["4610", "C–H Stretching & Deformation", "Protein / Collagen", "Side chain deformations"],
                    ["4310", "Sugar Ring Vibrations", "Proteoglycan", "ECM components like GAGs"],

                    ["3600–3200", "O–H Stretching", "Water, Hydroxyl", "Broad peak, indicates hydration or hydroxylation"],
                    ["3500–3300", "N–H Stretching", "Proteins", "Backbone amide stretching in proteins"],
                    ["3000–2800", "C–H Stretching", "Lipids, CH₂", "Fatty acids in lipids (symmetric and asymmetric)"],
                    ["1750–1650", "C=O Stretching", "Proteins, lipids", "Amide I band; carbonyls in esters/lipids"],
                    ["1550", "N–H Bending / C–N Stretch", "Amide II (Proteins)", "Collagen structure marker"],
                    ["1338", "CH₂ Side Chain Bending", "Collagen", "Indicator of collagen integrity"],
                    ["1100–900", "PO₄³⁻ Stretching", "Bone Mineral", "Hydroxyapatite phosphate groups"],
                    ["890–850", "CO₃²⁻ Bending", "Carbonate", "Carbonate substitution in apatite"],

                    ["1740", "C=O Stretching (Ester)", "Lipids, fatty acid esters",
                     "Indicates lipid presence in bone tissue"],
                    ["1650", "Amide I (C=O Stretch)", "Proteins (peptide bonds)",
                     "Backbone stretch in proteins (collagen-rich regions)"],
                    ["1630", "Water O–H Bending", "Water content", "Tissue hydration; can affect FTIR baselines"],
                    ["1550", "Amide II (C–N stretch + N–H bend)", "Proteins (peptide bonds)",
                     "Protein secondary structure (collagen fibrils)"],
                    ["1338", "CH₂ Side Chain Vibration", "Collagen", "Structural collagen-related bending mode"],
                    ["1200–1000", "C–O Stretching", "Alcohols, esters, ethers",
                     "Carbohydrates and organic matrix components"],
                    ["1115", "HPO₄²⁻ Stretching", "Bone mineral", "Immature mineral phase (hydrogen phosphate)"],
                    ["1060", "Sugar Ring C–O Stretch", "Carbohydrates", "From glycoproteins/proteoglycans"],
                    ["1030", "PO₄³⁻ Stretching", "Bone mineral (hydroxyapatite)", "Mature mineral content in bone"],
                    ["875", "CO₃²⁻ Bending", "Carbonates", "Substituted carbonate in apatite; affects stiffness"],
                    ["856", "C–S Bending", "Sulfated proteoglycans",
                     "ECM components important in growth plate / cartilage"],
                    ["800–900", "Aromatic C–H Bending (out-of-plane)", "Aromatic rings (fingerprint region)",
                     "Often present in structural proteins and metabolites"]
                ]
                vib_df = pd.DataFrame(vib_data,
                                      columns=["Wavenumber (cm⁻¹)", "Assignment / Mode", "Functional Group / Bond",
                                               "Tissue/Biochemical Relevance"])
                st.dataframe(vib_df.style.set_properties(**{
                    'background-color': '#f8f9fa',
                    'color': 'black',
                    'border-color': 'lightgray'
                }))

        ###################################### [FTIR spectral data Information ] ######################################

            st.markdown("""
            The FTIR spectral data from bone typically contains vibrational features associated with both organic and inorganic components:
        
            #### 🔬 Organic Matrix:
            - **1650 cm⁻¹ (Amide I)**: C=O stretching of proteins (mainly collagen).
            - **1550 cm⁻¹ (Amide II)**: N–H bending and C–N stretching.
            - **1338 cm⁻¹**: CH₂ side chain vibrations in collagen.
            - **1200–1000 cm⁻¹**: C–O stretching and other sugar ring vibrations.
        
            #### 🧱 Mineral Content:
            - **1115 cm⁻¹**: HPO₄²⁻ bending (mineral phase).
            - **1030 cm⁻¹**: PO₄³⁻ symmetric stretching – primary phosphate marker.
            - **875 cm⁻¹**: CO₃²⁻ bending – carbonate substitution in apatite.
            - **856 cm⁻¹**: C–S or proteoglycan components.
        
            #### 💧 Water and Other Features:
            - **1630 cm⁻¹**: Water O–H bending.
            - **7000–3600 cm⁻¹**: O–H stretching and bending of bound/free water.
            - **5800 cm⁻¹**: CH₂ stretching (lipids or side chains).
            - **5200–4300 cm⁻¹**: Combination bands and sugar ring-related vibrations.
        
            These bands help identify changes in bone tissue due to conditions like osteoporosis, collagen mutation (OIM), or hydration level.
            """)

        #####################################################################################################################

############################################# [ footer ] #############################################################################

    st.markdown(
        """
        <div style='
            position: fixed; left: 0; bottom: 0; width: 100%;
            text-align: center; font-size: 0.9em; color: gray;
            background: #ffffffdd; padding: 8px 0;'>🧬 AI-Powered Vibrational Spectroscopy for Bone Health.   <br>
            • Copyright © 2025 Komal Wavhal - Stevens Institute of Technology • 
        </div>
        """,
        unsafe_allow_html=True
    )

##########################################################################################################################
